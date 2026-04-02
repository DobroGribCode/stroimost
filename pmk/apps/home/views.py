# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
import datetime
import io
import random
import mimetypes
import os
import cherrypy
import openpyxl
import calendar
import tempfile
import dateutil
from dateutil import parser

from openpyxl import Workbook

from docx import Document
from docx.shared import Pt
from docx.shared import Inches
from docxtpl import DocxTemplate

from io import StringIO, BytesIO
from cherrypy.lib import file_generator

from random import randint
from random import randrange

from datetime import date
from calendar import monthrange

from django import template
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import reverse
from django.template import loader
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.decorators import login_required
from openpyxl.worksheet import worksheet
from django.db.models import Count
from django.utils import timezone
from django.http import HttpResponse, Http404
from openpyxl.styles import PatternFill, Border
from ..reference.forms import GazForm, TabForm, ObjsForm, TaskForm, AutoForm, GroupForm,PersoneForm, DriverForm, AgreementForm, TempFileForm
from ..reference.models import gazolin,gaz_input,gaz_tabel_head,gaz_tabel_body,task_number,driver_tax_lagacy, notifications,notificated,persone_salary,driver_tax,temp_file,tabel_group, category, dayapprove,dayapproved,persone, period,tabel_list, objs,task_head, task_body, group, automobile, driver, driver_list, agreement, agreement_accepter,agreement_step,accepter_list,agreementer_list, ender_list
from django.contrib.auth.models import User

@login_required(login_url="/login/")
def get_notificated_data(request):
    notificated_data = notificated.objects.filter(user=request.user).first()
    if notificated_data:
        data = {
            'user': notificated_data.user.id,
            'notificated': notificated_data.notificated
        }
    else:
        data = {
            'user': None,
            'notificated': None
        }

    return JsonResponse(data)
@login_required(login_url="/login/")
def get_notifications(request):
    try:
        notificat = notifications.objects.all().values().order_by('-id')
        return JsonResponse(list(notificat), safe=False)
    except:
        return JsonResponse({'error': 'Не удалось найти уведомления'}, status=418)

@login_required(login_url="/login/")
def update_notificated(request):
    if request.method == 'POST':
        notificated_instance = notificated.objects.get(user=request.user)
        notificated_instance.notificated = 0
        notificated_instance.save()
        return JsonResponse({'message': 'Значение поля notificated успешно обновлено'})
    else:
        return JsonResponse({'message': 'Метод не разрешен'}, status=405)

class tab:
    @login_required(login_url="/login/")
    def tabels(request):
        context = {'segment': 'tabels'}
        html_template = loader.get_template('tabels/tabels.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def month_info(request):
        return redirect('/tabels/')


    @login_required(login_url="/login/")
    def driver(request):
        persones = persone.objects.all().filter(driver=1)
        per = driver.objects.all()
        groups = tabel_group.objects.all()
        if groups:
            pass
        else:
            groups
        context = {'per': per, 'segment': 'driver', 'groups': groups,'persones':persones}
        html_template = loader.get_template('tabels/driver/driver.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_driver(request):
        dri = driver.objects.all()
        form = DriverForm
        persones = persone.objects.all().filter(date_leave=None,driver=1).order_by('full_name')
        if request.method == 'POST':
            forms = DriverForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                year =request.POST.get("ye")
                month = request.POST.get("mo")
                da = year + '-' + month + '-' + '01'
                post.date = da
                da = date(int(year),int(month), 1)
                post.save()
                try:
                    da = da.replace(day=1)
                    post.dw1 = da
                except:
                    pass
                try:
                    da = da.replace(day=2)
                    post.dw2 = da
                except:
                    pass
                try:
                    da = da.replace(day=3)
                    post.dw3 = da
                except:
                    pass
                try:
                    da = da.replace(day=4)
                    post.dw4 = da
                except:
                    pass
                try:
                    da = da.replace(day=5)
                    post.dw5 = da
                except:
                    pass
                try:
                    da = da.replace(day=6)
                    post.dw6 = da
                except:
                    pass
                try:
                    da = da.replace(day=7)
                    post.dw7 = da
                except:
                    pass
                try:
                    da = da.replace(day=8)
                    post.dw8 = da
                except:
                    pass
                try:
                    da = da.replace(day=9)
                    post.dw9 = da
                except:
                    pass
                try:
                    da = da.replace(day=10)
                    post.dw10 = da
                except:
                    pass
                try:
                    da = da.replace(day=11)
                    post.dw11 = da
                except:
                    pass
                try:
                    da = da.replace(day=12)
                    post.dw12 = da
                except:
                    pass
                try:
                    da = da.replace(day=13)
                    post.dw13 = da
                except:
                     pass
                try:
                    da = da.replace(day=14)
                    post.dw14 = da
                except:
                     pass
                try:
                    da = da.replace(day=15)
                    post.dw15 = da
                except:
                    pass
                try:
                    da = da.replace(day=16)
                    post.dw16 = da
                except:
                    pass
                try:
                    da = da.replace(day=17)
                    post.dw17 = da
                except:
                    pass
                try:
                    da = da.replace(day=18)
                    post.dw18 = da
                except:
                    pass
                try:
                    da = da.replace(day=19)
                    post.dw19 = da
                except:
                    pass
                try:
                    da = da.replace(day=20)
                    post.dw20 = da
                except:
                    pass
                try:
                    da = da.replace(day=21)
                    post.dw21 = da
                except:
                    pass
                try:
                    da = da.replace(day=22)
                    post.dw22 = da
                except:
                    pass
                try:
                    da = da.replace(day=23)
                    post.dw23 = da
                except:
                    pass
                try:
                    da = da.replace(day=24)
                    post.dw24 = da
                except:
                    pass
                try:
                    da = da.replace(day=25)
                    post.dw25 = da
                except:
                    pass
                try:
                    da = da.replace(day=26)
                    post.dw26 = da
                except:
                    pass
                try:
                    da = da.replace(day=27)
                    post.dw27 = da
                except:
                    pass
                try:
                    da = da.replace(day=28)
                    post.dw28 = da
                except:
                    pass
                try:
                    da = da.replace(day=29)
                    post.dw29 = da
                except:
                    pass
                try:
                    da = da.replace(day=30)
                    post.dw30 = da
                except:
                    pass
                try:
                    da = da.replace(day=31)
                    post.dw31 = da
                except:
                    pass
                post.save()
                dris = driver.objects.all().first()
                return redirect('view_driver/'+str(dris.id))
        else:
            context = {
                'segment': 'driver',
                'dri': dri,
                'form': form,
            }
            html_template = loader.get_template('tabels/driver/new_driver.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def group_move_tabelD(request, ids, id):
        per_old = driver.objects.get(id=ids)
        tabs = driver_list.objects.all().filter(period=per_old)
        per = driver.objects.get(id=id)
        for t in tabs:
            p = persone.objects.get(id=t.persone.id)
            driver_list(
                period=per,
                persone=p,
                full_name=p.full_name,
                position=p.position,
                salary=p.salary,
                output=p.output,
            ).save()
        return redirect('/view_driver/' + str(id))

    @login_required(login_url="/login/")
    def AddPersoneD(request, id, ids):
        per = driver.objects.get(id=ids)
        p = persone.objects.get(id=id)
        driver_list(
            period=per,
            persone=p,
            full_name=p.full_name,
            position=p.position,
            salary=p.salary,
            output=p.output,
        ).save()
        return redirect('/view_driver/' + str(ids))

    @login_required(login_url="/login/")
    def view_driver(request, id):
        per = driver.objects.get(id=id)
        pers = driver.objects.all()
        driver_taxs = driver_tax.objects.all()
        persones = persone.objects.all().filter(date_leave=None,driver=1).order_by('full_name')
        all = persone.objects.all().filter(date_leave=None).order_by('full_name')
        tabs = driver_list.objects.all().filter(period=per).order_by('full_name')
        autos = automobile.objects.all().order_by('name')
        approvers = dayapproved.objects.all().filter(per=per)
        html_template = loader.get_template('tabels/driver/view_driver.html')
        today = datetime.date.today()
        next_month = today.replace(day=1) + datetime.timedelta(days=32)
        if per.date.month == today.month or next_month.month == per.date.month:
            for t in tabs:
                if t.company:
                    if t.company != t.persone.company:
                        t.company = t.persone.company
                if t.full_name:
                    if t.full_name != t.persone.full_name:
                        t.full_name = t.persone.full_name
                if t.position:
                    if t.position != t.persone.position:
                        t.position = t.persone.position
        if per.date:
            month = datetime.date.today()
            year = str(month)[0:4]
            month = str(month)[5:7]
            pm = per.date
            py = str(pm)[0:4]
            pm = str(pm)[5:7]
            if int(month) > int(pm) and int(year) == int(py):
                per.fullblock = 1
                per.save()
            if per.fullblock != 1:
                today = datetime.date.today().day - 2
                for i in range(1, 32):
                    db_field = f'db{i}'
                    if today > i:
                        setattr(per, db_field, 1)
                        per.save()
            if per:
                for day in range(1, 32):
                    i = 0
                    for a in approvers:
                        if a.day == day:
                            i += 1
                        if i >= 2:
                            setattr(per, f'db{day}', 1)
                            per.save()
        if request.method == 'POST':
            for t in tabs:
                for i in range(1, 32):
                    da_field = request.POST.get(f'{t.id}d{i}')
                    if da_field:
                        words = ["ВВ", "Н", "М", "Б", "О", "Д", "В", "У"]
                        try:
                            if da_field.isdigit():
                                da_value = int(da_field)
                                if 0 < da_value <= 23:
                                    setattr(t, f'd{i}', da_value)
                            elif da_field.upper() in words:
                                setattr(t, f'd{i}', da_field.upper())
                            t.save()
                        except:
                            pass
                if t:
                    for i in range(1, 32):
                        if getattr(t, f'd{i}'):
                            try:
                                d = int(getattr(t, f'd{i}'))
                                if d > 10:
                                    setattr(t, f'd{i}c', "#00b300")
                                else:
                                    setattr(t, f'd{i}c', None)
                            except:
                                color_mapping = {"ВВ": "#f2e93f", "Н": "#808080", "М": "#ffa812", "Б": "#ffff00",
                                                 "О": "#ffff00", "Д": "#ffff00", "У": "#ffff00", "В": "#fa5555"}
                                if getattr(t, f'd{i}') in color_mapping:
                                    setattr(t, f'd{i}c', color_mapping[getattr(t, f'd{i}')])
                            t.save()
                if t:
                    for i in range(1, 32):
                        if getattr(t, f'd{i}'):
                            t.auto_block = 1
                    t.save()
                h = 0
                d = 0
                out = 0
                for i in range(1, 32):
                    value = getattr(t, f'd{i}')
                    try:
                        value = int(value)
                        if 0 < value < 18:
                            h += int(value)
                            d += 1
                    except:
                        if value == "ВВ":
                            out += 1
                            if out < 3:
                                h += int(t.persone.work_hours)
                                d += 1
                t.days = int(d)
                t.hours = int(h)
                t.save()
                auto = request.POST.get(str(t.id) + 'auto')
                try:
                    t.auto = automobile.objects.get(id=int(auto))
                except:
                    pass
                comment = request.POST.get("com" + str(t.id))
                if comment == '':
                    pass
                else:
                    if t.comment == None:
                        t.comment = ''
                        t.save()
                        comment = str(t.comment) + str(comment)  + ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                    else:
                        comment = str(t.comment) + str(comment)  + ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()

                try:
                    fine = request.POST.get(str(t.id)+"fine")
                    if fine:
                        t.fine = float(fine)
                    else:
                        t.fine = 0
                except:
                    pass
                try:
                    add = request.POST.get(str(t.id)+"add")
                    if add:
                        t.add = float(add)
                    else:
                        t.add = 0
                except:
                    pass
                hour = 300
                f = t.fine
                a = t.add
                today = datetime.date.today()
                if today.month == per.date.month:
                    for ta in driver_taxs:
                        if ta.driver == t.persone:
                            try:
                                t.sum = int(ta.tax)*int(t.hours)
                            except:
                                t.sum = int(t.persone.salary)/hour * int(t.hours)
                else:
                    tal = driver_tax_lagacy.objects.all().filter(driver=t.persone,date=per.date).order_by('-id').first()
                    if tal is None:
                        t.sum = 0
                    else:
                        t.sum = int(tal.tax) * int(t.hours)
                try:
                    sum = int(t.sum)
                except:
                    sum = 0
                t.salary = int(sum) - float(f) + float(a)
                t.save()
            return redirect('/view_driver/' + str(id))
        else:
            context = {
                'autos': autos,
                'persones': persones,
                'per': per,
                'segment': 'driver',
                'tabs': tabs,
                'pers':pers,
                'all':all,
                'approvers': approvers,
            }
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_driver_days(request, id):
        per = driver.objects.get(id=id)
        pers = driver.objects.all()
        driver_taxs = driver_tax.objects.all()
        persones = persone.objects.all().filter(date_leave=None, driver=1).order_by('full_name')
        all = persone.objects.all().filter(date_leave=None).order_by('full_name')
        tabs = driver_list.objects.all().filter(period=per).order_by('full_name')
        autos = automobile.objects.all()
        approvers = dayapproved.objects.all().filter(per=per)
        html_template = loader.get_template('tabels/driver/view_driver.html')
        today = datetime.date.today()
        next_month = today.replace(day=1) + datetime.timedelta(days=32)
        if per.date.month == today.month or next_month.month == per.date.month:
            for t in tabs:
                if t.company:
                    if t.company != t.persone.company:
                        t.company = t.persone.company
                if t.full_name:
                    if t.full_name != t.persone.full_name:
                        t.full_name = t.persone.full_name
                if t.position:
                    if t.position != t.persone.position:
                        t.position = t.persone.position
        if per.date:
            month = datetime.date.today()
            year = str(month)[0:4]
            month = str(month)[5:7]
            pm = per.date
            py = str(pm)[0:4]
            pm = str(pm)[5:7]
            if int(month) > int(pm) and int(year) == int(py):
                per.fullblock = 1
                per.save()
            if per.fullblock != 1:
                today = datetime.date.today().day - 2
                for i in range(26, 32):
                    db_field = f'db{i}'
                    if today > i:
                        setattr(per, db_field, 1)
                        per.save()
            if per:
                for day in range(26, 32):
                    i = 0
                    for a in approvers:
                        if a.day == day:
                            i += 1
                        if i >= 2:
                            setattr(per, f'db{day}', 1)
                            per.save()
        if request.method == 'POST':
            for t in tabs:
                for i in range(26, 32):
                    da_field = request.POST.get(f'{t.id}d{i}')
                    if da_field:
                        words = ["ВВ", "Н", "М", "Б", "О", "Д", "В", "У"]
                        try:
                            if da_field.isdigit():
                                da_value = int(da_field)
                                if 0 < da_value <= 23:
                                    setattr(t, f'd{i}', da_value)
                            elif da_field.upper() in words:
                                setattr(t, f'd{i}', da_field.upper())
                            t.save()
                        except:
                            pass
                if t:
                    for i in range(26, 32):
                        if getattr(t, f'd{i}'):
                            try:
                                d = int(getattr(t, f'd{i}'))
                                if d > 10:
                                    setattr(t, f'd{i}c', "#00b300")
                                else:
                                    setattr(t, f'd{i}c', None)
                            except:
                                color_mapping = {"ВВ": "#f2e93f", "Н": "#808080", "М": "#ffa812", "Б": "#ffff00",
                                                 "О": "#ffff00", "Д": "#ffff00", "У": "#ffff00", "В": "#fa5555"}
                                if getattr(t, f'd{i}') in color_mapping:
                                    setattr(t, f'd{i}c', color_mapping[getattr(t, f'd{i}')])
                            t.save()
                if t:
                    for i in range(26, 32):
                        if getattr(t, f'd{i}'):
                            t.auto_block = 1
                    t.save()
                h = 0
                d = 0
                out = 0
                for i in range(26, 32):
                    value = getattr(t, f'd{i}')
                    try:
                        value = int(value)
                        if 0 < value < 18:
                            h += int(value)
                            d += 1
                    except:
                        if value == "ВВ":
                            out += 1
                            if out < 3:
                                h += int(t.persone.work_hours)
                                d += 1
                t.days = int(d)
                t.hours = int(h)
                t.save()
                auto = request.POST.get(str(t.id) + 'auto')
                try:
                    t.auto = automobile.objects.get(id=int(auto))
                except:
                    pass
                comment = request.POST.get("com" + str(t.id))
                if comment == '':
                    pass
                else:
                    if t.comment == None:
                        t.comment = ''
                        t.save()
                        comment = str(t.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(
                            request.user.first_name) + ' ' + str(
                            datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                    else:
                        comment = str(t.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(
                            request.user.first_name) + ' ' + str(
                            datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()

                try:
                    fine = request.POST.get(str(t.id) + "fine")
                    if fine:
                        t.fine = float(fine)
                    else:
                        t.fine = 0
                except:
                    pass
                try:
                    add = request.POST.get(str(t.id) + "add")
                    if add:
                        t.add = float(add)
                    else:
                        t.add = 0
                except:
                    pass
                hour = 300
                f = t.fine
                a = t.add
                today = datetime.date.today()
                if today.month == per.date.month:
                    for ta in driver_taxs:
                        if ta.driver == t.persone:
                            try:
                                t.sum = int(ta.tax) * int(t.hours)
                            except:
                                t.sum = int(t.persone.salary) / hour * int(t.hours)
                else:
                    tal = driver_tax_lagacy.objects.all().filter(driver=t.persone, date=per.date).order_by(
                        '-id').first()
                    if tal is None:
                        t.sum = 0
                    else:
                        t.sum = int(tal.tax) * int(t.hours)
                try:
                    sum = int(t.sum)
                except:
                    sum = 0
                t.salary = int(sum) - float(f) + float(a)
                t.save()
            return redirect('/view_driver/' + str(id))
        else:
            context = {
                'autos': autos,
                'persones': persones,
                'per': per,
                'segment': 'driver',
                'tabs': tabs,
                'pers': pers,
                'all': all,
                'approvers': approvers,
            }
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def del_driverl(request, id, ids):
        tl = driver_list.objects.get(id=id)
        tl.delete()
        return redirect('/view_driver/' + str(ids))

    @login_required(login_url="/login/")
    def tabel(request):
        per = period.objects.all()
        persones = persone.objects.all()
        groups = tabel_group.objects.all()
        if groups:
            pass
        else:
            groups
        context = {'per': per, 'segment': 'tabel', 'groups': groups, 'persones':persones,}
        html_template = loader.get_template('tabels/worker/tabel.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def search(request, id):
        per = period.objects.all()
        searcher = persone.objects.get(id=id)
        persones = persone.objects.all()
        groups = tabel_group.objects.all()
        tabs = tabel_list.objects.all().filter(persone=searcher)
        tabsd = driver_list.objects.all().filter(persone=searcher)
        context = {'tabsd':tabsd, 'per': per, 'segment': 'tabel', 'groups': groups, 'persones': persones,'searcher':searcher,'tabs':tabs, }
        html_template = loader.get_template('tabels/worker/search.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_tabel(request):
        per = period.objects.all()
        form = TabForm
        persones = persone.objects.all().filter(date_leave=None).order_by('full_name')
        tl = tabel_list.objects.all()
        if request.method == 'POST':
            forms = TabForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                year =request.POST.get("ye")
                month = request.POST.get("mo")
                da = year + '-' + month + '-' + '01'
                post.date = da
                try:
                    gr = tabel_group.objects.get(date=post.date)
                except:
                    tabel_group(date=post.date).save()
                    gr = tabel_group.objects.get(date=post.date)
                post.group = gr
                da = date(int(year),int(month), 1)
                if post.type != None:
                    if post.graphic == 1:
                        days = monthrange(int(year),int(month))[1]
                        post.workdays = int(days)
                    post.save()
                    try:
                        da = da.replace(day=1)
                        post.dw1 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=2)
                        post.dw2 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=3)
                        post.dw3 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=4)
                        post.dw4 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=5)
                        post.dw5 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=6)
                        post.dw6 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=7)
                        post.dw7 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=8)
                        post.dw8 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=9)
                        post.dw9 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=10)
                        post.dw10 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=11)
                        post.dw11 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=12)
                        post.dw12 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=13)
                        post.dw13 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=14)
                        post.dw14 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=15)
                        post.dw15 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=16)
                        post.dw16 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=17)
                        post.dw17 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=18)
                        post.dw18 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=19)
                        post.dw19 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=20)
                        post.dw20 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=21)
                        post.dw21 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=22)
                        post.dw22 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=23)
                        post.dw23 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=24)
                        post.dw24 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=25)
                        post.dw25 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=26)
                        post.dw26 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=27)
                        post.dw27 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=28)
                        post.dw28 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=29)
                        post.dw29 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=30)
                        post.dw30 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=31)
                        post.dw31 = da
                    except:
                        pass
                    post.save()
                    return redirect('tabel')
                else:
                    error = 'Месяц или Год не выбраны'
                    context = {
                        'error': error,
                        'segment': 'tabel',
                        'per': per,
                        'form': form,
                    }
                    html_template = loader.get_template('tabels/worker/new_tabel.html')
                    return HttpResponse(html_template.render(context, request))
        else:
            context = {
                'segment': 'tabel',
                'persones':persones,
                'per': per,
                'form': form,
            }
            html_template = loader.get_template('tabels/worker/new_tabel.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def del_tabl(request, id, ids):
        tl = tabel_list.objects.get(id=id)
        tl.delete()
        return redirect('/view_tabel/' + str(ids))

    @login_required(login_url="/login/")
    def move_worker(request, id, ids, gid):
        g = group.objects.get(id=gid)
        tl = tabel_list.objects.get(id=id)
        tl.group = g
        tl.save()
        return redirect('/view_tabel/' + str(ids))

    @login_required(login_url="/login/")
    def AddPersone(request,id,ids):
        per = period.objects.get(id=ids)
        p = persone.objects.get(id=id)
        try:
            g = group.objects.get(id=p.group.id)
            if g:
                tabel_list(
                    period=per,
                    persone=p,
                    company=p.company,
                    snils=p.snils,
                    full_name=p.full_name,
                    group=g,
                    position=p.position,
                    salary=p.salary,
                    output=p.output,
                ).save()
            else:
                g = group.objects.get(name='None')
                tabel_list(
                    period=per,
                    persone=p,
                    company=p.company,
                    snils=p.snils,
                    full_name=p.full_name,
                    group=g,
                    position=p.position,
                    salary=p.salary,
                    output=p.output,
                ).save()
        except:
            g = group.objects.get(name='None')
            tabel_list(
                period=per,
                persone=p,
                company=p.company,
                snils=p.snils,
                full_name=p.full_name,
                group=g,
                position=p.position,
                salary=p.salary,
                output=p.output,
            ).save()
        return redirect('/view_tabel/' + str(ids))

    def check_day_in_other_table(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            ids = request.POST.get('ids')
            day = request.POST.get('day')
            type = request.POST.get('type')
            value = request.POST.get('value')
            if type == 'p':
                p = tabel_list.objects.get(id=id)
                per = period.objects.get(id=p.period.id)
                try:
                    if value not in ['Н', 'У', 'В', 'н', 'у', 'в']:
                        checklist = tabel_list.objects.filter(persone=p.persone, period__group=per.group).exclude(id=p.id)
                        for check in checklist:
                            day_value = getattr(check, f'd{day}')
                            if day_value and day_value not in ['Н', 'У', 'В']:  # Проверка на заполнение и значение "Н" или "У"
                                ote = f'День заполнен в другом табеле: {check.period.obj}'
                                return JsonResponse({'exists': True, 'other': ote})
                    else:
                        checklist = tabel_list.objects.filter(persone=p.persone, period__group=per.group).exclude(id=p.id)
                        for check in checklist:
                            day_value = getattr(check, f'd{day}')
                            if day_value and day_value not in ['Н', 'У', 'В']:  # Проверка на заполнение и значение "Н" или "У"
                                return JsonResponse({'error': 'Заполнен в другом дне, но Н,У,В допустимы.'}, status=418)
                except tabel_list.DoesNotExist:
                    return JsonResponse({'exists': False})
            elif type == 'D':
                p = driver_list.objects.get(id=id)
                per = driver.objects.get(id=p.period.id)
                try:
                    if value not in ['Н', 'У', 'В', 'н', 'у', 'в']:
                        checklist = driver_list.objects.filter(persone=p.persone, period__date=per.date).exclude(id=p.id)
                        for check in checklist:
                            day_value = getattr(check, f'd{day}')
                            if day_value and day_value not in ['Н', 'У', 'В']:  # Проверка на заполнение и значение "Н" или "У"
                                ote = f'День заполнен в другом табеле: {check.period.obj}'
                                return JsonResponse({'exists': True, 'other': ote})
                    else:
                        checklist = driver_list.objects.filter(persone=p.persone, period__date=per.date).exclude(id=p.id)
                        for check in checklist:
                            day_value = getattr(check, f'd{day}')
                            if day_value and day_value not in ['Н', 'У', 'В']:  # Проверка на заполнение и значение "Н" или "У"
                                return JsonResponse({'error': 'Заполнен в другом дне, но Н,У,В допустимы.'}, status=418)
                except tabel_list.DoesNotExist:
                    return JsonResponse({'exists': False})

        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    def save_input(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            day = request.POST.get('day')
            value = request.POST.get('value')
            type = request.POST.get('type')
            day_mapping = {str(i): f'd{i}' for i in range(1, 32)}  # Создаем словарь для дней от 1 до 31
            day_picker_mapping = {str(i): f'daypicker{i}' for i in range(1, 32)}
            picker = User.objects.get(id=request.user.id)
            if type == 'p':
                p = tabel_list.objects.get(id=id)
                if day in day_mapping:
                    if value in ["Б", "В", "Д", "У", "М", "Н", "ВВ","б", "в", "д", "у", "м", "н", "вв", "О", "о"] or (value.isdigit() and 0 <= int(value) <= 18):
                        try:
                            int(value)
                            if int(value) == 0:
                                pass
                                setattr(p, day_mapping[day], None)
                                setattr(p, day_picker_mapping[day], picker)
                                p.save()
                                return JsonResponse({'message': 'Данные успешно очищены'})
                            else:
                                setattr(p, day_mapping[day], value)
                                setattr(p, day_picker_mapping[day], picker)
                                p.save()
                                return JsonResponse({'message': 'Данные успешно сохранены'})
                        except:
                            setattr(p, day_mapping[day], value)
                            setattr(p, day_picker_mapping[day], picker)
                            p.save()
                            return JsonResponse({'message': 'Данные успешно сохранены'})
                    elif value == '':
                        return JsonResponse({'error': 'Null Недопустимое значение для поля Value'}, status=400)
                    else:
                        ote = f'Недопустимый значение для:'+str(day)+ 'дня у' +str(p.full_name)
                        return JsonResponse({'error': 'Недопустимое значение для поля Value','other': ote}, status=400)
            elif type == 'd':
                p = driver_list.objects.get(id=id)
                if day in day_mapping:
                    if value in ["Б", "В", "Д", "У", "М", "Н", "ВВ","б", "в", "д", "у", "м", "н", "вв", "О", "о"] or (value.isdigit() and 0 <= int(value) <= 18):
                        setattr(p, day_mapping[day], value)
                        setattr(p, day_picker_mapping[day], picker)
                        p.save()
                        return JsonResponse({'message': 'Данные успешно сохранены'})
                    elif value == '':
                        return JsonResponse({'error': 'Null Недопустимое значение для поля Value'}, status=400)
                    else:
                        ote = f'Недопустимый значение для:'+str(day)+ ' дня у ' +str(p.full_name)
                        return JsonResponse({'error': 'Недопустимое значение для поля Value','other': ote}, status=400)
            elif type == 'g':
                p = gaz_tabel_body.objects.get(id=id)
                if day in day_mapping:
                    if value.isdigit():
                        setattr(p, day_mapping[day], value)
                        setattr(p, day_picker_mapping[day], picker)
                        p.save()
                        return JsonResponse({'message': 'Данные успешно сохранены'})
                    elif value == '':
                        return JsonResponse({'error': 'Null Недопустимое значение для поля Value'}, status=400)
                    else:
                        ote = f'Недопустимый значение для:'+str(day)+ ' дня у ' +str(p.full_name)
                        return JsonResponse({'error': 'Недопустимое значение для поля Value','other': ote}, status=400)
            else:
                return JsonResponse({'error': 'Неверный день'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    def save_add(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            value = float(request.POST.get('value'))
            type = request.POST.get('type')
            if type == 'd':
                try:
                    p = driver_list.objects.get(id=id)
                    value = float(value)
                    p.add = value
                    p.save()
                    return JsonResponse({'message': 'Данные успешно сохранены'})
                except:
                    return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
            elif type == 'p':
                try:
                    p = tabel_list.objects.get(id=id)
                    value = float(value)
                    p.add = value
                    p.save()
                    return JsonResponse({'message': 'Данные успешно сохранены'})
                except:
                    return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    def save_auto(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            value = int(request.POST.get('value'))
            p = driver_list.objects.get(id=id)
            try:
                auto = automobile.objects.get(id=value)
                p.auto = auto
                p.save()
                return JsonResponse({'message': 'Данные успешно сохранены'})
            except:
                return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    @login_required(login_url="/login/")
    def view_tabel(request, id):
        groups = group.objects.all()
        per = period.objects.get(id=id)
        pers = period.objects.all()
        user = request.user
        approvers = dayapprove.objects.all().filter(per=per)
        persones = persone.objects.all().filter().order_by('full_name')
        tabs = tabel_list.objects.all().filter(period=per).order_by('full_name')
        html_template = loader.get_template('tabels/worker/view_tabel.html')
        today = datetime.date.today()
        next_month = today.replace(day=1) + datetime.timedelta(days=32)
        if per.date.month == today.month or next_month.month == per.date.month or request.user.username == 'KaratovaMadina' or request.user.username == 'DjavathanovaP':
            for t in tabs:
                if t.company:
                    if t.company != t.persone.company:
                        t.company = t.persone.company
                        t.save()
                if t.snils:
                    if t.snils != t.persone.snils:
                        t.snils = t.persone.snils
                        t.save()
                if t.full_name:
                    if t.full_name != t.persone.full_name:
                        t.full_name = t.persone.full_name
                        t.save()
                if t.graphic:
                    if t.graphic != t.persone.graphic:
                        t.graphic = t.persone.graphic
                        t.save()
                if t.group:
                    if t.group != t.persone.group:
                        t.group = t.persone.group
                        t.save()
                if t.position:
                    if t.position != t.persone.position:
                        t.position = t.persone.position
                        t.save()
        if per.date:
            month = datetime.date.today().month
            pm = per.date.month
            if month > pm:
                if datetime.date.today().year == per.date.year:
                    per.fullblock = 1
                    per.save()
            if per.fullblock != 1:
                today = datetime.date.today().day - 1
                for day in range(1, 32):
                    if today > day:
                        setattr(per, f'db{day}', 1)
                        per.save()
            if per:
                for day in range(1, 32):
                    i = 0
                    for a in approvers:
                        if a.day == day:
                            i += 1
                        if i >= 1:
                            setattr(per, f'da{day}', 1)
                            per.save()
        if request.method == 'POST':
            for t in tabs:
                out = 0
                for i in range(1, 32):
                    day_key = f'd{i}'
                    da = request.POST.get(str(t.id) + day_key)
                    if da:
                        try:
                            da = int(da)
                            if 0 < da <= 23:
                                setattr(t, day_key, da)
                        except ValueError:
                            da = da.upper()
                            if da in ["ВВ", "Н", "М", "Б", "О", "Д", "В", "У"]:
                                setattr(t, day_key, da)
                        t.save()
                h = 0
                d = 0
                for i in range(1, 32):
                    day_key = f'd{i}'
                    try:
                        h_day = int(getattr(t, day_key))
                        h += h_day
                        d += 1
                    except:
                        if getattr(t, day_key) == "ВВ":
                            out += 1
                            if out < 3:
                                h += int(t.persone.work_hours)
                                d += 1
                        elif getattr(t, day_key) == "М":
                            if str(t.group.name) == "ИТР" or str(t.group.name) == "ПТО":
                                h += int(t.persone.work_hours)
                                d += 1
                t.hours = int(h)
                t.days = int(d)
                t.save()
                comment = request.POST.get("com"+str(t.id))
                if comment == '':
                    pass
                else:
                    if t.comment == None:
                        t.comment = ''
                        t.save()
                        comment = str(t.comment) + str(comment)  + ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                    else:
                        comment = str(t.comment) + str(comment)  + ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                try:
                    fine = request.POST.get(str(t.id)+"fine")
                    if fine:
                        t.fine = float(fine)
                    else:
                        t.fine = 0
                except:
                    pass
                try:
                    add = request.POST.get(str(t.id)+"add")
                    if add:
                        t.add = float(add)
                    else:
                        t.add = 0
                except:
                    pass
                f = t.fine
                a = t.add
                today = datetime.date.today()
                sal = 0
                next_month = today.replace(day=1) + datetime.timedelta(days=32)
                if per.date.year == today.year and per.date.month == today.month or next_month.month == per.date.month:
                    if t.persone.sum_method == 1:
                        sal = int(t.persone.salary)
                    #else:
                    #    if str(t.persone.grade) == '2':
                    #        sal = 60900
                    #    elif str(t.persone.grade) == '3':
                    #        sal = 67800
                    #    elif str(t.persone.grade) == '4':
                    #        sal = 72700
                    #    elif str(t.persone.grade) == '5':
                    #        sal = 80000
                    #    elif str(t.persone.grade) == '6':
                    #        sal = 86250
                    #    elif t.persone.salary:
                    #        sal = int(t.persone.salary)
                    #    else:
                    #        sal = 0
                    else:
                        now = datetime.date(day=1,month=2,year=2025)
                        if per.date < now:
                             if str(t.persone.grade) == '2':
                                 sal = 60900
                             elif str(t.persone.grade) == '3':
                                 sal = 67800
                             elif str(t.persone.grade) == '4':
                                 sal = 72700
                             elif str(t.persone.grade) == '5':
                                 sal = 80000
                             elif str(t.persone.grade) == '6':
                                 sal = 86250
                             elif t.persone.salary:
                                 sal = int(t.persone.salary)
                             else:
                                 sal = 0
                        else:
                            if str(t.persone.grade) == '2':
                                sal = 80000
                            elif str(t.persone.grade) == '3':
                                sal = 87000
                            elif str(t.persone.grade) == '4':
                                sal = 92000
                            elif str(t.persone.grade) == '5':
                                sal = 100000
                            elif str(t.persone.grade) == '6':
                                sal = 106000
                            elif str(t.persone.grade) == '1':
                                sal = 60000
                            elif t.persone.salary:
                                sal = int(t.persone.salary)
                            else:
                                sal = 0
                    #if per.graphic == 1:
                    #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                    #else:
                    #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                else:
                    ps = persone_salary.objects.all().filter(persone=t.persone,date=per.date).order_by('-id').last()
                    if t.persone.sum_method == 1:
                        sal = int(ps.salary)
                    #else:
                    #    if ps.grade == '2':
                    #        sal = 60900
                    #    elif ps.grade == '3':
                    #        sal = 67800
                    #    elif ps.grade == '4':
                    #        sal = 72700
                    #    elif ps.grade == '5':
                    #        sal = 80000
                    #    elif ps.grade == '6':
                    #        sal = 86250
                    #    elif ps.salary:
                    #        sal = int(ps.salary)
                    #    else:
                    #        sal = 0
                    else:
                        now = datetime.date(day=1,month=2,year=2025)
                        if per.date < now:
                            if ps.grade == '2':
                                sal = 60900
                            elif ps.grade == '3':
                                sal = 67800
                            elif ps.grade == '4':
                                sal = 72700
                            elif ps.grade == '5':
                                sal = 80000
                            elif ps.grade == '6':
                                sal = 86250
                            elif ps.salary:
                                sal = int(ps.salary)
                            else:
                                sal = 0
                        else:
                            try:
                                if ps.grade == '2':
                                    sal = 80000
                                elif ps.grade == '3':
                                    sal = 87000
                                elif ps.grade == '4':
                                    sal = 92000
                                elif ps.grade == '5':
                                    sal = 100000
                                elif ps.grade == '6':
                                    sal = 106000
                                elif ps.grade == '1':
                                    sal = 60000
                                elif ps.salary:
                                    sal = int(ps.salary)
                                else:
                                    sal = 0
                            except:
                                sal = 0
                if t.graphic == 2:
                    t.sum = sal / int(per.workdays5) * t.days
                else:
                    t.sum = sal / int(per.workdays) * t.days
                if t.persone.exception == 0:
                    t.salary = t.sum - float(f) + float(a)
                else:
                    t.salary = 0
                t.save()
            wd = request.POST.get('wd');
            if int(wd) != per.workdays:
                per.workdays = int(wd)
                per.save()
            wd5 = request.POST.get('wd5');
            if int(wd5) != per.workdays5:
                per.workdays5 = int(wd5)
                per.save()
            return redirect('/view_tabel/' + str(id))
        else:
            context = {
                'user': user,
                'approvers':approvers,
                'persones': persones,
                'groups': groups,
                'per': per,
                'segment': 'tabel',
                'tabs': tabs,
                'pers': pers,
            }
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_tabel_days(request, id):
        groups = group.objects.all()
        per = period.objects.get(id=id)
        pers = period.objects.all()
        user = request.user
        approvers = dayapprove.objects.all().filter(per=per)
        persones = persone.objects.all().filter().order_by('full_name')
        tabs = tabel_list.objects.all().filter(period=per).order_by('full_name')
        html_template = loader.get_template('tabels/worker/view_tabel.html')
        today = datetime.date.today()
        next_month = today.replace(day=1) + datetime.timedelta(days=32)
        if per.date.month == today.month or next_month.month == per.date.month or request.user.username == 'KaratovaMadina' or request.user.username == 'DjavathanovaP':
            for t in tabs:
                if t.company:
                    if t.company != t.persone.company:
                        t.company = t.persone.company
                        t.save()
                if t.snils:
                    if t.snils != t.persone.snils:
                        t.snils = t.persone.snils
                        t.save()
                if t.full_name:
                    if t.full_name != t.persone.full_name:
                        t.full_name = t.persone.full_name
                        t.save()
                if t.graphic:
                    if t.graphic != t.persone.graphic:
                        t.graphic = t.persone.graphic
                        t.save()
                if t.group:
                    if t.group != t.persone.group:
                        t.group = t.persone.group
                        t.save()
                if t.position:
                    if t.position != t.persone.position:
                        t.position = t.persone.position
                        t.save()
        if per.date:
            month = datetime.date.today().month
            pm = per.date.month
            if month > pm:
                if datetime.date.today().year == per.date.year:
                    per.fullblock = 1
                    per.save()
            if per.fullblock != 1:
                today = datetime.date.today().day - 1
                for day in range(26, 32):
                    if today > day:
                        setattr(per, f'db{day}', 1)
                        per.save()
            if per:
                for day in range(26, 32):
                    i = 0
                    for a in approvers:
                        if a.day == day:
                            i += 1
                        if i >= 1:
                            setattr(per, f'da{day}', 1)
                            per.save()
        if request.method == 'POST':
            for t in tabs:
                out = 0
                for i in range(26, 32):
                    day_key = f'd{i}'
                    da = request.POST.get(str(t.id) + day_key)
                    if da:
                        try:
                            da = int(da)
                            if 0 < da <= 23:
                                setattr(t, day_key, da)
                        except ValueError:
                            da = da.upper()
                            if da in ["ВВ", "Н", "М", "Б", "О", "Д", "В", "У"]:
                                setattr(t, day_key, da)
                        t.save()
                h = 0
                d = 0
                for i in range(26, 32):
                    day_key = f'd{i}'
                    try:
                        h_day = int(getattr(t, day_key))
                        h += h_day
                        d += 1
                    except:
                        if getattr(t, day_key) == "ВВ":
                            out += 1
                            if out < 3:
                                h += int(t.persone.work_hours)
                                d += 1
                        elif getattr(t, day_key) == "М":
                            if str(t.group.name) == "ИТР" or str(t.group.name) == "ПТО":
                                h += int(t.persone.work_hours)
                                d += 1
                t.hours = int(h)
                t.days = int(d)
                t.save()
                comment = request.POST.get("com" + str(t.id))
                if comment == '':
                    pass
                else:
                    if t.comment == None:
                        t.comment = ''
                        t.save()
                        comment = str(t.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(
                            request.user.first_name) + ' ' + str(
                            datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                    else:
                        comment = str(t.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(
                            request.user.first_name) + ' ' + str(
                            datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                try:
                    fine = request.POST.get(str(t.id) + "fine")
                    if fine:
                        t.fine = float(fine)
                    else:
                        t.fine = 0
                except:
                    pass
                try:
                    add = request.POST.get(str(t.id) + "add")
                    if add:
                        t.add = float(add)
                    else:
                        t.add = 0
                except:
                    pass
                f = t.fine
                a = t.add
                today = datetime.date.today()
                sal = 0
                next_month = today.replace(day=1) + datetime.timedelta(days=32)
                if per.date.year == today.year and per.date.month == today.month or next_month.month == per.date.month:
                    if t.persone.sum_method == 1:
                        sal = int(t.persone.salary)
                    # else:
                    #    if str(t.persone.grade) == '2':
                    #        sal = 60900
                    #    elif str(t.persone.grade) == '3':
                    #        sal = 67800
                    #    elif str(t.persone.grade) == '4':
                    #        sal = 72700
                    #    elif str(t.persone.grade) == '5':
                    #        sal = 80000
                    #    elif str(t.persone.grade) == '6':
                    #        sal = 86250
                    #    elif t.persone.salary:
                    #        sal = int(t.persone.salary)
                    #    else:
                    #        sal = 0
                    else:
                        now = datetime.date(day=1, month=2, year=2025)
                        if per.date < now:
                            if str(t.persone.grade) == '2':
                                sal = 60900
                            elif str(t.persone.grade) == '3':
                                sal = 67800
                            elif str(t.persone.grade) == '4':
                                sal = 72700
                            elif str(t.persone.grade) == '5':
                                sal = 80000
                            elif str(t.persone.grade) == '6':
                                sal = 86250
                            elif t.persone.salary:
                                sal = int(t.persone.salary)
                            else:
                                sal = 0
                        else:
                            if str(t.persone.grade) == '2':
                                sal = 80000
                            elif str(t.persone.grade) == '3':
                                sal = 87000
                            elif str(t.persone.grade) == '4':
                                sal = 92000
                            elif str(t.persone.grade) == '5':
                                sal = 100000
                            elif str(t.persone.grade) == '6':
                                sal = 106000
                            elif str(t.persone.grade) == '1':
                                sal = 60000
                            elif t.persone.salary:
                                sal = int(t.persone.salary)
                            else:
                                sal = 0
                    # if per.graphic == 1:
                    #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                    # else:
                    #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                else:
                    ps = persone_salary.objects.all().filter(persone=t.persone, date=per.date).order_by('-id').last()
                    if t.persone.sum_method == 1:
                        sal = int(ps.salary)
                    # else:
                    #    if ps.grade == '2':
                    #        sal = 60900
                    #    elif ps.grade == '3':
                    #        sal = 67800
                    #    elif ps.grade == '4':
                    #        sal = 72700
                    #    elif ps.grade == '5':
                    #        sal = 80000
                    #    elif ps.grade == '6':
                    #        sal = 86250
                    #    elif ps.salary:
                    #        sal = int(ps.salary)
                    #    else:
                    #        sal = 0
                    else:
                        now = datetime.date(day=1, month=2, year=2025)
                        if per.date < now:
                            if ps.grade == '2':
                                sal = 60900
                            elif ps.grade == '3':
                                sal = 67800
                            elif ps.grade == '4':
                                sal = 72700
                            elif ps.grade == '5':
                                sal = 80000
                            elif ps.grade == '6':
                                sal = 86250
                            elif ps.salary:
                                sal = int(ps.salary)
                            else:
                                sal = 0
                        else:
                            try:
                                if ps.grade == '2':
                                    sal = 80000
                                elif ps.grade == '3':
                                    sal = 87000
                                elif ps.grade == '4':
                                    sal = 92000
                                elif ps.grade == '5':
                                    sal = 100000
                                elif ps.grade == '6':
                                    sal = 106000
                                elif ps.grade == '1':
                                    sal = 60000
                                elif ps.salary:
                                    sal = int(ps.salary)
                                else:
                                    sal = 0
                            except:
                                sal = 0
                if t.graphic == 2:
                    t.sum = sal / int(per.workdays5) * t.days
                else:
                    t.sum = sal / int(per.workdays) * t.days
                if t.persone.exception == 0:
                    t.salary = t.sum - float(f) + float(a)
                else:
                    t.salary = 0
                t.save()
            wd = request.POST.get('wd');
            if int(wd) != per.workdays:
                per.workdays = int(wd)
                per.save()
            wd5 = request.POST.get('wd5');
            if int(wd5) != per.workdays5:
                per.workdays5 = int(wd5)
                per.save()
            return redirect('/view_tabel/' + str(id))
        else:
            context = {
                'user': user,
                'approvers': approvers,
                'persones': persones,
                'groups': groups,
                'per': per,
                'segment': 'tabel',
                'tabs': tabs,
                'pers': pers,
            }
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def group_move_tabel(request, ids, id):
        per_old = period.objects.get(id=ids)
        tabs = tabel_list.objects.all().filter(period=per_old)
        per = period.objects.get(id=id)
        for t in tabs:
            p = persone.objects.get(id=t.persone.id)
            if p.group == None:
                g = group.objects.get(id=int(t.group.id))
                p.group = g
                p.save()
            tabel_list(
                period=per,
                persone=p,
                company=p.company,
                category=p.category,
                snils=p.snils,
                full_name=p.full_name,
                group=p.group,
                position=p.position,
                salary=p.salary,
                output=p.output,
            ).save()
        return redirect('/view_tabel/' + str(id))


    @login_required(login_url="/login/")
    def approve_table(request, day, ids):
        per = period.objects.get(id=ids)
        dayapprove(
            day=day,
            per=per,
            approver=request.user,
            time=datetime.datetime.now()
        ).save()
        return redirect('/view_tabel/'+str(ids))

    @login_required(login_url="/login/")
    def approve_tableD(request, day, ids):
        per = driver.objects.get(id=ids)
        dayapproved(
            day=day,
            per=per,
            approver=request.user,
            time=datetime.datetime.now()
        ).save()
        return redirect('/view_driver/' + str(ids))

    @login_required(login_url="/login/")
    def del_tabel(request, id):
        per = period.objects.get(id=id)
        tabs = tabel_list.objects.all().filter(period=per)
        das = dayapprove.objects.all().filter(per=per)
        for d in das:
            d.delete()
        for t in tabs:
            t.delete()
        per.delete()
        return redirect('tabel')

    @login_required(login_url="/login/")
    def del_driver(request, id):
        per = driver.objects.get(id=id)
        tabs = driver_list.objects.all().filter(period=per)
        das = dayapproved.objects.all().filter(per=per)
        for d in das:
            d.delete()
        for t in tabs:
            t.delete()
        per.delete()
        return redirect('driver')


    @login_required(login_url="/login/")
    @cherrypy.expose
    def hexp_tabel(request,name,id):
        per = period.objects.get(id=id)
        tabs = tabel_list.objects.all().filter(period=per)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('templ.xlsx')
        sheet : worksheet = book.worksheets[0]
        m = int(per.date.month)-1
        month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря']
        my = str(month[m])+' '+ str(per.date.strftime("%Y"))
        sheet["E3"].value = 'Табель от '+my
        sheet["K4"].value = str(per.obj)
        i = 8
        for t in tabs:
            sheet.insert_rows(i)
            sheet["A"+str(i)].value = ""+str(t.id)
            sheet["B" + str(i)].value = "" + str(t.group)
            sheet["C" + str(i)].value = ""+str(t.persone.snils)
            if t.persone.company == 1:
                sheet["D" + str(i)].value = "ПМК"
            else:
                sheet["D" + str(i)].value = "неоформ"
            sheet["E" + str(i)].value = ""+str(t.persone.full_name)
            sheet["F" + str(i)].value = ""+str(t.persone.position)
            if int(t.persone.salary) == 0:
                sal = t.persone.salary
                now = datetime.date(day=1, month=2, year=2025)
                if per.date < now:
                    if str(t.persone.grade) == '2':
                        sal = 60900
                    elif str(t.persone.grade) == '3':
                        sal = 67800
                    elif str(t.persone.grade) == '4':
                        sal = 72700
                    elif str(t.persone.grade) == '5':
                        sal = 80000
                    elif str(t.persone.grade) == '6':
                        sal = 86250
                else:
                    if str(t.persone.grade) == '1':
                        sal = 60000
                    elif str(t.persone.grade) == '2':
                        sal = 80000
                    elif str(t.persone.grade) == '3':
                        sal = 87000
                    elif str(t.persone.grade) == '4':
                        sal = 92000
                    elif str(t.persone.grade) == '5':
                        sal = 100000
                    elif str(t.persone.grade) == '6':
                        sal = 106000
                sheet["G" + str(i)].value = "" + str(t.persone.grade) + " разряд| "+str(sal)
            else:
                sheet["G" + str(i)].value = ""+str(int(float(t.persone.salary)))
            if t.d1:sheet["H" + str(i)].value = ""+str(t.d1)
            if t.d2:sheet["I" + str(i)].value = ""+str(t.d2)
            if t.d3:sheet["J" + str(i)].value = ""+str(t.d3)
            if t.d4:sheet["K" + str(i)].value = ""+str(t.d4)
            if t.d5:sheet["L" + str(i)].value = ""+str(t.d5)
            if t.d6:sheet["M" + str(i)].value = ""+str(t.d6)
            if t.d7:sheet["N" + str(i)].value = ""+str(t.d7)
            if t.d8:sheet["O" + str(i)].value = ""+str(t.d8)
            if t.d9:sheet["P" + str(i)].value = ""+str(t.d9)
            if t.d10:sheet["Q" + str(i)].value = ""+str(t.d10)
            if t.d11:sheet["R" + str(i)].value = ""+str(t.d11)
            if t.d12:sheet["S" + str(i)].value = ""+str(t.d12)
            if t.d13:sheet["T" + str(i)].value = ""+str(t.d13)
            if t.d14:sheet["U" + str(i)].value = ""+str(t.d14)
            if t.d15:sheet["V" + str(i)].value = ""+str(t.d15)
            if t.d16:sheet["W" + str(i)].value = ""+str(t.d16)
            if t.d17:sheet["X" + str(i)].value = ""+str(t.d17)
            if t.d18:sheet["Y" + str(i)].value = ""+str(t.d18)
            if t.d19:sheet["Z" + str(i)].value = ""+str(t.d19)
            if t.d20:sheet["AA" + str(i)].value = ""+str(t.d20)
            if t.d21:sheet["AB" + str(i)].value = ""+str(t.d21)
            if t.d22:sheet["AC" + str(i)].value = ""+str(t.d22)
            if t.d23:sheet["AD" + str(i)].value = ""+str(t.d23)
            if t.d24:sheet["AE" + str(i)].value = ""+str(t.d24)
            if t.d25:sheet["AF" + str(i)].value = ""+str(t.d25)
            if t.d26:sheet["AG" + str(i)].value = ""+str(t.d26)
            if t.d27:sheet["AH" + str(i)].value = ""+str(t.d27)
            if t.d28:sheet["AI" + str(i)].value = ""+str(t.d28)
            if t.d29:sheet["AJ" + str(i)].value = ""+str(t.d29)
            if t.d30:sheet["AK" + str(i)].value = ""+str(t.d30)
            if t.d31:sheet["AL" + str(i)].value = ""+str(t.d31)
            sheet["AM" + str(i)].value = "" + str(t.days)
            if request.user.has_perm('reference.view_temp_file'):
                if t.fine:sheet["AN" + str(i)].value = "" + str(int(float(t.fine)))
                if t.add:sheet["AO" + str(i)].value = "" + str(int(float(t.add)))
                if t.salary:sheet["AP" + str(i)].value = "" + str(int(float(t.salary)))
            else:
                sheet["AN5"].value = None
                sheet["AN6"].value = None
                sheet["AN7"].value = None
                sheet["AO5"].value = None
                sheet["AO6"].value = None
                sheet["AO7"].value = None
                sheet["AP5"].value = None
                sheet["AP6"].value = None
                sheet["AP7"].value = None
                sheet["AQ5"].value = None
                sheet["AQ6"].value = None
                sheet["AQ7"].value = None

                sheet["AN5"].fill = PatternFill(fill_type=None)
                sheet["AN6"].fill = PatternFill(fill_type=None)
                sheet["AN7"].fill = PatternFill(fill_type=None)
                sheet["AO5"].fill = PatternFill(fill_type=None)
                sheet["AO6"].fill = PatternFill(fill_type=None)
                sheet["AO7"].fill = PatternFill(fill_type=None)
                sheet["AP5"].fill = PatternFill(fill_type=None)
                sheet["AP6"].fill = PatternFill(fill_type=None)
                sheet["AP7"].fill = PatternFill(fill_type=None)
                sheet["AQ5"].fill = PatternFill(fill_type=None)
                sheet["AQ6"].fill = PatternFill(fill_type=None)
                sheet["AQ7"].fill = PatternFill(fill_type=None)

                sheet["AN5"].border = None
                sheet["AN6"].border = None
                sheet["AN7"].border = None
                sheet["AO5"].border = None
                sheet["AO6"].border = None
                sheet["AO7"].border = None
                sheet["AP5"].border = None
                sheet["AP6"].border = None
                sheet["AP7"].border = None
                sheet["AQ5"].border = None
                sheet["AQ6"].border = None
                sheet["AQ7"].border = None
            if t.comment: sheet["AQ" + str(i)].value = "" + str(t.d31)
            if request.user.username == 'KaratovaMadina' or request.user.username == 'DjavathanovaP' or request.user.is_superuser:
                sheet["AR6"].value = "Заметки"
                if t.persone.idea:sheet["AR" + str(i)].value = str(t.persone.idea)
            i+=1
        book.title='Табель от '+my
        book.save(response)
        return response
        return redirect('home')

    @login_required(login_url="/login/")
    @cherrypy.expose
    def hexp_driver(request,name,id):
        per = driver.objects.get(id=id)
        tabs = driver_list.objects.all().filter(period=per)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('templd.xlsx')
        sheet : worksheet = book.worksheets[0]
        m = int(per.date.month)-1
        month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря']
        my = str(month[m])+' '+ str(per.date.strftime("%Y"))
        sheet["D3"].value = 'Табель учета использования рабочего времени механизаторов от'+my
        sheet["Q4"].value = str(per.obj)
        i = 8
        for t in tabs:
            sheet.insert_rows(i)
            sheet["A" + str(i)].value = "" + str(t.persone.group)
            sheet["B" + str(i)].value = ""+str(t.persone.snils)
            if t.persone.company == 1:
                sheet["C" + str(i)].value = "ПМК"
            else:
                sheet["C" + str(i)].value = "Неоформл"
            sheet["D" + str(i)].value = ""+str(t.persone.full_name)
            if t.auto: sheet["E" + str(i)].value = ""+str(t.auto.name)+"["+str(t.auto.number)+"]"
            sheet["F" + str(i)].value = ""+str(t.persone.position)
            #sheet["F" + str(i)].value = ""+str(int(int(t.persone.salary)/300))
            if t.d1:sheet["G" + str(i)].value = ""+str(t.d1)
            if t.d2:sheet["H" + str(i)].value = ""+str(t.d2)
            if t.d3:sheet["I" + str(i)].value = ""+str(t.d3)
            if t.d4:sheet["J" + str(i)].value = ""+str(t.d4)
            if t.d5:sheet["K" + str(i)].value = ""+str(t.d5)
            if t.d6:sheet["L" + str(i)].value = ""+str(t.d6)
            if t.d7:sheet["M" + str(i)].value = ""+str(t.d7)
            if t.d8:sheet["N" + str(i)].value = ""+str(t.d8)
            if t.d9:sheet["O" + str(i)].value = ""+str(t.d9)
            if t.d10:sheet["P" + str(i)].value = ""+str(t.d10)
            if t.d11:sheet["Q" + str(i)].value = ""+str(t.d11)
            if t.d12:sheet["R" + str(i)].value = ""+str(t.d12)
            if t.d13:sheet["S" + str(i)].value = ""+str(t.d13)
            if t.d14:sheet["T" + str(i)].value = ""+str(t.d14)
            if t.d15:sheet["U" + str(i)].value = ""+str(t.d15)
            if t.d16:sheet["V" + str(i)].value = ""+str(t.d16)
            if t.d17:sheet["W" + str(i)].value = ""+str(t.d17)
            if t.d18:sheet["X" + str(i)].value = ""+str(t.d18)
            if t.d19:sheet["Y" + str(i)].value = ""+str(t.d19)
            if t.d20:sheet["Z" + str(i)].value = ""+str(t.d20)
            if t.d21:sheet["AA" + str(i)].value = ""+str(t.d21)
            if t.d22:sheet["AB" + str(i)].value = ""+str(t.d22)
            if t.d23:sheet["AC" + str(i)].value = ""+str(t.d23)
            if t.d24:sheet["AD" + str(i)].value = ""+str(t.d24)
            if t.d25:sheet["AE" + str(i)].value = ""+str(t.d25)
            if t.d26:sheet["AF" + str(i)].value = ""+str(t.d26)
            if t.d27:sheet["AG" + str(i)].value = ""+str(t.d27)
            if t.d28:sheet["AH" + str(i)].value = ""+str(t.d28)
            if t.d29:sheet["AI" + str(i)].value = ""+str(t.d29)
            if t.d30:sheet["AJ" + str(i)].value = ""+str(t.d30)
            if t.d31:sheet["AK" + str(i)].value = ""+str(t.d31)
            sheet["AL" + str(i)].value = "" + str(t.days)
            if request.user.has_perm('reference.view_temp_file'):
                if t.salary:sheet["AO" + str(i)].value = "" + str(int(float(t.salary)))
            else:
                sheet["AP5"].value = None
                sheet["AP6"].value = None
                sheet["AP7"].value = None
                sheet["AQ5"].value = None
                sheet["AQ6"].value = None
                sheet["AQ7"].value = None
                sheet["AR5"].value = None
                sheet["AR6"].value = None
                sheet["AR7"].value = None


                sheet["AP5"].fill = PatternFill(fill_type=None)
                sheet["AP6"].fill = PatternFill(fill_type=None)
                sheet["AP7"].fill = PatternFill(fill_type=None)
                sheet["AQ5"].fill = PatternFill(fill_type=None)
                sheet["AQ6"].fill = PatternFill(fill_type=None)
                sheet["AQ7"].fill = PatternFill(fill_type=None)
                sheet["AR5"].fill = PatternFill(fill_type=None)
                sheet["AR6"].fill = PatternFill(fill_type=None)
                sheet["AR7"].fill = PatternFill(fill_type=None)


                sheet["AQ5"].border = None
                sheet["AQ6"].border = None
                sheet["AQ7"].border = None
                sheet["AR5"].border = None
                sheet["AR6"].border = None
                sheet["AR7"].border = None
                sheet["AP5"].border = None
                sheet["AP6"].border = None
                sheet["AP7"].border = None

            i+=1
        book.title='Табель водителей от '+my
        book.save(response)
        return response
        return redirect('view_driver/'+str(id))

    # views.py
    from django.http import JsonResponse
    from django.views.decorators.csrf import csrf_exempt

    @csrf_exempt
    def save_input_value(request):
        if request.method == 'POST' and request.is_ajax():
            input_id = request.POST.get('inputId')
            input_value = request.POST.get('value')
            # Здесь вы можете сохранить значение в базе данных или файле
            # Например:
            # YourModel.objects.create(input_id=input_id, value=input_value)
            return JsonResponse({'message': 'Значение успешно сохранено'})
        else:
            return JsonResponse({'message': 'Недопустимый запрос'}, status=400)
    class gaz:

        def save_gaz(request):
            if request.method == 'POST':
                id = request.POST.get('id')
                value = int(request.POST.get('value'))
                p = gaz_tabel_head.objects.get(id=id)
                try:
                    gaz_input(
                        gaz_obj=p.gaz,
                        gaz_add=float(value),
                        date=datetime.date.today()
                    ).save()
                    ids = p.gaz.id
                    gaz = gazolin.objects.get(id=ids)
                    gaz.sum = gaz.sum + float(value)
                    gaz.save()
                    return JsonResponse({'message': 'Данные успешно сохранены'})
                except:
                    return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
            else:
                return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

        @login_required(login_url="/login/")
        def AddPersoneG(request, id, ids):
            per = gaz_tabel_head.objects.get(id=ids)
            p = persone.objects.get(id=id)
            gaz_tabel_body(
                tabel=per,
                persone=p,
                company=p.company,
                full_name=p.full_name,
                position=p.position,
            ).save()
            return redirect('/view_gaz/' + str(ids))

        @login_required(login_url="/login/")
        def gaz(request):
            persones = persone.objects.all()
            per = gaz_tabel_head.objects.all()
            groups = tabel_group.objects.all()
            if groups:
                pass
            else:
                groups
            context = {'per': per, 'segment': 'gaz', 'groups': groups, 'persones': persones}
            html_template = loader.get_template('tabels/gaz/gaz.html')
            return HttpResponse(html_template.render(context, request))

        @login_required(login_url="/login/")
        def view_gaz(request, id):
            per = gaz_tabel_head.objects.get(id=id)
            pers = gaz_tabel_head.objects.all()
            persones = persone.objects.all().filter(date_leave=None).order_by('full_name')
            all = persone.objects.all().filter(date_leave=None).order_by('full_name')
            tabs = gaz_tabel_body.objects.all().filter(tabel=per).order_by('full_name')
            autos = automobile.objects.all().order_by('name')
            html_template = loader.get_template('tabels/gaz/view_gaz.html')
            if per:
                s_minus = 0
                for t in tabs:
                    minus = 0
                    for i in range(1, 32):  # Проходим по диапазону от 1 до 31
                        attr_name = 'd' + str(i)  # Формируем название атрибута
                        if hasattr(t, attr_name) and getattr(t,
                                                             attr_name) is not None:  # Проверяем, существует ли атрибут и не является ли он None
                            try:
                                minus += float(
                                    getattr(t, attr_name))  # Пытаемся преобразовать значение в число и добавить к сумме
                            except (TypeError, ValueError):
                                print(f"Ошибка при преобразовании значения {attr_name} в число")

                    t.sum = minus
                    t.save()
                    s_minus += minus
                per.gaz_minus = s_minus
                per.save()
                sum_minus = 0
                for p in pers:
                    if p.gaz_minus:
                        sum_minus += p.gaz_minus
                gaz = gazolin.objects.get(id=per.gaz.id)
                gaz.gaz_sum = gaz.sum - sum_minus
                gaz.save()
            if request.method == 'POST':
                for t in tabs:
                    if t:
                        for i in range(1, 32):
                            if getattr(t, f'd{i}'):
                                t.auto_block = 1
                        t.save()
                    auto = request.POST.get(str(t.id) + 'auto')
                    try:
                        t.auto = automobile.objects.get(id=int(auto))
                    except:
                        pass
                    comment = request.POST.get("com" + str(t.id))
                    if comment == '':
                        pass
                    else:
                        if t.comment == None:
                            t.comment = ''
                            t.save()
                            comment = str(t.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(
                                request.user.first_name) + ' ' + str(
                                datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                            t.comment = comment
                            t.save()
                        else:
                            comment = str(t.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(
                                request.user.first_name) + ' ' + str(
                                datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                            t.comment = comment
                            t.save()
                    t.save()
                return redirect('/view_gaz/' + str(id))
            else:
                context = {
                    'autos': autos,
                    'persones': persones,
                    'per': per,
                    'segment': 'driver',
                    'tabs': tabs,
                    'pers': pers,
                    'all': all,
                }
                return HttpResponse(html_template.render(context, request))

        @login_required(login_url="/login/")
        def new_gaz(request):
            per = gaz_tabel_head.objects.all()
            form = GazForm
            persones = persone.objects.all().filter(date_leave=None).order_by('full_name')
            if request.method == 'POST':
                year = request.POST.get("ye")
                month = request.POST.get("mo")
                da = year + '-' + month + '-' + '01'
                gaz = gazolin.objects.get(id=int(request.POST.get('gaz')))
                gaz_tabel_head(
                    gaz = gaz,
                    date = da
                ).save()
                post = gaz_tabel_head.objects.last()
                da = date(int(year), int(month), 1)
                for i in range(1, 32):
                    try:
                        day_attr = f"dw{i}"
                        day_value = i
                        da = da.replace(day=day_value)
                        setattr(post, day_attr, da)
                    except:
                        pass
                post.save()
                return redirect('gaz')
                # else:
                #     error = 'Неправильно'
                #     context = {
                #         'error': error,
                #         'segment': 'tabel',
                #         'persones': persones,
                #         'per': per,
                #         'form': form,
                #     }
                #     html_template = loader.get_template('tabels/gaz/new_gaz.html')
                #     return HttpResponse(html_template.render(context, request))
            else:
                context = {
                    'segment': 'tabel',
                    'persones': persones,
                    'per': per,
                    'form': form,
                }
                html_template = loader.get_template('tabels/gaz/new_gaz.html')
                return HttpResponse(html_template.render(context, request))


class reference:
    @login_required(login_url="/login/")
    def ref(request):
        context = {'segment': 'ref'}
        html_template = loader.get_template('reference/ref.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def objs(request):
        obj = objs.objects.all()
        context = {'obj': obj, 'segment': 'task'}
        html_template = loader.get_template('reference/object/objects.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_object(request):
        obj = objs.objects.all()
        forms = ObjsForm
        if request.method == 'POST':
            forms = ObjsForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.save()
                return redirect('objs')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                           'obj': obj,
                           'forms': forms,
                           'erorr': erorr
                           }
                html_template = loader.get_template('reference/object/new_object.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                       'obj': obj,
                       'forms': forms
                       }
            html_template = loader.get_template('reference/object/new_object.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_object(request, id):
        r = objs.objects.get(id=id)
        context = {'r': r, 'segment': 'ref'}
        html_template = loader.get_template('reference/object/view_object.html')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
            name = request.POST.get("name")
            if name == str(r.name):
                pass
            else:
                r.name = name
                r.save()
            status = request.POST.get('status')
            if status == str(r.status):
                pass
            else:
                r.status = status
                r.save()
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def auto(request):
        auto = automobile.objects.all().order_by('name')
        context = {'auto': auto, 'segment': 'ref'}
        html_template = loader.get_template('reference/auto/auto.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_auto(request):
        auto = automobile.objects.all().order_by('name')
        forms = AutoForm
        if request.method == 'POST':
            forms = AutoForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.save()
                return redirect('auto')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                            'auto': auto,
                            'forms': forms,
                            'erorr': erorr
                            }
                html_template = loader.get_template('reference/auto/new_auto.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                        'auto': auto,
                        'forms': forms
                        }
            html_template = loader.get_template('reference/auto/new_auto.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_auto(request, id):
        r = automobile.objects.get(id=id)
        context = {'r': r, 'segment': 'ref'}
        html_template = loader.get_template('reference/auto/view_auto.html')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
            name = request.POST.get("name")
            owner = request.POST.get("owner")
            number = request.POST.get("number")

            if name != str(r.name):
                r.name = name
            if owner != str(r.owner):
                r.owner = owner
            if number != str(r.number):
                r.number = number
            r.save()
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def group(request):
        groups = group.objects.all()
        context = {'group': groups, 'segment': 'ref'}
        html_template = loader.get_template('reference/group/group.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_group(request):
        groups = group.objects.all()
        forms = GroupForm
        if request.method == 'POST':
            forms = GroupForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.save()
                return redirect('group')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                           'group': groups,
                           'forms': forms,
                           'erorr': erorr
                           }
                html_template = loader.get_template('reference/group/new_group.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                       'group': groups,
                       'forms': forms
                       }
            html_template = loader.get_template('reference/group/new_group.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_group(request, id):
        r = group.objects.get(id=id)
        context = {'r': r, 'segment': 'ref'}
        html_template = loader.get_template('reference/group/view_group.html')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
            name = request.POST.get("name")
            if name == str(r.name):
                pass
            else:
                r.name = name
                r.save()
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def persones(request):
        pers = persone.objects.all().order_by('-id')
        groups = group.objects.all()
        today = datetime.date.today()
        if misc.is_last_day_of_month(today):
            pre_month = datetime.date(day=31,month=today.month,year=today.year)
            for p in pers:
                persone_salary(
                    persone=p,
                    date=pre_month,
                    salary=p.salary,
                    grade=p.grade,
                ).save()
                try:
                    tax = driver_tax.objects.get(driver=p)
                    if tax:
                        driver_tax_lagacy(
                            driver=tax.driver,
                            tax=tax.tax,
                            date=pre_month
                        ).save()
                except:
                    pass
        context = {'pers': pers,'groups': groups, 'segment': 'ref'}
        html_template = loader.get_template('reference/persone/persones.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def persone_in_tabel(request):
        if request.method == 'POST':
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            start_date = datetime.date(2024,1,1)
            end_date = datetime.date.today()
        ptl = persone.objects.filter(tabel_list__period__date__range=(start_date, end_date)).annotate(num_tabel_list=Count('tabel_list')).filter(num_tabel_list__gt=0)
        pdl = persone.objects.filter(driver_list__period__date__range=(start_date, end_date)).annotate(num_driver_list=Count('driver_list')).filter(num_driver_list__gt=0)
        combined_list = (ptl | pdl).distinct().order_by('full_name')
        context = {
            'start_date':start_date,
            'end_date':end_date,
            'segment': 'ref',
            'persones': combined_list,
            }
        html_template = loader.get_template('reference/persone/persone_in_tabel.html')
        return HttpResponse(html_template.render(context, request))

    def hexp_pit(request,start_date,end_date, name):
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        ptl = persone.objects.filter(tabel_list__period__date__range=(start_date, end_date)).annotate(num_tabel_list=Count('tabel_list')).filter(num_tabel_list__gt=0)
        pdl = persone.objects.filter(driver_list__period__date__range=(start_date, end_date)).annotate(num_driver_list=Count('driver_list')).filter(num_driver_list__gt=0)
        tabs = (ptl | pdl).distinct().order_by('full_name')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('templ_pit.xlsx')
        sheet: worksheet = book.worksheets[0]
        sheet["B3"].value = 'Отчет по персаналу в табелях'
        i = 8
        for t in tabs:
            sheet.insert_rows(i)
            sheet["A" + str(i)].value = "" + str(i-7)
            if t.full_name:sheet["B" + str(i)].value = "" + str(t.full_name)
            if t.company == 1:
                org = 'ПМК'
            else:
                org = 'неоформ'
            sheet["C" + str(i)].value = "" + str(org)
            if t.idea:sheet["D" + str(i)].value = "" + str(t.idea)
            sheet["E" + str(i)].value = "" + str(t.position)
            if t.date_accept:
                accept = t.date_accept.strftime("%d.%m.%y")
            sheet["F" + str(i)].value = "" + str(accept)
            if t.leaved == 0:
                status = 'Работает'
            else:
                status = 'Уволен'
            sheet["G" + str(i)].value = "" + str(status)
            if t.phone_number:sheet["H" + str(i)].value = "" + str(t.phone_number)
            if t.resident:sheet["I" + str(i)].value = "" + str(t.resident)
            if t.comment:sheet["J" + str(i)].value = "" + str(t.comment)
            i += 1
        book.title = 'Отчет о персонажах'
        book.save(response)
        return response
        return redirect('/persones_in_tabel')
    def save_idea(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            value = request.POST.get('value')

            if not id or not value:
                return JsonResponse({'error': 'Отсутствуют необходимые параметры'}, status=400)

            try:
                value = str(value)
            except ValueError:
                return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)

            p = get_object_or_404(persone, id=id)

            if str(value) != str(p.idea):
                p.idea = str(value)
                p.save()
                return JsonResponse({'message': 'Данные успешно сохранены'})
            else:
                return JsonResponse({'error': 'Значения не совпадают'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    @login_required(login_url="/login/")
    def new_persone(request):
        pers = persone.objects.all().order_by('full_name')
        groups = group.objects.all().order_by('name')
        forms = PersoneForm
        if request.method == 'POST':
            forms = PersoneForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                p = persone.objects.all().last()
                try:
                    post.id = int(p.id)+1
                except:
                    pass
                post.position = request.POST.get("position")
                post.workgroup = request.POST.get("workgroup")
                post.idea = request.POST.get("idea")
                comments = request.POST.get("comment")
                comments = comments + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                post.comment = comments
                post.save()
                return redirect('persones')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                           'pers': pers,
                           'forms': forms,
                           'erorr': erorr,
                           'groups':groups
                           }
                html_template = loader.get_template('reference/persone/new_persone.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                       'pers': pers,
                       'forms': forms,
                       'groups': groups
                       }
            html_template = loader.get_template('reference/persone/new_persone.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def returning(request, id):
        return redirect()

    @login_required(login_url="/login/")
    def view_persone(request, id):
        r = persone.objects.get(id=id)
        try :
            tax = driver_tax.objects.get(driver=r)
        except:
            driver_tax(driver=r,tax=0).save()
            tax = driver_tax.objects.get(driver=r)
        groups = group.objects.all().order_by('name')
        cats = category.objects.all().order_by('name')
        persones = persone.objects.all().order_by('-id')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(comment) + ': ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                    r.comment = comment
                    r.save()
            if request.user.username == 'AhmedovMagomed' or request.user.username == 'IdrisovA' or request.user.is_superuser:
                try:
                    taxs = request.POST.get("tax")
                    if taxs:
                        tax.tax = taxs
                        tax.save()
                except:
                    pass
                try:
                    driver = request.POST.get('driver')
                    if driver == str(r.driver):
                        pass
                    else:
                        r.driver = int(driver)
                        r.save()
                except:
                    pass
                if request.user.is_superuser:
                    try:
                        res = request.POST.get('resident')
                        if res=='РФ' or res=='Иностранец' :
                            r.resident = res
                            r.save()
                    except:
                        pass
                    try:
                        method = request.POST.get("sum_method")
                        if method:
                            r.sum_method = '1'
                            r.save()
                        else:
                            r.sum_method = '0'
                            r.save()
                    except:
                        r.sum_method = '0'
                        r.save()
                    try:
                        exception = request.POST.get("exception")
                        if exception:
                            r.exception = '1'
                            r.save()
                        else:
                            r.exception = '0'
                            r.save()
                    except:
                        r.exception = '0'
                        r.save()
                    try:
                        cat = request.POST.get("category")
                        if cat:
                            cat = category.objects.get(id=int(cat))
                            r.category = cat
                            r.save()
                    except:
                        pass
                    try:
                        date_a = request.POST.get("date_a")
                        if date_a:
                            if date_a == r.date_accept:
                                pass
                            else:
                                r.date_accept = date_a
                                r.save()
                        else:
                            if date_a == r.date_accept:
                                pass
                            else:
                                r.date_accept = None
                                r.save()
                    except:
                        pass
                    try:
                        date_l = request.POST.get("date_l")
                        if date_l:
                            if date_l == r.date_leave:
                                pass
                            else:
                                r.date_leave = date_l
                                r.save()
                        else:
                            if date_l == r.date_leave:
                                pass
                            else:
                                r.date_leave = None
                                r.save()
                    except:
                        pass
                    try:
                        company = request.POST.get("company")
                        if company == str(r.company):
                            pass
                        else:
                            r.company = company
                            r.save()
                    except:
                        pass
                    try:
                        snils = request.POST.get('snils')
                        if snils == str(r.snils):
                            pass
                        else:
                            r.snils = snils
                            r.save()
                    except:
                        pass
                    try:
                        inn = request.POST.get('inn')
                        if snils == str(r.inn):
                            pass
                        else:
                            r.inn = inn
                            r.save()
                    except:
                        pass
                    try:
                        full_name = request.POST.get('full_name')
                        if full_name == str(r.full_name):
                            pass
                        else:
                            r.full_name = full_name
                            r.save()
                    except:
                        pass
                    try:
                        phone_number = request.POST.get('phone_number')
                        r.phone_number = phone_number
                        r.save()
                    except:
                        pass
                    try:
                        position = request.POST.get('position')
                        if position == str(r.position):
                            pass
                        else:
                            r.position = position
                            r.save()
                    except:
                        pass
                    try:
                        workgroup = request.POST.get('workgroup')
                        if workgroup == str(r.workgroup):
                            pass
                        else:
                            r.workgroup = workgroup
                            r.save()
                    except:
                        pass
                    try:
                        g = request.POST.get('group')
                        if r.group:
                            if int(g) == r.group.id:
                                pass
                            else:
                                g = int(g)
                                r.group = group.objects.get(id=g)
                                r.save()
                        else:
                            g = int(g)
                            r.group = group.objects.get(id=g)
                            r.save()
                    except:
                        pass
                    try:
                        grade = request.POST.get('grade')
                        if grade == str(r.grade):
                            pass
                        else:
                            r.grade = grade
                            r.save()
                    except:
                        pass
                    try:
                        salary = request.POST.get('salary')
                        if salary == str(r.salary):
                            pass
                        else:
                            try:
                                salary = int(salary)
                            except:
                                salary = salary.rpartition(',')[0]
                                salary = int(salary)
                            r.salary = float(salary)
                            r.save()
                    except:
                        pass
                    try:
                        graphic = request.POST.get('graphic')
                        if graphic == str(r.graphic):
                            pass
                        else:
                            r.graphic = int(graphic)
                            r.save()
                    except:
                        pass
                    try:
                        output = request.POST.get('output')
                        if output == str(r.output):
                            pass
                        else:
                            r.output = output
                            r.save()
                    except:
                        pass
                    try:
                        driver = request.POST.get('driver')
                        if driver == str(r.driver):
                            pass
                        else:
                            r.driver = int(driver)
                            r.save()
                    except:
                        pass
                    try:
                        leaved = request.POST.get('leaved')
                        if leaved == str(r.leaved):
                            pass
                        else:
                            r.leaved = int(leaved)
                            r.save()
                        work_hours = request.POST.get('work_hours')
                        if work_hours == str(r.work_hours):
                            pass
                        else:
                            r.work_hours = work_hours
                            r.save()
                    except:
                        pass
            else:
                try:
                    res = request.POST.get('resident')
                    if res == 'РФ' or res == 'Иностранец':
                        r.resident = res
                        r.save()
                except:
                    pass
                try:
                    cat = request.POST.get("category")
                    if cat:
                        cat = category.objects.get(id=int(cat))
                        r.category = cat
                        r.save()
                except:
                    pass
                try:
                    date_a = request.POST.get("date_a")
                    if date_a:
                        if date_a == r.date_accept:
                            pass
                        else:
                            r.date_accept = date_a
                            r.save()
                    else:
                        if date_a == r.date_accept:
                            pass
                        else:
                            r.date_accept = None
                            r.save()
                except:
                    pass
                try:
                    method = request.POST.get("sum_method")
                    if method:
                        r.sum_method = 1
                        r.save()
                    else:
                        r.sum_method = 0
                        r.save()
                except:
                    r.sum_method = 0
                    r.save()
                try:
                    exception = request.POST.get("exception")
                    if exception:
                        r.exception = '1'
                        r.save()
                    else:
                        r.exception = '0'
                        r.save()
                except:
                    r.exception = '0'
                    r.save()
                try:
                    date_l = request.POST.get("date_l")
                    if date_l:
                        if date_l == r.date_leave:
                            pass
                        else:
                            r.date_leave = date_l
                            r.save()
                    else:
                        if date_l == r.date_leave:
                            pass
                        else:
                            r.date_leave = None
                            r.save()
                except:
                    pass
                try:
                    company = request.POST.get("company")
                    if company == str(r.company):
                        pass
                    else:
                        r.company = company
                        r.save()
                except:
                    pass
                try:
                    snils = request.POST.get('snils')
                    if snils == str(r.snils):
                        pass
                    else:
                        r.snils = snils
                        r.save()
                except:
                    pass
                try:
                    inn = request.POST.get('inn')
                    if snils == str(r.inn):
                        pass
                    else:
                        r.inn = inn
                        r.save()
                except:
                    pass
                try:
                    full_name = request.POST.get('full_name')
                    if full_name == str(r.full_name):
                        pass
                    else:
                        r.full_name = full_name
                        r.save()
                except:
                    pass
                try:
                    phone_number = request.POST.get('phone_number')
                    r.phone_number = phone_number
                    r.save()
                except:
                    pass
                try:
                    position = request.POST.get('position')
                    if position == str(r.position):
                        pass
                    else:
                        r.position = position
                        r.save()
                except:
                    pass
                try:
                    workgroup = request.POST.get('workgroup')
                    if workgroup == str(r.workgroup):
                        pass
                    else:
                        r.workgroup = workgroup
                        r.save()
                except:
                    pass
                try:
                    g = request.POST.get('group')
                    if r.group:
                        if int(g) == r.group.id:
                            pass
                        else:
                            g = int(g)
                            r.group = group.objects.get(id=g)
                            r.save()
                    else:
                        g = int(g)
                        r.group = group.objects.get(id=g)
                        r.save()
                except:
                    pass
                try:
                    now = datetime.date(year=datetime.date.today().year, month=datetime.date.today().month, day=1)
                    ps = persone_salary.objects.all().filter(persone=r, date=now).last()
                except:
                    pass
                try:
                    grade = request.POST.get('grade')
                    if grade == str(r.grade):
                        pass
                    else:
                        r.grade = grade
                        r.save()
                        ps.grade = r.grade
                        ps.save()
                except:
                    pass
                try:
                    salary = request.POST.get('salary')
                    if salary == str(r.salary):
                        pass
                    else:
                        try:
                            salary = int(salary)
                        except:
                            salary = salary.rpartition(',')[0]
                            salary = int(salary)
                        r.salary = float(salary)
                        r.save()
                        ps.salary = r.salary
                        ps.save()
                except:
                    pass
                try:
                    graphic = request.POST.get('graphic')
                    if graphic == str(r.graphic):
                        pass
                    else:
                        r.graphic = int(graphic)
                        r.save()
                except:
                    pass
                try:
                    output = request.POST.get('output')
                    if output == str(r.output):
                        pass
                    else:
                        r.output = output
                        r.save()
                except:
                    pass
                try:
                    driver = request.POST.get('driver')
                    if driver == str(r.driver):
                        pass
                    else:
                        r.driver = int(driver)
                        r.save()
                except:
                    pass
                try:
                    leaved = request.POST.get('leaved')
                    if leaved == str(r.leaved):
                        pass
                    else:
                        r.leaved = int(leaved)
                        r.save()
                    work_hours = request.POST.get('work_hours')
                    if work_hours == str(r.work_hours):
                        pass
                    else:
                        r.work_hours = work_hours
                        r.save()
                except:
                    pass
            return redirect('/view_persone/'+str(id))
        else:
            context = {'r': r, 'groups': groups, 'segment': 'ref', 'persones': persones, 'cats': cats, 'tax': tax, }
            html_template = loader.get_template('reference/persone/view_persone.html')
            return HttpResponse(html_template.render(context, request))

class task:
    @login_required(login_url="/login/")
    def task_main(request):
        tasks = task_head.objects.all()
        context = {'segment': 'task', 'tasks': tasks}
        html_template = loader.get_template('task/task_main.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_task(request):
        tb = task_body.objects.all()
        obj = objs.objects.all()
        forms = TaskForm
        if request.method == 'POST':
            forms = TaskForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.date_create = datetime.datetime.now()
                post.creator = request.user
                n = int(request.POST.get('tbn'))
                i = 0
                post.obj = objs.objects.get(id=int(request.POST.get('obj')))
                try:
                    number = task_number.objects.get(obj=post.obj)
                except:
                    task_number(
                        obj=objs.objects.get(id=int(request.POST.get('obj'))),
                        number=1
                    ).save()
                    number = task_number.objects.all().filter(obj=post.obj).last()
                post.id_name = str(post.obj.name) + '_' + str(number.number)
                number.number += 1
                number.save()
                post.save()
                if n == 0:
                    tn = request.POST.get('tn' + str(i))
                    if tn == '':
                        tn = 'Не заполнено'
                    tc = request.POST.get('tc' + str(i))
                    if tc == '':
                        tc = 'Не заполнено'
                    te = request.POST.get('te' + str(i))
                    if te == '':
                        te = 'Не заполнено'
                    td = request.POST.get('td' + str(i))
                    if td == '':
                        td = datetime.date.today() + datetime.timedelta(days=5)
                    task_body(
                        head=post,
                        name=tn,
                        number=tc,
                        ci=te,
                        need_date=td
                    ).save()
                    post.body = 1
                else:
                    n = n-1
                    while i <= n:
                        tn = request.POST.get('tn' + str(i))
                        if tn == '':
                            tn = 'Не заполнено'
                        tc = request.POST.get('tc' + str(i))
                        if tc == '':
                            tc = 'Не заполнено'
                        te = request.POST.get('te' + str(i))
                        if te == '':
                            te = 'Не заполнено'
                        td = request.POST.get('td' + str(i))
                        if td == '':
                            td = datetime.date.today()+datetime.timedelta(days=5)
                        task_body(
                            head=post,
                            name=tn,
                            number=tc,
                            ci=te,
                            need_date=td
                        ).save()
                        i=i+1
                        post.body = i
                post.save()
                return redirect('task_main')
            else:
                error = 'Некорректные данные'
                context = {'segment': 'task', 'forms': forms,'error': error}
                html_template = loader.get_template('task/new_task.html')
                return HttpResponse(html_template.render(context, request))
        context = {'segment': 'task', 'forms':forms, 'obj':obj}
        html_template = loader.get_template('task/new_task.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_task(request, id):
        th = task_head.objects.get(id=id)
        tb = task_body.objects.all().filter(head=th)
        if request.method == 'POST':
            try:
                stat = request.POST.get('status')
                if stat:
                    if th.status != stat:
                        th.status = stat
                        th.status_changer = request.user
                    th.save()
            except:
                pass
            s = th.status
            comment = request.POST.get('comm')
            if comment != '':
                if th.comment == None:
                    th.comment = ' '
                    th.save()
                comment = str(th.comment) + str(comment)+ ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M"))+' '+ '\r\n'
                th.comment = comment
                th.status = s
                th.save()
            try:
                ph = 0
                i = 0
                for t in tb:
                    tch = request.POST.get('tch' + str(t.id))
                    tac = request.POST.get('tac' + str(t.id))
                    tag = request.POST.get('tagreem'+str(t.id))
                    if tag:
                        tag = int(tag)
                        if tag == 1:
                            t.agreem = 2
                            t.agreemer = None
                            comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' изменил статус одобрения запроса "' + str(t.name) + '" - "На рассмотрение"' + '\r\n'
                            th.comment = comment
                            th.save()
                        if tag == 2:
                            t.agreem = 1
                            t.agreemer = request.user
                            t.save()
                            comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' изменил статус одобрения запроса "' + str(t.name) + '" - "Одобрено"' + '\r\n'
                            th.comment = comment
                            th.save()
                    else:
                        if t.agreem == 1:
                            t.agreem = 2
                            t.agreemer = None
                            comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' изменил статус одобрения запроса "' + str(t.name) + '" - "На рассмотрение"' + '\r\n'
                            th.comment = comment
                            th.save()
                    if tch:
                        t.change = int(tch)
                    else:
                        t.change = None
                    try:
                        if tac:
                            tac = int(tac)
                            if tac == 1:
                                t.accept = 2
                                comment = str(th.comment) + str(
                                    datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' изменил статус приемки "' + str(
                                    t.name) + '" - "В ожидании"' + '\r\n'
                                th.comment = comment
                                th.save()
                            if tac == 2:
                                t.accept = 1
                                t.save()
                                comment = str(th.comment) + str(
                                    datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' изменил статус приемки "' + str(
                                    t.name) + '" - "Получено"' + '\r\n'
                                th.comment = comment
                                th.save()
                        else:
                            if t.accept == 1:
                                t.accept = 2
                                comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' изменил статус приемки "' + str(
                                    t.name) + '" - "В ожидании"' + '\r\n'
                                th.comment = comment
                                th.save()
                    except:
                        pass
                    try:
                        if t.procent <= 99:
                            TestDay = datetime.date.today()
                            if TestDay >= t.need_date:
                                t.expired = 1
                                th.expired = 1
                                t.save()
                    except:
                        pass
                    t.save()
                    ph = ph + t.procent
                    i +=1
                endph = 0
                for t in tb:
                    if t.procent <= 100:
                        endph=endph + t.procent
                    else:
                        endph=endph + 1000
                th.procent = endph/int(th.body)
                th.save()
            except:
                pass
            return redirect('/view_task/'+str(id))
        else:
            context = {'segment': 'task', 'tb':tb, 'th': th}
            html_template = loader.get_template('task/view_task.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def edit_task(request, id):
        th = task_head.objects.get(id=id)
        tb = task_body.objects.all().filter(head=th).order_by('id')
        context = {'segment': 'task', 'tb': tb, 'th': th}
        if request.method == 'POST':
            n = int(request.POST.get('tbn'))
            i = 0
            for t in tb:
                name = request.POST.get('name'+str(t.id))
                number = request.POST.get('number'+str(t.id))
                ci = request.POST.get('ci'+str(t.id))
                need_date = request.POST.get('need_date'+str(t.id))
                t.name = name
                t.number = number
                t.ci = ci
                t.need_date = need_date
                t.save()
            if int(n)!=0:
                n = n - 1
                while i <= n:
                    tn = request.POST.get('tn' + str(i))
                    if tn == '':
                        tn = 'Не заполнено'
                    tc = request.POST.get('tc' + str(i))
                    if tc == '':
                        tc = 'Не заполнено'
                    te = request.POST.get('te' + str(i))
                    if te == '':
                        te = 'Не заполнено'
                    td = request.POST.get('td' + str(i))
                    if td == '':
                        td = datetime.date.today() + datetime.timedelta(days=5)
                    task_body(
                        head=th,
                        name=tn,
                        number=tc,
                        ci=te,
                        need_date=td
                    ).save()
                    i+=1
            name = request.POST.get('name')
            if name != th.type:
                if name:
                    th.type = name
            th.body = th.body + i
            th.save()
            return redirect('/edit_task/'+str(id))
        html_template = loader.get_template('task/edit_task.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def del_tb(request, id,ids):
        tb = task_body.objects.get(id=id)
        tb.delete()
        response = redirect('/view_task/'+str(ids))
        return response


    @login_required(login_url="/login/")
    def del_task(request, id):
        th = task_head.objects.get(id=id)
        tb = task_body.objects.all().filter(head=th)
        for t in tb:
            t.delete()
        th.delete()
        response = redirect('task_main')
        return response

    @login_required(login_url="/login/")
    def close_task(request, id):
        th = task_head.objects.get(id=id)
        th.closed = 1
        th.system += ' Заявка закрыта ' + str() + '- ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + '. ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
        th.save()
        response = redirect('task_main')
        return response

    @login_required(login_url="/login/")
    def open_task(request, id):
        th = task_head.objects.get(id=id)
        th.closed = None
        th.system += ' Заявка открыта ' + str() + '- ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + '. ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
        th.save()
        response = redirect('task_main')
        return response

    @login_required(login_url="/login/")
    def accept_save(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            value = request.POST.get('value')
            taskb = task_body.objects.get(id=int(id))
            taskh = task_head.objects.get(id=taskb.head.id)
            if value:
                taskb.accept = int(value)
                if value == '2':
                    stage = 'Закрыто'
                elif value == '1':
                    stage = 'Мой склад'
                elif value == '0':
                    stage = 'В процессе'
                if taskh.system is None:
                    taskh.system = ''
                taskh.system +=' Стадия заявки - ' + str(stage) + '- ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + '. ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) +  '\r\n'
                taskh.save()
                taskb.save()
                return JsonResponse({'message': 'Данные успешно сохранены'})
            elif value == '':
                return JsonResponse({'error': 'Null Недопустимое значение для поля Value'}, status=400)
            else:
                return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    @login_required(login_url="/login/")
    def change_save(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            value = request.POST.get('value')
            taskb = task_body.objects.get(id=int(id))
            taskh = task_head.objects.get(id=taskb.head.id)
            if value:
                taskb.change = int(value)
                taskb.status_changer = request.user
                taskb.save()
                taskh.system += str(taskh.system) +'Изменение кол-ва - ' + str(request.user.last_name)+' '+ str(request.user.first_name) + '. ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) +   '\r\n'
                taskh.save()
                user = ''
                if taskb.status_changer.last_name:
                    user += taskb.status_changer.last_name
                    if taskb.status_changer.first_name:
                        user += ' '
                        user += taskb.status_changer.first_name
                elif taskb.status_changer.first_name:
                        user += taskb.status_changer.first_name
                else:
                    user +=  taskb.status_changer
                return JsonResponse({'message': 'Данные успешно сохранены', 'user': user})
            elif value == '':
                return JsonResponse({'error': 'Null Недопустимое значение для поля Value'}, status=400)
            else:
                return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    def agreem_save(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            value = request.POST.get('value')
            taskh = task_head.objects.get(id=int(id))
            if value:
                taskh.status = value
                if taskh.system:
                    taskh.system += 'Статус изменен на - '+ str(value)+ ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                else:
                    taskh.system = ''
                    taskh.system += 'Статус изменен на - '+ str(value)+ ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                taskh.status_changer = request.user
                taskh.save()
                return JsonResponse({'message': 'Данные успешно сохранены'})
            elif value == '':
                return JsonResponse({'error': 'Null Недопустимое значение для поля Value'}, status=400)
            else:
                return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    def control_save(request):
        if request.method == 'POST':
            id = request.POST.get('id')
            value = request.POST.get('value')
            taskh = task_head.objects.get(id=int(id))
            if value:
                taskh.control = value
                if taskh.system:
                    taskh.system += 'Контроль изменен на - '+ str(value)+ ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                else:
                    taskh.system = ''
                    taskh.system += 'Контроль изменен на - '+ str(value)+ ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                taskh.control_changer = request.user
                taskh.save()
                return JsonResponse({'message': 'Данные успешно сохранены'})
            elif value == '':
                return JsonResponse({'error': 'Null Недопустимое значение для поля Value'}, status=400)
            else:
                return JsonResponse({'error': 'Недопустимое значение для поля Value'}, status=400)
        else:
            return JsonResponse({'error': 'Метод не поддерживается'}, status=405)
    @login_required(login_url="/login/")
    def taskexport(request, obj, name):
        obj_s = objs.objects.get(name=obj)
        th = task_head.objects.all().filter(obj=obj_s)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('taskexp.xlsx')
        t_list: worksheet = book.worksheets[0]
        i = 3
        fill = PatternFill(start_color="bc5d58", end_color="bc5d58", fill_type="solid")
        for p in th:
            n = 1
            tbs = task_body.objects.all().filter(head=p).order_by('name')
            t_list.insert_rows(i)
            t_list["A" + str(i)].fill = fill
            t_list["B" + str(i)].fill = fill
            t_list["C" + str(i)].fill = fill
            t_list["D" + str(i)].fill = fill
            t_list["A" + str(i)].value = "" + str(p.type)
            t_list["B" + str(i)].value = "" + str(p.creator.last_name)+" "+str(p.creator.first_name)
            t_list["C" + str(i)].value = "" + str(p.date_create)
            t_list["D" + str(i)].value = "" + str(p.status)
            i+=1
            for t in tbs:
                t_list.insert_rows(i)
                t_list["A" + str(i)].value = "" + str(n)
                t_list["B" + str(i)].value = "" + str(t.name)
                t_list["C" + str(i)].value = "" + str(t.number)
                t_list["D" + str(i)].value = "" + str(t.ci)
                i += 1
                n+=1
        book.title = 'Отчет общий'
        book.save(response)
        return response
        return redirect('/task_main/')










class Agreement:
    @login_required(login_url="/login/")
    def agreement_main(request):
        agreements = agreement.objects.all()
        context = {'segment': 'agreement', 'agreements': agreements}
        html_template = loader.get_template('agreement/agreement_main.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_agreement(request):
        if request.method == 'POST':
            form = AgreementForm(request.POST, request.FILES)
            if form.is_valid():
                post = form.save(commit=False)
                post.date_create = datetime.datetime.now()
                post.name = request.POST.get('name')
                post.creator = request.user
                post.status = 'На проверке'
                post.save()
                agreem = agreement.objects.get(id=post.id)
                agreement_step(
                    step='Проверка',
                    agreem=agreem,
                    accepted=0
                ).save()
                accepters = accepter_list.objects.all()
                for accepter in accepters:
                    accept_user = User.objects.get(id=accepter.user.id)
                    step = agreement_step.objects.all().last()
                    agreement_accepter(accepting=step,user=accept_user).save()
                return HttpResponseRedirect(reverse('agreements'))
        else:
            forms = AgreementForm
            agreem = agreement.objects.all()
            context = {'segment': 'agreement', 'agreem': agreem,'forms': forms}
            html_template = loader.get_template('agreement/new_agreement.html')
            return HttpResponse(html_template.render(context, request))


    @login_required(login_url="/login/")
    def view_agreement(request, id):
        agreem = agreement.objects.get(id=id)
        if request.method == 'POST':
            agreem.comment = str(agreem.comment)+'\r\n'+ str(request.POST.get('com'))+': '+ str(request.user.last_name)+' '+ str(request.user.first_name) +' '+ str(datetime.datetime.now().strftime("%d.%m.%y %H:%M"))
            agreem.save()
        forms = TempFileForm
        steps = agreement_step.objects.all().filter(agreem=agreem)
        accepters = agreement_accepter.objects.all()
        context = {'segment': 'agreement', 'agreem': agreem,'steps':steps,'accepters':accepters,'forms':forms}
        html_template = loader.get_template('agreement/view_agreement.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def correct_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        creator = User.objects.get(id=agreem.creator.id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted += 1
        close.save()
        agreement_step(
            step = 'Корректировка',
            agreem = agreem,
            accepted = 0,
            need = 1
        ).save()
        step = agreement_step.objects.all().last()
        agreement_accepter(
            accepting=step,
            user=creator,
        ).save()
        agreem.status = 'На корректировке'
        agreem.save()
        return redirect('/view_agreement/'+str(id))

    @login_required(login_url="/login/")
    def accept_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        creator = User.objects.get(id=agreem.creator.id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted+=1
        close.save()
        if close.accepted == close.need:
            n = 0
            list = agreementer_list.objects.all()
            for a in list:
                n+=1
            agreement_step(
                step='Согласование',
                agreem=agreem,
                accepted=0,
                need = n
            ).save()
            step = agreement_step.objects.all().last()
            for a in list:
                agreement_accepter(
                    accepting=step,
                    user=creator,
                ).save()
            agreem.status = 'На согласование'
            agreem.save()
        return redirect('/view_agreement/' + str(id))

    @login_required(login_url="/login/")
    def agreemented(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        creator = User.objects.get(id=agreem.creator.id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted += 1

        close.save()
        if close.accepted == close.need:
            n = 0
            list = ender_list.objects.all()
            for a in list:
                n += 1
            agreement_step(
                step='Завершение',
                agreem=agreem,
                accepted=0,
                need=n
            ).save()
            step = agreement_step.objects.all().last()
            for a in list:
                agreement_accepter(
                    accepting=step,
                    user=creator,
                ).save()
            agreem.status = 'Завершен'
            agreem.save()
        return redirect('/view_agreement/' + str(id))

    @login_required(login_url="/login/")
    def close_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted += 1
        close.save()
        agreem.closed = 1
        agreem.closer = request.user
        agreem.save()
        agreem.status = 'Завершен'
        agreem.save()
        return redirect('/view_agreement/' + str(id))



    @login_required(login_url="/login/")
    def new_file_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        if request.method == 'POST':
            form = TempFileForm(request.POST, request.FILES)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                agreem.file = post.temp
                agreem.status = 'Проверка'
                agreem.save()
                accept = agreement_accepter.objects.get(id=ida)
                accept.accept = 1
                accept.save()
                close = agreement_step.objects.get(id=ids)
                close.accepted += 1
                close.save()
                n=0
                accepters = accepter_list.objects.all()
                for accepter in accepters:
                    n+=1
                agreement_step(
                    step='Проверка',
                    agreem=agreem,
                    accepted=0,
                    need=n
                ).save()
                for accepter in accepters:
                    accept_user = User.objects.get(id=accepter.user.id)
                    step = agreement_step.objects.all().last()
                    agreement_accepter(accepting=step, user=accept_user).save()
                return redirect('/view_agreement/' + str(id))
        form = TempFileForm
        context = {'segment': 'agreement', 'form': form}
        html_template = loader.get_template('agreement/corrector_agreement.html')
        return HttpResponse(html_template.render(context, request))





 #   close_agreement
    @login_required(login_url="/login/")
    def agreement_file(request, id):
        document = get_object_or_404(agreement, id=id)
        response = HttpResponse(document.file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{document.file.name}"'
        return response

class misc:

    @login_required(login_url="/login/")
    def month_info_half(request, years, month, name):
        date = datetime.date(years, month, 1)
        pers = period.objects.all().filter(date=date)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('mix.xlsx')
        t_list: worksheet = book.worksheets[0]
        i = 8
        n = 1
        for p in pers:
            tls = tabel_list.objects.all().filter(period=p)
            m = int(p.date.month) - 1
            month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября',
                     'Ноября', 'Декабря']
            my = str(month[m]) + ' ' + str(p.date.strftime("%Y"))
            for t in tls:
                if t.persone.company == 1:
                    org = 'ПМК'
                else:
                    org = 'неоформ'
                t_list.insert_rows(i)
                t_list["A" + str(i)].value = "" + str(n)
                t_list["B" + str(i)].value = "" + str(org)
                t_list["C" + str(i)].value = "" + str(t.persone.snils)
                t_list["D" + str(i)].value = "" + str(t.full_name)
                t_list["E" + str(i)].value = "" + str(t.position)
                t_list["F" + str(i)].value = "" + str(t.period.obj)
                t_list["G" + str(i)].value = "" + str(t.group)
                if t.persone.grade and t.persone.grade != '0':
                    if t.persone.sum_method == 1:
                        t_list["H" + str(i)].value = "" + str(round(t.persone.salary))
                    else:
                        if str(t.persone.grade) == '1':
                            t_list["H" + str(i)].value = "60000/1"
                        elif str(t.persone.grade) == '2':
                            t_list["H" + str(i)].value = "80000/2"
                        elif str(t.persone.grade) == '3':
                            t_list["H" + str(i)].value = "87000/3"
                        elif str(t.persone.grade) == '4':
                            t_list["H" + str(i)].value = "92000/4"
                        elif str(t.persone.grade) == '5':
                            t_list["H" + str(i)].value = "100000/5"
                        elif str(t.persone.grade) == '6':
                            t_list["H" + str(i)].value = "106000/6"
                else:
                    t_list["H" + str(i)].value = "" + str(round(t.persone.salary))
                if t.d1: t_list["J" + str(i)].value = "" + str(t.d1)
                if t.d2: t_list["K" + str(i)].value = "" + str(t.d2)
                if t.d3: t_list["L" + str(i)].value = "" + str(t.d3)
                if t.d4: t_list["M" + str(i)].value = "" + str(t.d4)
                if t.d5: t_list["N" + str(i)].value = "" + str(t.d5)
                if t.d6: t_list["O" + str(i)].value = "" + str(t.d6)
                if t.d7: t_list["P" + str(i)].value = "" + str(t.d7)
                if t.d8: t_list["Q" + str(i)].value = "" + str(t.d8)
                if t.d9: t_list["R" + str(i)].value = "" + str(t.d9)
                if t.d10: t_list["S" + str(i)].value = "" + str(t.d10)
                if t.d11: t_list["T" + str(i)].value = "" + str(t.d11)
                if t.d12: t_list["U" + str(i)].value = "" + str(t.d12)
                if t.d13: t_list["V" + str(i)].value = "" + str(t.d13)
                if t.d14: t_list["W" + str(i)].value = "" + str(t.d14)
                if t.d15: t_list["X" + str(i)].value = "" + str(t.d15)
                if t.days: t_list["AO" + str(i)].value = "" + str(t.days)
                if t.add: t_list["AP" + str(i)].value = "" + str(round(t.add))
                if t.fine: t_list["AQ" + str(i)].value = "" + str(round(t.fine))
                if t.sum: t_list["AR" + str(i)].value = "" + str(round(t.sum))
                if t.comment: t_list["AS" + str(i)].value = "" + str(t.comment)
                if request.user.username == 'KaratovaMadina' or request.user.username == 'DjavathanovaP' or request.user.is_superuser:
                    t_list["AT6"].value = "Заметки"
                    if t.persone.idea: t_list["AT" + str(i)].value = str(t.persone.idea)
                i += 1
                n += 1
        book.title = 'Отчет общий'
        book.save(response)
        return response
        return redirect('/tabels/')

    @login_required(login_url="/login/")
    def view_tabel_half(request, id):
        groups = group.objects.all()
        per = period.objects.get(id=id)
        pers = period.objects.all()
        user = request.user
        approvers = dayapprove.objects.all().filter(per=per)
        persones = persone.objects.all().filter().order_by('full_name')
        tabs = tabel_list.objects.all().filter(period=per).order_by('full_name')
        html_template = loader.get_template('tabels/worker/view_tabel_half.html')
        today = datetime.date.today()
        next_month = today.replace(day=1) + datetime.timedelta(days=16)
        if per.date.month == today.month or next_month.month == per.date.month or request.user.username == 'KaratovaMadina' or request.user.username == 'DjavathanovaP':
            for t in tabs:
                if t.company:
                    if t.company != t.persone.company:
                        t.company = t.persone.company
                        t.save()
                if t.snils:
                    if t.snils != t.persone.snils:
                        t.snils = t.persone.snils
                        t.save()
                if t.full_name:
                    if t.full_name != t.persone.full_name:
                        t.full_name = t.persone.full_name
                        t.save()
                if t.graphic:
                    if t.graphic != t.persone.graphic:
                        t.graphic = t.persone.graphic
                        t.save()
                if t.group:
                    if t.group != t.persone.group:
                        t.group = t.persone.group
                        t.save()
                if t.position:
                    if t.position != t.persone.position:
                        t.position = t.persone.position
                        t.save()
        if per.date:
            month = datetime.date.today().month
            pm = per.date.month
            if month > pm:
                if datetime.date.today().year == per.date.year:
                    per.fullblock = 1
                    per.save()
            if per.fullblock != 1:
                today = datetime.date.today().day - 1
                for day in range(1, 16):
                    if today > day:
                        setattr(per, f'db{day}', 1)
                        per.save()
            if per:
                for day in range(1, 16):
                    i = 0
                    for a in approvers:
                        if a.day == day:
                            i += 1
                        if i >= 1:
                            setattr(per, f'da{day}', 1)
                            per.save()
        if request.method == 'POST':
            for t in tabs:
                out = 0
                for i in range(1, 16):
                    day_key = f'd{i}'
                    da = request.POST.get(str(t.id) + day_key)
                    if da:
                        try:
                            da = int(da)
                            if 0 < da <= 23:
                                setattr(t, day_key, da)
                        except ValueError:
                            da = da.upper()
                            if da in ["ВВ", "Н", "М", "Б", "О", "Д", "В", "У"]:
                                setattr(t, day_key, da)
                        t.save()
                h = 0
                d = 0
                for i in range(1, 16):
                    day_key = f'd{i}'
                    try:
                        h_day = int(getattr(t, day_key))
                        h += h_day
                        d += 1
                    except:
                        if getattr(t, day_key) == "ВВ":
                            out += 1
                            if out < 3:
                                h += int(t.persone.work_hours)
                                d += 1
                        elif getattr(t, day_key) == "М":
                            if str(t.group.name) == "ИТР" or str(t.group.name) == "ПТО":
                                h += int(t.persone.work_hours)
                                d += 1
                t.hours = int(h)
                t.days = int(d)
                t.save()
                comment = request.POST.get("com"+str(t.id))
                if comment == '':
                    pass
                else:
                    if t.comment == None:
                        t.comment = ''
                        t.save()
                        comment = str(t.comment) + str(comment)  + ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                    else:
                        comment = str(t.comment) + str(comment)  + ': ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' ' + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + '\r\n'
                        t.comment = comment
                        t.save()
                try:
                    fine = request.POST.get(str(t.id)+"fine")
                    if fine:
                        t.fine = float(fine)
                    else:
                        t.fine = 0
                except:
                    pass
                try:
                    add = request.POST.get(str(t.id)+"add")
                    if add:
                        t.add = float(add)
                    else:
                        t.add = 0
                except:
                    pass
                f = t.fine
                a = t.add
                today = datetime.date.today()
                sal = 0
                next_month = today.replace(day=1) + datetime.timedelta(days=16)
                if per.date.year == today.year and per.date.month == today.month or next_month.month == per.date.month:
                    if t.persone.sum_method == 1:
                        sal = int(t.persone.salary)
                    #else:
                    #    if str(t.persone.grade) == '2':
                    #        sal = 60900
                    #    elif str(t.persone.grade) == '3':
                    #        sal = 67800
                    #    elif str(t.persone.grade) == '4':
                    #        sal = 72700
                    #    elif str(t.persone.grade) == '5':
                    #        sal = 80000
                    #    elif str(t.persone.grade) == '6':
                    #        sal = 86250
                    #    elif t.persone.salary:
                    #        sal = int(t.persone.salary)
                    #    else:
                    #        sal = 0
                    else:
                        now = datetime.date(day=1,month=2,year=2025)
                        if per.date < now:
                             if str(t.persone.grade) == '2':
                                 sal = 60900
                             elif str(t.persone.grade) == '3':
                                 sal = 67800
                             elif str(t.persone.grade) == '4':
                                 sal = 72700
                             elif str(t.persone.grade) == '5':
                                 sal = 80000
                             elif str(t.persone.grade) == '6':
                                 sal = 86250
                             elif t.persone.salary:
                                 sal = int(t.persone.salary)
                             else:
                                 sal = 0
                        else:
                            if str(t.persone.grade) == '2':
                                sal = 80000
                            elif str(t.persone.grade) == '3':
                                sal = 87000
                            elif str(t.persone.grade) == '4':
                                sal = 92000
                            elif str(t.persone.grade) == '5':
                                sal = 100000
                            elif str(t.persone.grade) == '6':
                                sal = 106000
                            elif str(t.persone.grade) == '1':
                                sal = 60000
                            elif t.persone.salary:
                                sal = int(t.persone.salary)
                            else:
                                sal = 0
                    #if per.graphic == 1:
                    #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                    #else:
                    #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                else:
                    ps = persone_salary.objects.all().filter(persone=t.persone,date=per.date).order_by('-id').last()
                    if t.persone.sum_method == 1:
                        sal = int(ps.salary)
                    #else:
                    #    if ps.grade == '2':
                    #        sal = 60900
                    #    elif ps.grade == '3':
                    #        sal = 67800
                    #    elif ps.grade == '4':
                    #        sal = 72700
                    #    elif ps.grade == '5':
                    #        sal = 80000
                    #    elif ps.grade == '6':
                    #        sal = 86250
                    #    elif ps.salary:
                    #        sal = int(ps.salary)
                    #    else:
                    #        sal = 0
                    else:
                        now = datetime.date(day=1,month=2,year=2025)
                        if per.date < now:
                            if ps.grade == '2':
                                sal = 60900
                            elif ps.grade == '3':
                                sal = 67800
                            elif ps.grade == '4':
                                sal = 72700
                            elif ps.grade == '5':
                                sal = 80000
                            elif ps.grade == '6':
                                sal = 86250
                            elif ps.salary:
                                sal = int(ps.salary)
                            else:
                                sal = 0
                        else:
                            try:
                                if ps.grade == '2':
                                    sal = 80000
                                elif ps.grade == '3':
                                    sal = 87000
                                elif ps.grade == '4':
                                    sal = 92000
                                elif ps.grade == '5':
                                    sal = 100000
                                elif ps.grade == '6':
                                    sal = 106000
                                elif ps.grade == '1':
                                    sal = 60000
                                elif ps.salary:
                                    sal = int(ps.salary)
                                else:
                                    sal = 0
                            except:
                                sal = 0
                if t.graphic == 2:
                    t.sum = sal / int(per.workdays5) * t.days
                else:
                    t.sum = sal / int(per.workdays) * t.days
                if t.persone.exception == 0:
                    t.salary = t.sum - float(f) + float(a)
                else:
                    t.salary = 0
                t.save()
            wd = request.POST.get('wd');
            if int(wd) != per.workdays:
                per.workdays = int(wd)
                per.save()
            wd5 = request.POST.get('wd5');
            if int(wd5) != per.workdays5:
                per.workdays5 = int(wd5)
                per.save()
            return redirect('/view_tabel_half/' + str(id))
        else:
            context = {
                'user': user,
                'approvers':approvers,
                'persones': persones,
                'groups': groups,
                'per': per,
                'segment': 'tabel',
                'tabs': tabs,
                'pers': pers,
            }
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def converts(request):
        if request.method == 'POST':
            pers = persone.objects.all()
            feb = request.POST.get('convert_date')
            feb = datetime.datetime.strptime(feb, '%Y-%m-%d').date()
            for p in pers:
                persone_salary(
                    persone=p,
                    date=feb,
                    salary=p.salary,
                    grade=p.grade,
                ).save()
                try:
                    tax = driver_tax.objects.get(driver=p)
                    if tax:
                        driver_tax_lagacy(
                            driver=tax.driver,
                            tax=tax.tax,
                            date=feb
                        ).save()
                except:
                    pass
        else:
            context = {
                'segment': 'profile',
            }
            html_template = loader.get_template('home/save.html')
            return HttpResponse(html_template.render(context, request))
        return redirect('/persones')

    def notification_add(requst):
        user_list = User.objects.all()
        for user_name in user_list:
            try:
                notificat = notificated.objects.get(user=user_name)
                notificat.notificated = 1
                notificat.save()
            except:
                notificated(
                    user=user_name,
                    notificated=1
                ).save()

        return redirect('/')

    def convert6(requst):
        pers = persone.objects.all()
        feb = datetime.date(2026,3,1)
        for p in pers:
            persone_salary(
                persone=p,
                date=feb,
                salary=p.salary,
                grade=p.grade,
            ).save()
            try:
                tax = driver_tax.objects.get(driver=p)
                if tax:
                    driver_tax_lagacy(
                        driver=tax.driver,
                        tax=tax.tax,
                        date=feb
                    ).save()
            except:
                pass
        return redirect('/persones')


    def is_last_day_of_month(date):
        last_day = calendar.monthrange(date.year, date.month)
        return date.day == last_day

    def convert5(request):
        tbl = driver.objects.all()
        for t in tbl:
            if str(t.persone.company) == '1' or str(t.persone.company) == 'ПМК':
                t.company = 1
            else:
                t.company = 2
            t.save()
        return redirect('/persones')

    def convert8(request):
        persones = persone.objects.all()
        for t in persones:
            if t.group == None:
                g = group.objects.get(name='None')
                t.group = g
                t.save()
        return redirect('/persones/')

    @login_required(login_url="/login/")
    def persones_info(request, name):
        persones = persone.objects.all()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('persones.xlsx')
        t_list: worksheet = book.worksheets[0]
        i = 5
        for p in persones:
            if p.company == 1:
                org = 'ПМК'
            else:
                org = 'неоформ'
            t_list.insert_rows(i)
            t_list["A" + str(i)].value = "" + str(p.id)
            t_list["B" + str(i)].value = "" + str(org)
            if p.snils:
                t_list["C" + str(i)].value = "" + str(p.snils)
            if p.inn:
                t_list["D" + str(i)].value = "" + str(p.inn)
            t_list["E" + str(i)].value = "" + str(p.full_name)
            t_list["F" + str(i)].value = "" + str(p.category)
            if p.group:
                t_list["G" + str(i)].value = "" + str(p.group)
            if p.workgroup:
                t_list["H" + str(i)].value = "" + str(p.workgroup)
            if p.position:
                t_list["I" + str(i)].value = "" + str(p.position)
            if p.salary:
                t_list["J" + str(i)].value = "" + str(p.salary)
            if p.grade:
                t_list["K" + str(i)].value = "" + str(p.grade)
            if p.graphic:
                t_list["L" + str(i)].value = "" + str(p.graphic)
            if p.date_accept:
                t_list["M" + str(i)].value = "" + str(p.date_accept)
            if p.date_leave:
                t_list["N" + str(i)].value = "" + str(p.date_leave)
            if p.phone_number:
                t_list["O" + str(i)].value = "" + str(p.phone_number)
            if p.idea:
                t_list["P" + str(i)].value = "" + str(p.idea)
            if p.comment:
                t_list["Q" + str(i)].value = "" + str(p.comment)
            i += 1
        book.title = 'Отчет по персоналу'
        book.save(response)
        return response
        return redirect('/persones/')


    @login_required(login_url="/login/")
    def month_info(request, years, month, name):
        date = datetime.date(years,month,1)
        pers = period.objects.all().filter(date=date)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('mix.xlsx')
        t_list: worksheet = book.worksheets[0]
        i=8
        n=1
        for p in pers:
            tls = tabel_list.objects.all().filter(period=p)
            m = int(p.date.month)-1
            month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября',
                     'Ноября', 'Декабря']
            my = str(month[m]) + ' ' + str(p.date.strftime("%Y"))
            for t in tls:
                if t.persone.company == 1:
                    org = 'ПМК'
                else:
                    org = 'неоформ'
                t_list.insert_rows(i)
                t_list["A" + str(i)].value = "" + str(n)
                t_list["B" + str(i)].value = "" + str(org)
                t_list["C" + str(i)].value = "" + str(t.persone.snils)
                t_list["D" + str(i)].value = "" + str(t.full_name)
                t_list["E" + str(i)].value = "" + str(t.position)
                t_list["F" + str(i)].value = "" + str(t.period.obj)
                t_list["G" + str(i)].value = "" + str(t.group)
                if t.persone.grade and t.persone.grade != '0':
                    if t.persone.sum_method == 1:
                        t_list["H" + str(i)].value = "" + str(round(t.persone.salary))
                    else:
                        if str(t.persone.grade) == '1':
                            t_list["H" + str(i)].value = "60000/1"
                        elif str(t.persone.grade) == '2':
                            t_list["H" + str(i)].value = "80000/2"
                        elif str(t.persone.grade) == '3':
                            t_list["H" + str(i)].value = "87000/3"
                        elif str(t.persone.grade) == '4':
                            t_list["H" + str(i)].value = "92000/4"
                        elif str(t.persone.grade) == '5':
                            t_list["H" + str(i)].value = "100000/5"
                        elif str(t.persone.grade) == '6':
                            t_list["H" + str(i)].value = "106000/6"
                else:
                    t_list["H" + str(i)].value = "" + str(round(t.persone.salary))
                if t.d1: t_list["J" + str(i)].value = "" + str(t.d1)
                if t.d2: t_list["K" + str(i)].value = "" + str(t.d2)
                if t.d3: t_list["L" + str(i)].value = "" + str(t.d3)
                if t.d4: t_list["M" + str(i)].value = "" + str(t.d4)
                if t.d5: t_list["N" + str(i)].value = "" + str(t.d5)
                if t.d6: t_list["O" + str(i)].value = "" + str(t.d6)
                if t.d7: t_list["P" + str(i)].value = "" + str(t.d7)
                if t.d8: t_list["Q" + str(i)].value = "" + str(t.d8)
                if t.d9: t_list["R" + str(i)].value = "" + str(t.d9)
                if t.d10: t_list["S" + str(i)].value = "" + str(t.d10)
                if t.d11: t_list["T" + str(i)].value = "" + str(t.d11)
                if t.d12: t_list["U" + str(i)].value = "" + str(t.d12)
                if t.d13: t_list["V" + str(i)].value = "" + str(t.d13)
                if t.d14: t_list["W" + str(i)].value = "" + str(t.d14)
                if t.d15: t_list["X" + str(i)].value = "" + str(t.d15)
                if t.d16: t_list["Y" + str(i)].value = "" + str(t.d16)
                if t.d17: t_list["Z" + str(i)].value = "" + str(t.d17)
                if t.d18: t_list["AA" + str(i)].value = "" + str(t.d18)
                if t.d19: t_list["AB" + str(i)].value = "" + str(t.d19)
                if t.d20: t_list["AC" + str(i)].value = "" + str(t.d20)
                if t.d21: t_list["AD" + str(i)].value = "" + str(t.d21)
                if t.d22: t_list["AE" + str(i)].value = "" + str(t.d22)
                if t.d23: t_list["AF" + str(i)].value = "" + str(t.d23)
                if t.d24: t_list["AG" + str(i)].value = "" + str(t.d24)
                if t.d25: t_list["AH" + str(i)].value = "" + str(t.d25)
                if t.d26: t_list["AI" + str(i)].value = "" + str(t.d26)
                if t.d27: t_list["AJ" + str(i)].value = "" + str(t.d27)
                if t.d28: t_list["AK" + str(i)].value = "" + str(t.d28)
                if t.d29: t_list["AL" + str(i)].value = "" + str(t.d29)
                if t.d30: t_list["AM" + str(i)].value = "" + str(t.d30)
                if t.d31: t_list["AN" + str(i)].value = "" + str(t.d31)
                if t.days: t_list["AP" + str(i)].value = "" + str(t.days)
                if t.add: t_list["AQ" + str(i)].value = "" + str(round(t.add))
                if t.fine: t_list["AR" + str(i)].value = "" + str(round(t.fine))
                if t.salary: t_list["AS" + str(i)].value = "" + str(round(t.salary))
                if t.comment: t_list["AT" + str(i)].value = "" + str(t.comment)
                if request.user.username == 'KaratovaMadina' or request.user.username == 'DjavathanovaP' or request.user.is_superuser:
                    t_list["AU6"].value = "Заметки"
                    if t.persone.idea: t_list["AU" + str(i)].value = str(t.persone.idea)
                i += 1
                n += 1
        book.title = 'Отчет общий'
        book.save(response)
        return response
        return redirect('/tabels/')

    @login_required(login_url="/login/")
    def month_info_driver(request, years, month, name):
        date = datetime.date(years, month, 1)
        pers = driver.objects.all().filter(date=date)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('mix.xlsx')
        t_list: worksheet = book.worksheets[0]
        i = 8
        n = 1
        for p in pers:
            tls = driver_list.objects.all().filter(period=p)
            m = int(p.date.month) - 1
            month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября',
                     'Ноября', 'Декабря']
            my = str(month[m]) + ' ' + str(p.date.strftime("%Y"))
            for t in tls:
                if t.persone:
                    if t.persone.company == 1 or t.persone.company == 'ПМК' or t.persone.company == 'пмк':
                        org = 'ПМК'
                    else:
                        org = 'неоформ'
                    today = datetime.date.today()
                    if p.date.month == today.month:
                      try:
                          tax = driver_tax.objects.get(driver=t.persone)
                      except:
                          tax = driver_tax_lagacy.objects.all().filter(driver=t.persone, date=p.date).last()
                    else:
                      try:
                          tax = driver_tax_lagacy.objects.all().filter(driver=t.persone, date=p.date).last()
                      except:
                          tax = driver_tax.objects.get(driver=t.persone)
                    t_list.insert_rows(i)
                    if tax is None:
                        t_list["H" + str(i)].value = "Н/Д"
                    else:
                        t_list["H" + str(i)].value = "" + str(round(tax.tax))
                    t_list["A" + str(i)].value = "" + str(n)
                    t_list["B" + str(i)].value = "" + str(org)
                    t_list["C" + str(i)].value = "" + str(t.persone.snils)
                    t_list["D" + str(i)].value = "" + str(t.full_name)
                    t_list["E" + str(i)].value = "" + str(t.position)
                    t_list["F" + str(i)].value = "(В)" + str(t.period.obj)
                    t_list["G" + str(i)].value = "ОГМ"
                    t_list["I" + str(i)].value = "" + str(t.auto)
                    if t.d1: t_list["J" + str(i)].value = "" + str(t.d1)
                    if t.d2: t_list["K" + str(i)].value = "" + str(t.d2)
                    if t.d3: t_list["L" + str(i)].value = "" + str(t.d3)
                    if t.d4: t_list["M" + str(i)].value = "" + str(t.d4)
                    if t.d5: t_list["N" + str(i)].value = "" + str(t.d5)
                    if t.d6: t_list["O" + str(i)].value = "" + str(t.d6)
                    if t.d7: t_list["P" + str(i)].value = "" + str(t.d7)
                    if t.d8: t_list["Q" + str(i)].value = "" + str(t.d8)
                    if t.d9: t_list["R" + str(i)].value = "" + str(t.d9)
                    if t.d10: t_list["S" + str(i)].value = "" + str(t.d10)
                    if t.d11: t_list["T" + str(i)].value = "" + str(t.d11)
                    if t.d12: t_list["U" + str(i)].value = "" + str(t.d12)
                    if t.d13: t_list["V" + str(i)].value = "" + str(t.d13)
                    if t.d14: t_list["W" + str(i)].value = "" + str(t.d14)
                    if t.d15: t_list["X" + str(i)].value = "" + str(t.d15)
                    if t.d16: t_list["Y" + str(i)].value = "" + str(t.d16)
                    if t.d17: t_list["Z" + str(i)].value = "" + str(t.d17)
                    if t.d18: t_list["AA" + str(i)].value = "" + str(t.d18)
                    if t.d19: t_list["AB" + str(i)].value = "" + str(t.d19)
                    if t.d20: t_list["AC" + str(i)].value = "" + str(t.d20)
                    if t.d21: t_list["AD" + str(i)].value = "" + str(t.d21)
                    if t.d22: t_list["AE" + str(i)].value = "" + str(t.d22)
                    if t.d23: t_list["AF" + str(i)].value = "" + str(t.d23)
                    if t.d24: t_list["AG" + str(i)].value = "" + str(t.d24)
                    if t.d25: t_list["AH" + str(i)].value = "" + str(t.d25)
                    if t.d26: t_list["AI" + str(i)].value = "" + str(t.d26)
                    if t.d27: t_list["AJ" + str(i)].value = "" + str(t.d27)
                    if t.d28: t_list["AK" + str(i)].value = "" + str(t.d28)
                    if t.d29: t_list["AL" + str(i)].value = "" + str(t.d29)
                    if t.d30: t_list["AM" + str(i)].value = "" + str(t.d30)
                    if t.d31: t_list["AN" + str(i)].value = "" + str(t.d31)
                    if t.hours: t_list["AO" + str(i)].value = "" + str(t.hours)
                    if t.days: t_list["AP" + str(i)].value = "" + str(t.days)
                    if t.add: t_list["AQ" + str(i)].value = "" + str(round(t.add))
                    if t.fine: t_list["AR" + str(i)].value = "" + str(round(t.fine))
                    if t.salary: t_list["AS" + str(i)].value = "" + str(round(t.salary))
                    if t.comment: t_list["AT" + str(i)].value = "" + str(t.comment)
                    if request.user.username == 'KaratovaMadina' or request.user.username == 'DjavathanovaP' or request.user.is_superuser:
                        t_list["AU6"].value = "Заметки"
                        if t.persone.idea: t_list["AT" + str(i)].value = str(t.persone.idea)
                    i += 1
                    n += 1
        book.title = 'Отчет общий'
        book.save(response)
        return response
        return redirect('/tabels/')


@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))
