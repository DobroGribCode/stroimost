# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
import datetime
import io
import random
import mimetypes
import os
import cherrypy
import openpyxl
import tempfile
import dateutil
from dateutil import parser

from openpyxl import Workbook

from docx import Document
from docx.shared import Pt
from docx.shared import Inches
from docxtpl import DocxTemplate

from io import StringIO, BytesIO
from cherrypy.lib import file_generator

from random import randint
from random import randrange

from datetime import date
from calendar import monthrange
from wsgiref.util import FileWrapper
import mimetypes
import os, os.path

from django.core.files.base import File
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django import template
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.template import loader
from django.shortcuts import render, redirect,get_object_or_404
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.files.base import File
from django.contrib.auth.decorators import login_required
from openpyxl.worksheet import worksheet
from django.conf import settings
from django.http import HttpResponse, Http404
from openpyxl.styles import PatternFill, Border
from ..reference.forms import TabForm, ObjsForm, TaskForm, AutoForm, GroupForm,PersoneForm, DriverForm, AgreementForm, TempFileForm
from ..reference.models import temp_file,tabel_group, category, dayapprove,dayapproved,persone, period,tabel_list, objs,task_head, task_body, group, automobile, driver, driver_list, agreement, agreement_accepter,agreement_step,accepter_list,agreementer_list, ender_list
from django.contrib.auth.models import User

class tab:
    @login_required(login_url="/login/")
    def tabels(request):
        context = {'segment': 'tabels'}
        html_template = loader.get_template('tabels/tabels.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def month_info(request):
        return redirect('/tabels/')


    @login_required(login_url="/login/")
    def driver(request):
        per = driver.objects.all()
        groups = tabel_group.objects.all()
        if groups:
            pass
        else:
            groups
        context = {'per': per, 'segment': 'driver', 'groups': groups}
        html_template = loader.get_template('tabels/driver/driver.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_driver(request):
        dri = driver.objects.all()
        form = DriverForm
        persones = persone.objects.all().filter(date_leave=None,driver=1)
        if request.method == 'POST':
            forms = DriverForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                year =request.POST.get("ye")
                month = request.POST.get("mo")
                da = year + '-' + month + '-' + '01'
                post.date = da
                da = date(int(year),int(month), 1)
                post.save()
                try:
                    da = da.replace(day=1)
                    post.dw1 = da
                except:
                    pass
                try:
                    da = da.replace(day=2)
                    post.dw2 = da
                except:
                    pass
                try:
                    da = da.replace(day=3)
                    post.dw3 = da
                except:
                    pass
                try:
                    da = da.replace(day=4)
                    post.dw4 = da
                except:
                    pass
                try:
                    da = da.replace(day=5)
                    post.dw5 = da
                except:
                    pass
                try:
                    da = da.replace(day=6)
                    post.dw6 = da
                except:
                    pass
                try:
                    da = da.replace(day=7)
                    post.dw7 = da
                except:
                    pass
                try:
                    da = da.replace(day=8)
                    post.dw8 = da
                except:
                    pass
                try:
                    da = da.replace(day=9)
                    post.dw9 = da
                except:
                    pass
                try:
                    da = da.replace(day=10)
                    post.dw10 = da
                except:
                    pass
                try:
                    da = da.replace(day=11)
                    post.dw11 = da
                except:
                    pass
                try:
                    da = da.replace(day=12)
                    post.dw12 = da
                except:
                    pass
                try:
                    da = da.replace(day=13)
                    post.dw13 = da
                except:
                     pass
                try:
                    da = da.replace(day=14)
                    post.dw14 = da
                except:
                     pass
                try:
                    da = da.replace(day=15)
                    post.dw15 = da
                except:
                    pass
                try:
                    da = da.replace(day=16)
                    post.dw16 = da
                except:
                    pass
                try:
                    da = da.replace(day=17)
                    post.dw17 = da
                except:
                    pass
                try:
                    da = da.replace(day=18)
                    post.dw18 = da
                except:
                    pass
                try:
                    da = da.replace(day=19)
                    post.dw19 = da
                except:
                    pass
                try:
                    da = da.replace(day=20)
                    post.dw20 = da
                except:
                    pass
                try:
                    da = da.replace(day=21)
                    post.dw21 = da
                except:
                    pass
                try:
                    da = da.replace(day=22)
                    post.dw22 = da
                except:
                    pass
                try:
                    da = da.replace(day=23)
                    post.dw23 = da
                except:
                    pass
                try:
                    da = da.replace(day=24)
                    post.dw24 = da
                except:
                    pass
                try:
                    da = da.replace(day=25)
                    post.dw25 = da
                except:
                    pass
                try:
                    da = da.replace(day=26)
                    post.dw26 = da
                except:
                    pass
                try:
                    da = da.replace(day=27)
                    post.dw27 = da
                except:
                    pass
                try:
                    da = da.replace(day=28)
                    post.dw28 = da
                except:
                    pass
                try:
                    da = da.replace(day=29)
                    post.dw29 = da
                except:
                    pass
                try:
                    da = da.replace(day=30)
                    post.dw30 = da
                except:
                    pass
                try:
                    da = da.replace(day=31)
                    post.dw31 = da
                except:
                    pass
                post.save()
                dris = driver.objects.all().first()
                return redirect('view_driver/'+str(dris.id))
        else:
            context = {
                'segment': 'driver',
                'dri': dri,
                'form': form,
            }
            html_template = loader.get_template('tabels/driver/new_driver.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def group_move_tabelD(request, ids, id):
        per_old = driver.objects.get(id=ids)
        tabs = tabel_list.objects.all().filter(period=per_old)
        per = driver.objects.get(id=id)
        for t in tabs:
            p = persone.objects.get(id=t.persone.id)
            driver_list(
                period=per,
                persone=p,
                full_name=p.full_name,
                position=p.position,
                salary=p.salary,
                output=p.output,
            ).save()
        return redirect('/view_driver/' + str(id))

    @login_required(login_url="/login/")
    def AddPersoneD(request, id, ids):
        per = driver.objects.get(id=ids)
        p = persone.objects.get(id=id)
        driver_list(
            period=per,
            persone=p,
            full_name=p.full_name,
            position=p.position,
            salary=p.salary,
            output=p.output,
        ).save()
        return redirect('/view_driver/' + str(ids))

    @login_required(login_url="/login/")
    def view_driver(request, id):
        per = driver.objects.get(id=id)
        pers = driver.objects.all()
        persones = persone.objects.all().filter(date_leave=None,driver=1)
        all = persone.objects.all().filter(date_leave=None)
        tabs = driver_list.objects.all().filter(period=per).order_by('full_name')
        autos = automobile.objects.all()
        approvers = dayapproved.objects.all().filter(per=per)
        html_template = loader.get_template('tabels/driver/view_driver.html')
        if per.date:
            month = datetime.date.today()
            year = str(month)[0:4]
            month = str(month)[5:7]
            pm = per.date
            py = str(pm)[0:4]
            pm = str(pm)[5:7]
            if int(month) > int(pm):
                if int(year) == int(py):
                    per.fullblock = 1;
                    per.save()
            if per.fullblock == 1:
                pass
                today = 32
            else:
                today = datetime.date.today()
                today = str(today)[8:]
                today = int(today) - 2
            if int(month) >= int(pm):
                if today > 1:
                    per.db1 = 1
                    per.save()
                if today > 2:
                    per.db2 = 1
                    per.save()
                if today > 3:
                    per.db3 = 1
                    per.save()
                if today > 4:
                    per.db4 = 1
                    per.save()
                if today > 5:
                    per.db5 = 1
                    per.save()
                if today > 6:
                    per.db6 = 1
                    per.save()
                if today > 7:
                    per.db7 = 1
                    per.save()
                if today > 8:
                    per.db8 = 1
                    per.save()
                if today > 9:
                    per.db9 = 1
                    per.save()
                if today > 10:
                    per.db10 = 1
                    per.save()
                if today > 11:
                    per.db11 = 1
                    per.save()
                if today > 12:
                    per.db12 = 1
                    per.save()
                if today > 13:
                    per.db13 = 1
                    per.save()
                if today > 14:
                    per.db14 = 1
                    per.save()
                if today > 15:
                    per.db15 = 1
                    per.save()
                if today > 16:
                    per.db16 = 1
                    per.save()
                if today > 17:
                    per.db17 = 1
                    per.save()
                if today > 18:
                    per.db18 = 1
                    per.save()
                if today > 19:
                    per.db19 = 1
                    per.save()
                if today > 20:
                    per.db20 = 1
                    per.save()
                if today > 21:
                    per.db21 = 1
                    per.save()
                if today > 22:
                    per.db22 = 1
                    per.save()
                if today > 23:
                    per.db23 = 1
                    per.save()
                if today > 24:
                    per.db24 = 1
                    per.save()
                if today > 25:
                    per.db25 = 1
                    per.save()
                if today > 26:
                    per.db26 = 1
                    per.save()
                if today > 27:
                    per.db27 = 1
                    per.save()
                if today > 28:
                    per.db28 = 1
                    per.save()
                if today > 29:
                    per.db29 = 1
                    per.save()
                if today > 30:
                    per.db30 = 1
                    per.save()
                if today > 31:
                    per.db31 = 1
                    per.save()
            if per:
                i1 = 0
                i2 = 0
                i3 = 0
                i4 = 0
                i5 = 0
                i6 = 0
                i7 = 0
                i8 = 0
                i9 = 0
                i10 = 0
                i11 = 0
                i12 = 0
                i13 = 0
                i14 = 0
                i15 = 0
                i16 = 0
                i17 = 0
                i18 = 0
                i19 = 0
                i20 = 0
                i21 = 0
                i22 = 0
                i23 = 0
                i24 = 0
                i25 = 0
                i26 = 0
                i27 = 0
                i28 = 0
                i29 = 0
                i30 = 0
                i31 = 0
            for a in approvers:
                if a.day == 1:
                    i1 += 1
                if i1 >= 2:
                    per.da1 = 1
                if a.day == 2:
                    i2 += 1
                if i2 >= 2:
                    per.da2 = 1
                if a.day == 3:
                    i3 += 1
                if i3 >= 2:
                    per.da3 = 1
                if a.day == 4:
                    i4 += 1
                if i4 >= 2:
                    per.da4 = 1
                if a.day == 5:
                    i5 += 1
                if i5 >= 2:
                    per.da5 = 1
                if a.day == 6:
                    i6 += 1
                if i6 >= 2:
                    per.da6 = 1
                if a.day == 7:
                    i7 += 1
                if i7 >= 2:
                    per.da7 = 1
                if a.day == 8:
                    i8 += 1
                if i8 >= 2:
                    per.da8 = 1
                if a.day == 9:
                    i9 += 1
                if i9 >= 2:
                    per.da9 = 1
                if a.day == 10:
                    i10 += 1
                if i10 >= 2:
                    per.da10 = 1
                if a.day == 11:
                    i11 += 1
                if i11 >= 2:
                    per.da11 = 1
                if a.day == 12:
                    i12 += 1
                if i12 >= 2:
                    per.da12 = 1
                if a.day == 13:
                    i13 += 1
                if i13 >= 2:
                    per.da13 = 1
                if a.day == 14:
                    i14 += 1
                if i14 >= 2:
                    per.da14 = 1
                if a.day == 15:
                    i15 += 1
                if i15 >= 2:
                    per.da15 = 1
                if a.day == 16:
                    i16 += 1
                if i16 >= 2:
                    per.da16 = 1
                if a.day == 17:
                    i17 += 1
                if i17 >= 2:
                    per.da17 = 1
                if a.day == 18:
                    i18 += 1
                if i18 >= 2:
                    per.da18 = 1
                if a.day == 19:
                    i19 += 1
                if i19 >= 2:
                    per.da19 = 1
                if a.day == 20:
                    i20 += 1
                if i20 >= 2:
                    per.da20 = 1
                if a.day == 21:
                    i21 += 1
                if i21 >= 2:
                    per.da21 = 1
                if a.day == 22:
                    i22 += 1
                if i22 >= 2:
                    per.da22 = 1
                if a.day == 23:
                    i23 += 1
                if i23 >= 2:
                    per.da23 = 1
                if a.day == 24:
                    i24 += 1
                if i24 >= 2:
                    per.da24 = 1
                if a.day == 25:
                    i25 += 1
                if i25 >= 2:
                    per.da25 = 1
                if a.day == 26:
                    i26 += 1
                if i26 >= 2:
                    per.da26 = 1
                if a.day == 27:
                    i27 += 1
                if i27 >= 2:
                    per.da27 = 1
                if a.day == 28:
                    i28 += 1
                if i28 >= 2:
                    per.da28 = 1
                if a.day == 29:
                    i29 += 1
                if i29 >= 2:
                    per.da29 = 1
                if a.day == 30:
                    i30 += 1
                if i30 >= 2:
                    per.da30 = 1
                if a.day == 31:
                    i31 += 1
                if i31 >= 2:
                    per.da31 = 1
                per.save()
        if request.method == 'POST':
            for t in tabs:
                da1 = request.POST.get(str(t.id) + 'd1');
                da2 = request.POST.get(str(t.id) + 'd2');
                da3 = request.POST.get(str(t.id) + 'd3');
                da4 = request.POST.get(str(t.id) + 'd4');
                da5 = request.POST.get(str(t.id) + 'd5');
                da6 = request.POST.get(str(t.id) + 'd6');
                da7 = request.POST.get(str(t.id) + 'd7');
                da8 = request.POST.get(str(t.id) + 'd8');
                da9 = request.POST.get(str(t.id) + 'd9');
                da10 = request.POST.get(str(t.id) + 'd10');
                da11 = request.POST.get(str(t.id) + 'd11');
                da12 = request.POST.get(str(t.id) + 'd12');
                da13 = request.POST.get(str(t.id) + 'd13');
                da14 = request.POST.get(str(t.id) + 'd14');
                da15 = request.POST.get(str(t.id) + 'd15');
                da16 = request.POST.get(str(t.id) + 'd16');
                da17 = request.POST.get(str(t.id) + 'd17');
                da18 = request.POST.get(str(t.id) + 'd18');
                da19 = request.POST.get(str(t.id) + 'd19');
                da20 = request.POST.get(str(t.id) + 'd20');
                da21 = request.POST.get(str(t.id) + 'd21');
                da22 = request.POST.get(str(t.id) + 'd22');
                da23 = request.POST.get(str(t.id) + 'd23');
                da24 = request.POST.get(str(t.id) + 'd24');
                da25 = request.POST.get(str(t.id) + 'd25');
                da26 = request.POST.get(str(t.id) + 'd26');
                da27 = request.POST.get(str(t.id) + 'd27');
                da28 = request.POST.get(str(t.id) + 'd28');
                da29 = request.POST.get(str(t.id) + 'd29');
                da30 = request.POST.get(str(t.id) + 'd30');
                da31 = request.POST.get(str(t.id) + 'd31');
                # if t:
                #     try:
                #         check_per = driver.objects.all().filter(date=per.date)
                #         for cp in check_per:
                #             if cp.id != per.id:
                #                 check_tl = driver_list.objects.all().filter(period=cp)
                #                 for ct in check_tl:
                #                     if ct.d1:
                #                         if da1:
                #                             error_d = 1
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d2:
                #                         if da2:
                #                             error_d = 2
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d3:
                #                         if da3:
                #                             error_d = 3
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d4:
                #                         if da4:
                #                             error_d = 4
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d5:
                #                         if da5:
                #                             error_d = 5
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d6:
                #                         if da6:
                #                             error_d = 6
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d7:
                #                         if da7:
                #                             error_d = 7
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d8:
                #                         if da8:
                #                             error_d = 8
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d9:
                #                         if da9:
                #                             error_d = 9
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d10:
                #                         if da10:
                #                             error_d = 10
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d11:
                #                         if da11:
                #                             error_d = 11
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d12:
                #                         if da12:
                #                             error_d = 12
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d13:
                #                         if da13:
                #                             error_d = 13
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d14:
                #                         if da14:
                #                             error_d = 14
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d15:
                #                         if da15:
                #                             error_d = 15
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d16:
                #                         if da16:
                #                             error_d = 16
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d17:
                #                         if da17:
                #                             error_d = 17
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d18:
                #                         if da18:
                #                             error_d = 18
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d19:
                #                         if da19:
                #                             error_d = 19
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d20:
                #                         if da20:
                #                             error_d = 20
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d21:
                #                         if da21:
                #                             error_d = 21
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d22:
                #                         if da22:
                #                             error_d = 22
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d23:
                #                         if da23:
                #                             error_d = 23
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d24:
                #                         if da24:
                #                             error_d = 24
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d25:
                #                         if da25:
                #                             error_d = 25
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d26:
                #                         if da26:
                #                             error_d = 26
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d27:
                #                         if da27:
                #                             error_d = 27
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d28:
                #                         if da28:
                #                             error_d = 28
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d29:
                #                         if da29:
                #                             error_d = 29
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d30:
                #                         if da30:
                #                             error_d = 30
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d31:
                #                         if da31:
                #                             error_d = 31
                #                             error_p = ct.persone
                #                             context = {
                #                                 'persones': persones,
                #
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #     except:
                #         pass
                if t:
                    words = ["ВВ","Н","М","Б","О","Д","В","У"]
                    if da1:
                        try:
                            da1 = int(da1)
                            if da1 > 0 and da1 <= 23:
                                t.d1 = da1
                        except:
                            da1 = da1.upper()
                            if da1 in words:
                                t.d1 = da1
                                t.save()
                        t.save()
                    if da2:
                        try:
                            da2 = int(da2)
                            if da2 > 0 and da2 <= 23:
                                t.d2 = da2
                        except:
                            da2 = da2.upper()
                            if da2 in words:
                                t.d2 = da2
                                t.save()
                        t.save()
                    if da3:
                        try:
                            da3 = int(da3)
                            if da3 > 0 and da3 <= 23:
                                t.d3 = da3
                        except:
                            da3 = da3.upper()
                            if da3 in words:
                                t.d3 = da3
                                t.save()
                        t.save()
                    if da4:
                        try:
                            da4 = int(da4)
                            if da4 > 0 and da4 <= 23:
                                t.d4 = da4
                        except:
                            da4 = da4.upper()
                            if da4 in words:
                                t.d4 = da4
                                t.save()
                        t.save()
                    if da5:
                        try:
                            da5 = int(da5)
                            if da5 > 0 and da5 <= 23:
                                t.d5 = da5
                        except:
                            da5 = da5.upper()
                            if da5 in words:
                                t.d5 = da5
                                t.save()
                        t.save()
                    if da6:
                        try:
                            da6 = int(da6)
                            if da6 > 0 and da6 <= 23:
                                t.d6 = da6
                        except:
                            da6 = da6.upper()
                            if da6 in words:
                                t.d6 = da6
                                t.save()
                        t.save()
                    if da7:
                        try:
                            da7 = int(da7)
                            if da7 > 0 and da7 <= 23:
                                t.d7 = da7
                        except:
                            da7 = da7.upper()
                            if da7 in words:
                                t.d7 = da7
                                t.save()
                        t.save()
                    if da8:
                        try:
                            da8 = int(da8)
                            if da8 > 0 and da8 <= 23:
                                t.d8 = da8
                        except:
                            da8 = da8.upper()
                            if da8 in words:
                                t.d8 = da8
                                t.save()
                        t.save()
                    if da9:
                        try:
                            da9 = int(da9)
                            if da9 > 0 and da9 <= 23:
                                t.d9 = da9
                        except:
                            da9 = da9.upper()
                            if da9 in words:
                                t.d9 = da9
                                t.save()
                        t.save()
                    if da10:
                        try:
                            da10 = int(da10)
                            if da10 > 0 and da10 <= 23:
                                t.d10 = da10
                        except:
                            da10 = da10.upper()
                            if da10 in words:
                                t.d10 = da10
                                t.save()
                        t.save()
                    if da11:
                        try:
                            da11 = int(da11)
                            if da11 > 0 and da11 <= 23:
                                t.d11 = da11
                        except:
                            da11 = da11.upper()
                            if da11 in words:
                                t.d11 = da11
                                t.save()
                        t.save()
                    if da12:
                        try:
                            da12 = int(da12)
                            if da12 > 0 and da12 <= 23:
                                t.d12 = da12
                        except:
                            da12 = da12.upper()
                            if da12 in words:
                                t.d12 = da12
                                t.save()
                        t.save()
                    if da13:
                        try:
                            da13 = int(da13)
                            if da13 > 0 and da13 <= 23:
                                t.d13 = da13
                        except:
                            da13 = da13.upper()
                            if da13 in words:
                                t.d13 = da13
                                t.save()
                        t.save()
                    if da14:
                        try:
                            da14 = int(da14)
                            if da14 > 0 and da14 <= 23:
                                t.d14 = da14
                        except:
                            da14 = da14.upper()
                            if da14 in words:
                                t.d14 = da14
                                t.save()
                        t.save()
                    if da15:
                        try:
                            da15 = int(da15)
                            if da15 > 0 and da15 <= 23:
                                t.d15 = da15
                        except:
                            da15 = da15.upper()
                            if da15 in words:
                                t.d15 = da15
                                t.save()
                        t.save()
                    if da16:
                        try:
                            da16 = int(da16)
                            if da16 > 0 and da16 <= 23:
                                t.d16 = da16
                        except:
                            da16 = da16.upper()
                            if da16 in words:
                                t.d16 = da16
                                t.save()
                        t.save()
                    if da17:
                        try:
                            da17 = int(da17)
                            if da17 > 0 and da17 <= 23:
                                t.d17 = da17
                        except:
                            da17 = da17.upper()
                            if da17 in words:
                                t.d17 = da17
                                t.save()
                        t.save()
                    if da18:
                        try:
                            da18 = int(da18)
                            if da18 > 0 and da18 <= 23:
                                t.d18 = da18
                        except:
                            da18 = da18.upper()
                            if da18 in words:
                                t.d18 = da18
                                t.save()
                        t.save()
                    if da19:
                        try:
                            da19 = int(da19)
                            if da19 > 0 and da19 <= 23:
                                t.d19 = da19
                        except:
                            da19 = da19.upper()
                            if da19 in words:
                                t.d19 = da19
                                t.save()
                        t.save()
                    if da20:
                        try:
                            da20 = int(da20)
                            if da20 > 0 and da20 <= 23:
                                t.d20 = da20
                        except:
                            da20 = da20.upper()
                            if da20 in words:
                                t.d20 = da20
                                t.save()
                        t.save()
                    if da21:
                        try:
                            da21 = int(da21)
                            if da21 > 0 and da21 <= 23:
                                t.d21 = da21
                        except:
                            da21 = da21.upper()
                            if da21 in words:
                                t.d21 = da21
                                t.save()
                        t.save()
                    if da22:
                        try:
                            da22 = int(da22)
                            if da22 > 0 and da22 <= 23:
                                t.d22 = da22
                        except:
                            da22 = da22.upper()
                            if da22 in words:
                                t.d22 = da22
                                t.save()
                        t.save()
                    if da23:
                        try:
                            da23 = int(da23)
                            if da23 > 0 and da23 <= 23:
                                t.d23 = da23
                        except:
                            da23 = da23.upper()
                            if da23 in words:
                                t.d23 = da23
                                t.save()
                        t.save()
                    if da24:
                        try:
                            da24 = int(da24)
                            if da24 > 0 and da24 <= 23:
                                t.d24 = da24
                        except:
                            da24 = da24.upper()
                            if da24 in words:
                                t.d24 = da24
                                t.save()
                        t.save()
                    if da25:
                        try:
                            da25 = int(da25)
                            if da25 > 0 and da25 <= 23:
                                t.d25 = da25
                        except:
                            da25 = da25.upper()
                            if da25 in words:
                                t.d25 = da25
                                t.save()
                        t.save()
                    if da26:
                        try:
                            da26 = int(da26)
                            if da26 > 0 and da26 <= 23:
                                t.d26 = da26
                        except:
                            da26 = da26.upper()
                            if da26 in words:
                                t.d26 = da26
                                t.save()
                        t.save()
                    if da27:
                        try:
                            da27 = int(da27)
                            if da27 > 0 and da27 <= 23:
                                t.d27 = da27
                        except:
                            da27 = da27.upper()
                            if da27 in words:
                                t.d27 = da27
                                t.save()
                        t.save()
                    if da28:
                        try:
                            da28 = int(da28)
                            if da28 > 0 and da28 <= 23:
                                t.d28 = da28
                        except:
                            da28 = da28.upper()
                            if da28 in words:
                                t.d28 = da28
                                t.save()
                        t.save()
                    if da29:
                        try:
                            da29 = int(da29)
                            if da29 > 0 and da29 <= 23:
                                t.d29 = da29
                        except:
                            da29 = da29.upper()
                            if da29 in words:
                                t.d29 = da29
                                t.save()
                        t.save()
                    if da30:
                        try:
                            da30 = int(da30)
                            if da30 > 0 and da30 <= 23:
                                t.d30 = da30
                        except:
                            da30 = da30.upper()
                            if da30 in words:
                                t.d30 = da30
                                t.save()
                        t.save()
                    if da31:
                        try:
                            da31 = int(da31)
                            if da31 > 0 and da31 <= 23:
                                t.d31 = da31
                        except:
                            da31 = da31.upper()
                            if da31 in words:
                                t.d31 = da31
                                t.save()
                        t.save()
                if t:
                    if t.d1:
                        try:
                            d = int(t.d1)
                            if t.d1 > 10:
                                t.d1c = "#00b300"
                        except:
                            if t.d1 == "ВВ": t.d1c = "#f2e93f"
                            if t.d1 == "Н": t.d1c = "#808080"
                            if t.d1 == "М": t.d1c = "#ffa812"
                            if t.d1 == "Б" or t.d1 == "О" or t.d1 == "Д" or t.d1 == "У": t.d1c = "#ffff00"
                            if t.d1 == "В": t.d1c = "#fa5555"
                        t.save()
                    if t.d2:
                        try:
                            d = int(t.d2)
                            if t.d2 > 10:
                                t.d2c = "#00b300"
                        except:
                            if t.d2 == "ВВ": t.d2c = "#f2e93f"
                            if t.d2 == "Н": t.d2c = "#808080"
                            if t.d2 == "М": t.d2c = "#ffa812"
                            if t.d2 == "Б" or t.d2 == "О" or t.d2 == "Д" or t.d2 == "У": t.d2c = "#ffff00"
                            if t.d2 == "В": t.d2c = "#fa5555"
                        t.save()
                    if t.d3:
                        try:
                            d = int(t.d3)
                            if t.d3 > 10:
                                t.d3c = "#00b300"
                        except:
                            if t.d3 == "ВВ": t.d3c = "#f2e93f"
                            if t.d3 == "Н": t.d3c = "#808080"
                            if t.d3 == "М": t.d3c = "#ffa812"
                            if t.d3 == "Б" or t.d3 == "О" or t.d3 == "Д" or t.d3 == "У": t.d3c = "#ffff00"
                            if t.d3 == "В": t.d3c = "#fa5555"
                        t.save()
                    if t.d4:
                        try:
                            d = int(t.d4)
                            if t.d4 > 10:
                                t.d4c = "#00b300"
                        except:
                            if t.d4 == "ВВ": t.d4c = "#f2e93f"
                            if t.d4 == "Н": t.d4c = "#808080"
                            if t.d4 == "М": t.d4c = "#ffa812"
                            if t.d4 == "Б" or t.d4 == "О" or t.d4 == "Д" or t.d4 == "У": t.d4c = "#ffff00"
                            if t.d4 == "В": t.d4c = "#fa5555"
                        t.save()
                    if t.d5:
                        try:
                            d = int(t.d5)
                            if t.d5 > 10:
                                t.d5c = "#00b300"
                        except:
                            if t.d5 == "ВВ": t.d5c = "#f2e93f"
                            if t.d5 == "Н": t.d5c = "#808080"
                            if t.d5 == "М": t.d5c = "#ffa812"
                            if t.d5 == "Б" or t.d5 == "О" or t.d5 == "Д" or t.d5 == "У": t.d5c = "#ffff00"
                            if t.d5 == "В": t.d5c = "#fa5555"
                        t.save()
                    if t.d6:
                        try:
                            d = int(t.d6)
                            if t.d6 > 10:
                                t.d6c = "#00b300"
                        except:
                            if t.d6 == "ВВ": t.d6c = "#f2e93f"
                            if t.d6 == "Н": t.d6c = "#808080"
                            if t.d6 == "М": t.d6c = "#ffa812"
                            if t.d6 == "Б" or t.d6 == "О" or t.d6 == "Д" or t.d6 == "У": t.d6c = "#ffff00"
                            if t.d6 == "В": t.d6c = "#fa5555"
                        t.save()
                    if t.d7:
                        try:
                            d = int(t.d7)
                            if t.d7 > 10:
                                t.d7c = "#00b300"
                        except:
                            if t.d7 == "ВВ": t.d7c = "#f2e93f"
                            if t.d7 == "Н": t.d7c = "#808080"
                            if t.d7 == "М": t.d7c = "#ffa812"
                            if t.d7 == "Б" or t.d7 == "О" or t.d7 == "Д" or t.d7 == "У": t.d7c = "#ffff00"
                            if t.d7 == "В": t.d7c = "#fa5555"
                        t.save()
                    if t.d8:
                        try:
                            d = int(t.d8)
                            if t.d8 > 10:
                                t.d8c = "#00b300"
                        except:
                            if t.d8 == "ВВ": t.d8c = "#f2e93f"
                            if t.d8 == "Н": t.d8c = "#808080"
                            if t.d8 == "М": t.d8c = "#ffa812"
                            if t.d8 == "Б" or t.d8 == "О" or t.d8 == "Д" or t.d8 == "У": t.d8c = "#ffff00"
                            if t.d8 == "В": t.d8c = "#fa5555"
                        t.save()
                    if t.d9:
                        try:
                            d = int(t.d9)
                            if t.d9 > 10:
                                t.d9c = "#00b300"
                        except:
                            if t.d9 == "ВВ": t.d9c = "#f2e93f"
                            if t.d9 == "Н": t.d9c = "#808080"
                            if t.d9 == "М": t.d9c = "#ffa812"
                            if t.d9 == "Б" or t.d9 == "О" or t.d9 == "Д" or t.d9 == "У": t.d9c = "#ffff00"
                            if t.d9 == "В": t.d9c = "#fa5555"
                        t.save()
                    if t.d10:
                        try:
                            d = int(t.d10)
                            if t.d10 > 10:
                                t.d10c = "#00b300"
                        except:
                            if t.d10 == "ВВ": t.d10c = "#f2e93f"
                            if t.d10 == "Н": t.d10c = "#808080"
                            if t.d10 == "М": t.d10c = "#ffa812"
                            if t.d10 == "Б" or t.d10 == "О" or t.d10 == "Д" or t.d10 == "У": t.d10c = "#ffff00"
                            if t.d10 == "В": t.d10c = "#fa5555"
                        t.save()
                    if t.d11:
                        try:
                            d = int(t.d11)
                            if t.d11 > 10:
                                t.d11c = "#00b300"
                        except:
                            if t.d11 == "ВВ": t.d11c = "#f2e93f"
                            if t.d11 == "Н": t.d11c = "#808080"
                            if t.d11 == "М": t.d11c = "#ffa812"
                            if t.d11 == "Б" or t.d11 == "О" or t.d11 == "Д" or t.d11 == "У": t.d11c = "#ffff00"
                            if t.d11 == "В": t.d11c = "#fa5555"
                        t.save()
                    if t.d12:
                        try:
                            d = int(t.d12)
                            if t.d12 > 10:
                                t.d12c = "#00b300"
                        except:
                            if t.d12 == "ВВ": t.d12c = "#f2e93f"
                            if t.d12 == "Н": t.d12c = "#808080"
                            if t.d12 == "М": t.d12c = "#ffa812"
                            if t.d12 == "Б" or t.d12 == "О" or t.d12 == "Д" or t.d12 == "У": t.d12c = "#ffff00"
                            if t.d12 == "В": t.d12c = "#fa5555"
                        t.save()
                    if t.d13:
                        try:
                            d = int(t.d13)
                            if t.d13 > 10:
                                t.d13c = "#00b300"
                        except:
                            if t.d13 == "ВВ": t.d13c = "#f2e93f"
                            if t.d13 == "Н": t.d13c = "#808080"
                            if t.d13 == "М": t.d13c = "#ffa812"
                            if t.d13 == "Б" or t.d13 == "О" or t.d13 == "Д" or t.d13 == "У": t.d13c = "#ffff00"
                            if t.d13 == "В": t.d13c = "#fa5555"
                        t.save()
                    if t.d14:
                        try:
                            d = int(t.d14)
                            if t.d14 > 10:
                                t.d14c = "#00b300"
                        except:
                            if t.d14 == "ВВ": t.d14c = "#f2e93f"
                            if t.d14 == "Н": t.d14c = "#808080"
                            if t.d14 == "М": t.d14c = "#ffa812"
                            if t.d14 == "Б" or t.d14 == "О" or t.d14 == "Д" or t.d14 == "У": t.d14c = "#ffff00"
                            if t.d14 == "В": t.d14c = "#fa5555"
                        t.save()
                    if t.d15:
                        try:
                            d = int(t.d15)
                            if t.d15 > 10:
                                t.d15c = "#00b300"
                        except:
                            if t.d15 == "ВВ": t.d15c = "#f2e93f"
                            if t.d15 == "Н": t.d15c = "#808080"
                            if t.d15 == "М": t.d15c = "#ffa812"
                            if t.d15 == "Б" or t.d15 == "О" or t.d15 == "Д" or t.d15 == "У": t.d15c = "#ffff00"
                            if t.d15 == "В": t.d15c = "#fa5555"
                        t.save()
                    if t.d16:
                        try:
                            d = int(t.d16)
                            if t.d16 > 10:
                                t.d16c = "#00b300"
                        except:
                            if t.d16 == "ВВ": t.d16c = "#f2e93f"
                            if t.d16 == "Н": t.d16c = "#808080"
                            if t.d16 == "М": t.d16c = "#ffa812"
                            if t.d16 == "Б" or t.d16 == "О" or t.d16 == "Д" or t.d16 == "У": t.d16c = "#ffff00"
                            if t.d16 == "В": t.d16c = "#fa5555"
                        t.save()
                    if t.d17:
                        try:
                            d = int(t.d17)
                            if t.d17 > 10:
                                t.d17c = "#00b300"
                        except:
                            if t.d17 == "ВВ": t.d17c = "#f2e93f"
                            if t.d17 == "Н": t.d17c = "#808080"
                            if t.d17 == "М": t.d17c = "#ffa812"
                            if t.d17 == "Б" or t.d17 == "О" or t.d17 == "Д" or t.d17 == "У": t.d17c = "#ffff00"
                            if t.d17 == "В": t.d17c = "#fa5555"
                        t.save()
                    if t.d18:
                        try:
                            d = int(t.d18)
                            if t.d18 > 10:
                                t.d18c = "#00b300"
                        except:
                            if t.d18 == "ВВ": t.d18c = "#f2e93f"
                            if t.d18 == "Н": t.d18c = "#808080"
                            if t.d18 == "М": t.d18c = "#ffa812"
                            if t.d18 == "Б" or t.d18 == "О" or t.d18 == "Д" or t.d18 == "У": t.d18c = "#ffff00"
                            if t.d18 == "В": t.d18c = "#fa5555"
                        t.save()
                    if t.d19:
                        try:
                            d = int(t.d19)
                            if t.d19 > 10:
                                t.d19c = "#00b300"
                        except:
                            if t.d19 == "ВВ": t.d19c = "#f2e93f"
                            if t.d19 == "Н": t.d19c = "#808080"
                            if t.d19 == "М": t.d19c = "#ffa812"
                            if t.d19 == "Б" or t.d19 == "О" or t.d19 == "Д" or t.d19 == "У": t.d19c = "#ffff00"
                            if t.d19 == "В": t.d19c = "#fa5555"
                        t.save()
                    if t.d20:
                        try:
                            d = int(t.d20)
                            if t.d20 > 10:
                                t.d20c = "#00b300"
                        except:
                            if t.d20 == "ВВ": t.d20c = "#f2e93f"
                            if t.d20 == "Н": t.d20c = "#808080"
                            if t.d20 == "М": t.d20c = "#ffa812"
                            if t.d20 == "Б" or t.d20 == "О" or t.d20 == "Д" or t.d20 == "У": t.d20c = "#ffff00"
                            if t.d20 == "В": t.d20c = "#fa5555"
                        t.save()
                    if t.d21:
                        try:
                            d = int(t.d21)
                            if t.d21 > 10:
                                t.d21c = "#00b300"
                        except:
                            if t.d21 == "ВВ": t.d21c = "#f2e93f"
                            if t.d21 == "Н": t.d21c = "#808080"
                            if t.d21 == "М": t.d21c = "#ffa812"
                            if t.d21 == "Б" or t.d21 == "О" or t.d21 == "Д" or t.d21 == "У": t.d21c = "#ffff00"
                            if t.d21 == "В": t.d21c = "#fa5555"
                        t.save()
                    if t.d22:
                        try:
                            d = int(t.d22)
                            if t.d22 > 10:
                                t.d22c = "#00b300"
                        except:
                            if t.d22 == "ВВ": t.d22c = "#f2e93f"
                            if t.d22 == "Н": t.d22c = "#808080"
                            if t.d22 == "М": t.d22c = "#ffa812"
                            if t.d22 == "Б" or t.d22 == "О" or t.d22 == "Д" or t.d22 == "У": t.d22c = "#ffff00"
                            if t.d22 == "В": t.d22c = "#fa5555"
                        t.save()
                    if t.d23:
                        try:
                            d = int(t.d23)
                            if t.d23 > 10:
                                t.d23c = "#00b300"
                        except:
                            if t.d23 == "ВВ": t.d23c = "#f2e93f"
                            if t.d23 == "Н": t.d23c = "#808080"
                            if t.d23 == "М": t.d23c = "#ffa812"
                            if t.d23 == "Б" or t.d23 == "О" or t.d23 == "Д" or t.d23 == "У": t.d23c = "#ffff00"
                            if t.d23 == "В": t.d23c = "#fa5555"
                        t.save()
                    if t.d24:
                        try:
                            d = int(t.d24)
                            if t.d24 > 10:
                                t.d24c = "#00b300"
                        except:
                            if t.d24 == "ВВ": t.d24c = "#f2e93f"
                            if t.d24 == "Н": t.d24c = "#808080"
                            if t.d24 == "М": t.d24c = "#ffa812"
                            if t.d24 == "Б" or t.d24 == "О" or t.d24 == "Д" or t.d24 == "У": t.d24c = "#ffff00"
                            if t.d24 == "В": t.d24c = "#fa5555"
                        t.save()
                    if t.d25:
                        try:
                            d = int(t.d25)
                            if t.d25 > 10:
                                t.d25c = "#00b300"
                        except:
                            if t.d25 == "ВВ": t.d25c = "#f2e93f"
                            if t.d25 == "Н": t.d25c = "#808080"
                            if t.d25 == "М": t.d25c = "#ffa812"
                            if t.d25 == "Б" or t.d25 == "О" or t.d25 == "Д" or t.d25 == "У": t.d25c = "#ffff00"
                            if t.d25 == "В": t.d25c = "#fa5555"
                        t.save()
                    if t.d26:
                        try:
                            d = int(t.d26)
                            if t.d26 > 10:
                                t.d26c = "#00b300"
                        except:
                            if t.d26 == "ВВ": t.d26c = "#f2e93f"
                            if t.d26 == "Н": t.d26c = "#808080"
                            if t.d26 == "М": t.d26c = "#ffa812"
                            if t.d26 == "Б" or t.d26 == "О" or t.d26 == "Д" or t.d26 == "У": t.d26c = "#ffff00"
                            if t.d26 == "В": t.d26c = "#fa5555"
                        t.save()
                    if t.d27:
                        try:
                            d = int(t.d27)
                            if t.d27 > 10:
                                t.d27c = "#00b300"
                        except:
                            if t.d27 == "ВВ": t.d27c = "#f2e93f"
                            if t.d27 == "Н": t.d27c = "#808080"
                            if t.d27 == "М": t.d27c = "#ffa812"
                            if t.d27 == "Б" or t.d27 == "О" or t.d27 == "Д" or t.d27 == "У": t.d27c = "#ffff00"
                            if t.d27 == "В": t.d27c = "#fa5555"
                        t.save()
                    if t.d28:
                        try:
                            d = int(t.d28)
                            if t.d28 > 10:
                                t.d28c = "#00b300"
                        except:
                            if t.d28 == "ВВ": t.d28c = "#f2e93f"
                            if t.d28 == "Н": t.d28c = "#808080"
                            if t.d28 == "М": t.d28c = "#ffa812"
                            if t.d28 == "Б" or t.d28 == "О" or t.d28 == "Д" or t.d28 == "У": t.d28c = "#ffff00"
                            if t.d28 == "В": t.d28c = "#fa5555"
                        t.save()
                    if t.d29:
                        try:
                            d = int(t.d29)
                            if t.d29 > 10:
                                t.d29c = "#00b300"
                        except:
                            if t.d29 == "ВВ": t.d29c = "#f2e93f"
                            if t.d29 == "Н": t.d29c = "#808080"
                            if t.d29 == "М": t.d29c = "#ffa812"
                            if t.d29 == "Б" or t.d29 == "О" or t.d29 == "Д" or t.d29 == "У": t.d29c = "#ffff00"
                            if t.d29 == "В": t.d29c = "#fa5555"
                        t.save()
                    if t.d30:
                        try:
                            d = int(t.d30)
                            if t.d30 > 10:
                                t.d30c = "#00b300"
                        except:
                            if t.d30 == "ВВ": t.d30c = "#f2e93f"
                            if t.d30 == "Н": t.d30c = "#808080"
                            if t.d30 == "М": t.d30c = "#ffa812"
                            if t.d30 == "Б" or t.d30 == "О" or t.d30 == "Д" or t.d30 == "У": t.d30c = "#ffff00"
                            if t.d30 == "В": t.d30c = "#fa5555"
                        t.save()
                    if t.d31:
                        try:
                            d = int(t.d31)
                            if t.d31 > 10:
                                t.d31c = "#00b300"
                        except:
                            if t.d31 == "ВВ": t.d31c = "#f2e93f"
                            if t.d31 == "Н": t.d31c = "#808080"
                            if t.d31 == "М": t.d31c = "#ffa812"
                            if t.d31 == "Б" or t.d31 == "О" or t.d31 == "Д" or t.d31 == "У": t.d31c = "#ffff00"
                            if t.d31 == "В": t.d31c = "#fa5555"
                        t.save()
                if t:
                    if t.d1:
                        t.auto_block = 1
                    if t.d2:
                        t.auto_block = 1
                    if t.d3:
                        t.auto_block = 1
                    if t.d4:
                        t.auto_block = 1
                    if t.d5:
                        t.auto_block = 1
                    if t.d6:
                        t.auto_block = 1
                    if t.d7:
                        t.auto_block = 1
                    if t.d8:
                        t.auto_block = 1
                    if t.d9:
                        t.auto_block = 1
                    if t.d10:
                        t.auto_block = 1
                    if t.d11:
                        t.auto_block = 1
                    if t.d12:
                        t.auto_block = 1
                    if t.d13:
                        t.auto_block = 1
                    if t.d14:
                        t.auto_block = 1
                    if t.d15:
                        t.auto_block = 1
                    if t.d16:
                        t.auto_block = 1
                    if t.d17:
                        t.auto_block = 1
                    if t.d18:
                        t.auto_block = 1
                    if t.d19:
                        t.auto_block = 1
                    if t.d20:
                        t.auto_block = 1
                    if t.d21:
                        t.auto_block = 1
                    if t.d22:
                        t.auto_block = 1
                    if t.d23:
                        t.auto_block = 1
                    if t.d24:
                        t.auto_block = 1
                    if t.d25:
                        t.auto_block = 1
                    if t.d26:
                        t.auto_block = 1
                    if t.d27:
                        t.auto_block = 1
                    if t.d28:
                        t.auto_block = 1
                    if t.d29:
                        t.auto_block = 1
                    if t.d30:
                        t.auto_block = 1
                    if t.d31:
                        t.auto_block = 1
                    t.save()
                h = 0
                d = 0
                try:
                    try:
                        h1 = int(t.d1)
                        h = h + h1
                        d += 1
                    except:
                        if t.d1 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h2 = int(t.d2)
                        h = h + h2
                        d += 1
                    except:
                        if t.d2 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h3 = int(t.d3)
                        h = h + h3
                        d += 1
                    except:
                        if t.d3 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h4 = int(t.d4)
                        h = h + h4
                        d += 1
                    except:
                        if t.d4 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h5 = int(t.d5)
                        h = h + h5
                        d += 1
                    except:
                        if t.d5 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h6 = int(t.d6)
                        h = h + h6
                        d += 1
                    except:
                        if t.d6 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h7 = int(t.d7)
                        h = h + h7
                        d += 1
                    except:
                        if t.d7 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h8 = int(t.d8)
                        h = h + h8
                        d += 1
                    except:
                        if t.d8 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h9 = int(t.d9)
                        h = h + h9
                        d += 1
                    except:
                        if t.d9 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h10 = int(t.d10)
                        h = h + h10
                        d += 1
                    except:
                        if t.d10 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h11 = int(t.d11)
                        h = h + h11
                        d += 1
                    except:
                        if t.d11 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h12 = int(t.d12)
                        h = h + h12
                        d += 1
                    except:
                        if t.d12 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h13 = int(t.d13)
                        h = h + h13
                        d += 1
                    except:
                        if t.d13 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h14 = int(t.d14)
                        h = h + h14
                        d += 1
                    except:
                        if t.d14 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h15 = int(t.d15)
                        h = h + h15
                        d += 1
                    except:
                        if t.d15 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h16 = int(t.d16)
                        h = h + h16
                        d += 1
                    except:
                        if t.d16 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h17 = int(t.d17)
                        h = h + h17
                        d += 1
                    except:
                        if t.d17 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h18 = int(t.d18)
                        h = h + h18
                        d += 1
                    except:
                        if t.d18 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h19 = int(t.d19)
                        h = h + h19
                        d += 1
                    except:
                        if t.d19 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h20 = int(t.d20)
                        h = h + h20
                        d += 1
                    except:
                        if t.d20 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h21 = int(t.d21)
                        h = h + h21
                        d += 1
                    except:
                        if t.d21 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h22 = int(t.d22)
                        h = h + h22
                        d += 1
                    except:
                        if t.d22 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h23 = int(t.d23)
                        h = h + h23
                        d += 1
                    except:
                        if t.d23 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h24 = int(t.d24)
                        h = h + h24
                        d += 1
                    except:
                        if t.d24 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h25 = int(t.d25)
                        h = h + h25
                        d += 1
                    except:
                        if t.d25 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h26 = int(t.d26)
                        h = h + h26
                        d += 1
                    except:
                        if t.d26 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h27 = int(t.d27)
                        h = h + h27
                        d += 1
                    except:
                        if t.d27 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h28 = int(t.d28)
                        h = h + h28
                        d += 1
                    except:
                        if t.d28 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h29 = int(t.d29)
                        h = h + h29
                        d += 1
                    except:
                        if t.d29 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h30 = int(t.d30)
                        h = h + h30
                        d += 1
                    except:
                        if t.d30 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h31 = int(t.d31)
                        h = h + h31
                        d += 1
                    except:
                        if t.d31 == "ВВ":
                            h = h + int(t.persone.work_hours)
                except:
                    pass
                t.hours = int(h)
                t.days = int(d)
                t.save()
                auto = request.POST.get(str(t.id) + 'auto')
                try:
                    t.auto = automobile.objects.get(id=int(auto))
                except:
                    pass
                comment = request.POST.get("com" + str(t.id))
                if comment == '':
                    pass
                else:
                    if t.comment == None:
                        t.comment = ''
                        t.save()
                        comment = str(t.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                        t.comment = comment
                        t.save()
                    else:
                        comment = str(t.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                        t.comment = comment
                        t.save()

                try:
                    fine = request.POST.get(str(t.id)+"fine")
                    if fine:
                        t.fine = float(fine)
                    else:
                        t.fine = 0
                except:
                    pass
                try:
                    add = request.POST.get(str(t.id)+"add")
                    if add:
                        t.add = float(add)
                    else:
                        t.add = 0
                except:
                    pass
                hour = 300
                f = t.fine
                a = t.add
                t.sum = int(t.persone.salary)/hour * int(t.hours)
                t.salary = int(t.sum) - float(f) + float(a)
                t.save()
            return redirect('/view_driver/' + str(id))
        else:
            context = {
                'autos': autos,
                'persones': persones,
                'per': per,
                'segment': 'driver',
                'tabs': tabs,
                'pers':pers,
                'all':all,
                'approvers': approvers,
            }
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def del_driverl(request, id, ids):
        tl = driver_list.objects.get(id=id)
        tl.delete()
        return redirect('/view_driver/' + str(ids))

    @login_required(login_url="/login/")
    def tabel(request):
        per = period.objects.all()
        groups = tabel_group.objects.all()
        if groups:
            pass
        else:
            groups
        context = {'per': per, 'segment': 'tabel', 'groups': groups}
        html_template = loader.get_template('tabels/worker/tabel.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_tabel(request):
        per = period.objects.all()
        form = TabForm
        persones = persone.objects.all().filter(date_leave=None)
        tl = tabel_list.objects.all()
        if request.method == 'POST':
            forms = TabForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                year =request.POST.get("ye")
                month = request.POST.get("mo")
                da = year + '-' + month + '-' + '01'
                post.date = da
                try:
                    gr = tabel_group.objects.get(date=post.date)
                except:
                    tabel_group(date=post.date).save()
                    gr = tabel_group.objects.get(date=post.date)
                post.group = gr
                da = date(int(year),int(month), 1)
                if post.type != None:
                    if post.graphic == 1:
                        days = monthrange(int(year),int(month))[1]
                        post.workdays = int(days)
                    post.save()
                    try:
                        da = da.replace(day=1)
                        post.dw1 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=2)
                        post.dw2 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=3)
                        post.dw3 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=4)
                        post.dw4 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=5)
                        post.dw5 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=6)
                        post.dw6 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=7)
                        post.dw7 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=8)
                        post.dw8 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=9)
                        post.dw9 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=10)
                        post.dw10 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=11)
                        post.dw11 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=12)
                        post.dw12 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=13)
                        post.dw13 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=14)
                        post.dw14 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=15)
                        post.dw15 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=16)
                        post.dw16 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=17)
                        post.dw17 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=18)
                        post.dw18 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=19)
                        post.dw19 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=20)
                        post.dw20 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=21)
                        post.dw21 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=22)
                        post.dw22 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=23)
                        post.dw23 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=24)
                        post.dw24 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=25)
                        post.dw25 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=26)
                        post.dw26 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=27)
                        post.dw27 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=28)
                        post.dw28 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=29)
                        post.dw29 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=30)
                        post.dw30 = da
                    except:
                        pass
                    try:
                        da = da.replace(day=31)
                        post.dw31 = da
                    except:
                        pass
                    post.save()
                    return redirect('tabel')
                else:
                    error = 'Месяц или Год не выбраны'
                    context = {
                        'error': error,
                        'segment': 'tabel',
                        'per': per,
                        'form': form,
                    }
                    html_template = loader.get_template('tabels/worker/new_tabel.html')
                    return HttpResponse(html_template.render(context, request))
        else:
            context = {
                'segment': 'tabel',
                'persones':persones,
                'per': per,
                'form': form,
            }
            html_template = loader.get_template('tabels/worker/new_tabel.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def del_tabl(request, id, ids):
        tl = tabel_list.objects.get(id=id)
        tl.delete()
        return redirect('/view_tabel/' + str(ids))

    @login_required(login_url="/login/")
    def move_worker(request, id, ids, gid):
        g = group.objects.get(id=gid)
        tl = tabel_list.objects.get(id=id)
        tl.group = g
        tl.save()
        return redirect('/view_tabel/' + str(ids))

    @login_required(login_url="/login/")
    def AddPersone(request,id,ids):
        per = period.objects.get(id=ids)
        p = persone.objects.get(id=id)
        try:
            g = group.objects.get(id=p.group.id)
            if g:
                tabel_list(
                    period=per,
                    persone=p,
                    company=p.company,
                    snils=p.snils,
                    full_name=p.full_name,
                    group=g,
                    position=p.position,
                    salary=p.salary,
                    output=p.output,
                ).save()
            else:
                g = group.objects.get(name='None')
                tabel_list(
                    period=per,
                    persone=p,
                    company=p.company,
                    snils=p.snils,
                    full_name=p.full_name,
                    group=g,
                    position=p.position,
                    salary=p.salary,
                    output=p.output,
                ).save()
        except:
            g = group.objects.get(name='None')
            tabel_list(
                period=per,
                persone=p,
                company=p.company,
                snils=p.snils,
                full_name=p.full_name,
                group=g,
                position=p.position,
                salary=p.salary,
                output=p.output,
            ).save()
        return redirect('/view_tabel/' + str(ids))

    @login_required(login_url="/login/")
    def view_tabel(request, id):
        groups = group.objects.all()
        per = period.objects.get(id=id)
        pers = period.objects.all()
        user = request.user
        approvers = dayapprove.objects.all().filter(per=per)
        persones = persone.objects.all().filter().order_by('full_name')
        tabs = tabel_list.objects.all().filter(period=per).order_by('full_name')
        html_template = loader.get_template('tabels/worker/view_tabel.html')
        if per.date:
            month = datetime.date.today()
            year = str(month)[0:4]
            month = str(month)[5:7]
            pm = per.date
            py = str(pm)[0:4]
            pm = str(pm)[5:7]
            if int(month)>int(pm):
                if int(year)==int(py):
                    per.fullblock = 1;
                    per.save()
            if per.fullblock == 1:
                pass
                today = 32
            else:
                today = datetime.date.today()
                today = str(today)[8:]
                today = int(today) - 2
            if int(month)>=int(pm):
                if today > 1:
                    per.db1 = 1
                    per.save()
                if today > 2:
                    per.db2 = 1
                    per.save()
                if today > 3:
                    per.db3 = 1
                    per.save()
                if today > 4:
                    per.db4 = 1
                    per.save()
                if today > 5:
                    per.db5 = 1
                    per.save()
                if today > 6:
                    per.db6 = 1
                    per.save()
                if today > 7:
                    per.db7 = 1
                    per.save()
                if today > 8:
                    per.db8 = 1
                    per.save()
                if today > 9:
                    per.db9 = 1
                    per.save()
                if today > 10:
                    per.db10 = 1
                    per.save()
                if today > 11:
                    per.db11 = 1
                    per.save()
                if today > 12:
                    per.db12 = 1
                    per.save()
                if today > 13:
                    per.db13 = 1
                    per.save()
                if today > 14:
                    per.db14 = 1
                    per.save()
                if today > 15:
                    per.db15 = 1
                    per.save()
                if today > 16:
                    per.db16 = 1
                    per.save()
                if today > 17:
                    per.db17 = 1
                    per.save()
                if today > 18:
                    per.db18 = 1
                    per.save()
                if today > 19:
                    per.db19 = 1
                    per.save()
                if today > 20:
                    per.db20 = 1
                    per.save()
                if today > 21:
                    per.db21 = 1
                    per.save()
                if today > 22:
                    per.db22 = 1
                    per.save()
                if today > 23:
                    per.db23 = 1
                    per.save()
                if today > 24:
                    per.db24 = 1
                    per.save()
                if today > 25:
                    per.db25 = 1
                    per.save()
                if today > 26:
                    per.db26 = 1
                    per.save()
                if today > 27:
                    per.db27 = 1
                    per.save()
                if today > 28:
                    per.db28 = 1
                    per.save()
                if today > 29:
                    per.db29 = 1
                    per.save()
                if today > 30:
                    per.db30 = 1
                    per.save()
                if today > 31:
                    per.db31 = 1
                    per.save()
            if per:
                i1 = 0
                i2 = 0
                i3 = 0
                i4 = 0
                i5 = 0
                i6 = 0
                i7 = 0
                i8 = 0
                i9 = 0
                i10 = 0
                i11 = 0
                i12 = 0
                i13 = 0
                i14 = 0
                i15 = 0
                i16 = 0
                i17 = 0
                i18 = 0
                i19 = 0
                i20 = 0
                i21 = 0
                i22 = 0
                i23 = 0
                i24 = 0
                i25 = 0
                i26 = 0
                i27 = 0
                i28 = 0
                i29 = 0
                i30 = 0
                i31 = 0
            for a in approvers:
                if a.day == 1:
                    i1 += 1
                if i1 >= 2:
                    per.da1 = 1
                if a.day == 2:
                    i2 += 1
                if i2 >= 2:
                    per.da2 = 1
                if a.day == 3:
                    i3 += 1
                if i3 >= 2:
                    per.da3 = 1
                if a.day == 4:
                    i4 += 1
                if i4 >= 2:
                    per.da4 = 1
                if a.day == 5:
                    i5 += 1
                if i5 >= 2:
                    per.da5 = 1
                if a.day == 6:
                    i6 += 1
                if i6 >= 2:
                    per.da6 = 1
                if a.day == 7:
                    i7 += 1
                if i7 >= 2:
                    per.da7 = 1
                if a.day == 8:
                    i8 += 1
                if i8 >= 2:
                    per.da8 = 1
                if a.day == 9:
                    i9 += 1
                if i9 >= 2:
                    per.da9 = 1
                if a.day == 10:
                    i10 += 1
                if i10 >= 2:
                    per.da10 = 1
                if a.day == 11:
                    i11 += 1
                if i11 >= 2:
                    per.da11 = 1
                if a.day == 12:
                    i12 += 1
                if i12 >= 2:
                    per.da12 = 1
                if a.day == 13:
                    i13 += 1
                if i13 >= 2:
                    per.da13 = 1
                if a.day == 14:
                    i14 += 1
                if i14 >= 2:
                    per.da14 = 1
                if a.day == 15:
                    i15 += 1
                if i15 >= 2:
                    per.da15 = 1
                if a.day == 16:
                    i16 += 1
                if i16 >= 2:
                    per.da16 = 1
                if a.day == 17:
                    i17 += 1
                if i17 >= 2:
                    per.da17 = 1
                if a.day == 18:
                    i18 += 1
                if i18 >= 2:
                    per.da18 = 1
                if a.day == 19:
                    i19 += 1
                if i19 >= 2:
                    per.da19 = 1
                if a.day == 20:
                    i20 += 1
                if i20 >= 2:
                    per.da20 = 1
                if a.day == 21:
                    i21 += 1
                if i21 >= 2:
                    per.da21 = 1
                if a.day == 22:
                    i22 += 1
                if i22 >= 2:
                    per.da22 = 1
                if a.day == 23:
                    i23 += 1
                if i23 >= 2:
                    per.da23 = 1
                if a.day == 24:
                    i24 += 1
                if i24 >= 2:
                    per.da24 = 1
                if a.day == 25:
                    i25 += 1
                if i25 >= 2:
                    per.da25 = 1
                if a.day == 26:
                    i26 += 1
                if i26 >= 2:
                    per.da26 = 1
                if a.day == 27:
                    i27 += 1
                if i27 >= 2:
                    per.da27 = 1
                if a.day == 28:
                    i28 += 1
                if i28 >= 2:
                    per.da28 = 1
                if a.day == 29:
                    i29 += 1
                if i29 >= 2:
                    per.da29 = 1
                if a.day == 30:
                    i30 += 1
                if i30 >= 2:
                    per.da30 = 1
                if a.day == 31:
                    i31 += 1
                if i31 >= 2:
                    per.da31 = 1
                per.save()
        if request.method == 'POST':
            for t in tabs:
                da1 = request.POST.get(str(t.id)+'d1');
                da2 = request.POST.get(str(t.id)+'d2');
                da3 = request.POST.get(str(t.id)+'d3');
                da4 = request.POST.get(str(t.id)+'d4');
                da5 = request.POST.get(str(t.id)+'d5');
                da6 = request.POST.get(str(t.id)+'d6');
                da7 = request.POST.get(str(t.id)+'d7');
                da8 = request.POST.get(str(t.id)+'d8');
                da9 = request.POST.get(str(t.id)+'d9');
                da10 = request.POST.get(str(t.id)+'d10');
                da11 = request.POST.get(str(t.id)+'d11');
                da12 = request.POST.get(str(t.id)+'d12');
                da13 = request.POST.get(str(t.id)+'d13');
                da14 = request.POST.get(str(t.id)+'d14');
                da15 = request.POST.get(str(t.id)+'d15');
                da16 = request.POST.get(str(t.id)+'d16');
                da17 = request.POST.get(str(t.id)+'d17');
                da18 = request.POST.get(str(t.id)+'d18');
                da19 = request.POST.get(str(t.id)+'d19');
                da20 = request.POST.get(str(t.id)+'d20');
                da21 = request.POST.get(str(t.id)+'d21');
                da22 = request.POST.get(str(t.id)+'d22');
                da23 = request.POST.get(str(t.id)+'d23');
                da24 = request.POST.get(str(t.id)+'d24');
                da25 = request.POST.get(str(t.id)+'d25');
                da26 = request.POST.get(str(t.id)+'d26');
                da27 = request.POST.get(str(t.id)+'d27');
                da28 = request.POST.get(str(t.id)+'d28');
                da29 = request.POST.get(str(t.id)+'d29');
                da30 = request.POST.get(str(t.id)+'d30');
                da31 = request.POST.get(str(t.id)+'d31');
                # if t:
                #     try:
                #         check_per = period.objects.all().filter(date=per.date)
                #         for cp in check_per:
                #             if int(cp.id) != int(id):
                #                 check_tl = tabel_list.objects.all().filter(period=cp)
                #                 for ct in check_tl:
                #                     if ct.d1:
                #                         if da1:
                #                             error_d = 1
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d2:
                #                         if da2:
                #                             error_d = 2
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d3:
                #                         if da3:
                #                             error_d = 3
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d4:
                #                         if da4:
                #                             error_d = 4
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d5:
                #                         if da5:
                #                             error_d = 5
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d6:
                #                         if da6:
                #                             error_d = 6
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d7:
                #                         if da7:
                #                             error_d = 7
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d8:
                #                         if da8:
                #                             error_d = 8
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d9:
                #                         if da9:
                #                             error_d = 9
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d10:
                #                         if da10:
                #                             error_d = 10
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d11:
                #                         if da11:
                #                             error_d = 11
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d12:
                #                         if da12:
                #                             error_d = 12
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d13:
                #                         if da13:
                #                             error_d = 13
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d14:
                #                         if da14:
                #                             error_d = 14
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d15:
                #                         if da15:
                #                             error_d = 15
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d16:
                #                         if da16:
                #                             error_d = 16
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d17:
                #                         if da17:
                #                             error_d = 17
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d18:
                #                         if da18:
                #                             error_d = 18
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d19:
                #                         if da19:
                #                             error_d = 19
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d20:
                #                         if da20:
                #                             error_d = 20
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d21:
                #                         if da21:
                #                             error_d = 21
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d22:
                #                         if da22:
                #                             error_d = 22
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d23:
                #                         if da23:
                #                             error_d = 23
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d24:
                #                         if da24:
                #                             error_d = 24
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d25:
                #                         if da25:
                #                             error_d = 25
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d26:
                #                         if da26:
                #                             error_d = 26
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d27:
                #                         if da27:
                #                             error_d = 27
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d28:
                #                         if da28:
                #                             error_d = 28
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d29:
                #                         if da29:
                #                             error_d = 29
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d30:
                #                         if da30:
                #                             error_d = 30
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if ct.d31:
                #                         if da31:
                #                             error_d = 31
                #                             error_p = ct.persone
                #                             context = {
                #                                 'user': user,
                #                                 'persones': persones,
                #                                 'groups': groups,
                #                                 'per': per,
                #                                 'segment': 'tabel',
                #                                 'tabs': tabs,
                #                                 'pers': pers,
                #                                 'all': all,
                #                                 'approvers': approvers,
                #                                 'error_d': error_d,
                #                                 'error_p': error_p,
                #                             }
                #                             return HttpResponse(html_template.render(context, request))
                #                     if error_d:
                #                         context = {
                #                                 'user': user,
                #                             'persones': persones,
                #                             'groups': groups,
                #                             'per': per,
                #                             'segment': 'tabel',
                #                             'tabs': tabs,
                #                             'pers':pers,
                #                             'all': all,
                #                             'approvers':approvers,
                #                             'error_d': error_d,
                #                             'error_p': error_p,
                #                         }
                #                         return HttpResponse(html_template.render(context, request))
                #     except:
                #         pass
                if t:
                    if da1:
                        try:
                            da1 = int(da1)
                            if da1 > 0 and da1 <= 23:
                                t.d1 = da1
                        except:
                            da1 = da1.upper()
                            if da1 == "ВВ" or da1 == "Н" or da1 == "М" or da1 == "Б" or da1 == "О" or da1 == "Д" or da1 == "В" or da1 == "У":
                                t.d1 = da1
                                t.save()
                        t.save()
                    if da2:
                        try:
                            da2 = int(da2)
                            if da2 > 0 and da2 <= 23:
                                t.d2 = da2
                        except:
                            da2 = da2.upper()
                            if da2 == "ВВ" or da2 == "Н" or da2 == "М" or da2 == "Б" or da2 == "О" or da2 == "Д" or da2 == "В" or da2 == "У":
                                t.d2 = da2
                                t.save()
                        t.save()
                    if da3:
                        try:
                            da3 = int(da3)
                            if da3 > 0 and da3 <= 23:
                                t.d3 = da3
                        except:
                            da3 = da3.upper()
                            if da3 == "ВВ" or da3 == "Н" or da3 == "М" or da3 == "Б" or da3 == "О" or da3 == "Д" or da3 == "В" or da3 == "У":
                                t.d3 = da3
                                t.save()
                        t.save()
                    if da4:
                        try:
                            da4 = int(da4)
                            if da4 > 0 and da4 <= 23:
                                t.d4 = da4
                        except:
                            da4 = da4.upper()
                            if da4 == "ВВ" or da4 == "Н" or da4 == "М" or da4 == "Б" or da4 == "О" or da4 == "Д" or da4 == "В" or da4 == "У":
                                t.d4 = da4
                                t.save()
                        t.save()
                    if da5:
                        try:
                            da5 = int(da5)
                            if da5 > 0 and da5 <= 23:
                                t.d5 = da5
                        except:
                            da5 = da5.upper()
                            if da5 == "ВВ" or da5 == "Н" or da5 == "М" or da5 == "Б" or da5 == "О" or da5 == "Д" or da5 == "В" or da5 == "У":
                                t.d5 = da5
                                t.save()
                        t.save()
                    if da6:
                        try:
                            da6 = int(da6)
                            if da6 > 0 and da6 <= 23:
                                t.d6 = da6
                        except:
                            da6 = da6.upper()
                            if da6 == "ВВ" or da6 == "Н" or da6 == "М" or da6 == "Б" or da6 == "О" or da6 == "Д" or da6 == "В" or da6 == "У":
                                t.d6 = da6
                                t.save()
                        t.save()
                    if da7:
                        try:
                            da7 = int(da7)
                            if da7 > 0 and da7 <= 23:
                                t.d7 = da7
                        except:
                            da7 = da7.upper()
                            if da7 == "ВВ" or da7 == "Н" or da7 == "М" or da7 == "Б" or da7 == "О" or da7 == "Д" or da7 == "В" or da7 == "У":
                                t.d7 = da7
                                t.save()
                        t.save()
                    if da8:
                        try:
                            da8 = int(da8)
                            if da8 > 0 and da8 <= 23:
                                t.d8 = da8
                        except:
                            da8 = da8.upper()
                            if da8 == "ВВ" or da8 == "Н" or da8 == "М" or da8 == "Б" or da8 == "О" or da8 == "Д" or da8 == "В" or da8 == "У":
                                t.d8 = da8
                                t.save()
                        t.save()
                    if da9:
                        try:
                            da9 = int(da9)
                            if da9 > 0 and da9 <= 23:
                                t.d9 = da9
                        except:
                            da9 = da9.upper()
                            if da9 == "ВВ" or da9 == "Н" or da9 == "М" or da9 == "Б" or da9 == "О" or da9 == "Д" or da9 == "В" or da9 == "У":
                                t.d9 = da9
                                t.save()
                        t.save()
                    if da10:
                        try:
                            da10 = int(da10)
                            if da10 > 0 and da10 <= 23:
                                t.d10 = da10
                        except:
                            da10 = da10.upper()
                            if da10 == "ВВ" or da10 == "Н" or da10 == "М" or da10 == "Б" or da10 == "О" or da10 == "Д" or da10 == "В" or da10 == "У":
                                t.d10 = da10
                                t.save()
                        t.save()
                    if da11:
                        try:
                            da11 = int(da11)
                            if da11 > 0 and da11 <= 23:
                                t.d11 = da11
                        except:
                            da11 = da11.upper()
                            if da11 == "ВВ" or da11 == "Н" or da11 == "М" or da11 == "Б" or da11 == "О" or da11 == "Д" or da11 == "В" or da11 == "У":
                                t.d11 = da11
                                t.save()
                        t.save()
                    if da12:
                        try:
                            da12 = int(da12)
                            if da12 > 0 and da12 <= 23:
                                t.d12 = da12
                        except:
                            da12 = da12.upper()
                            if da12 == "ВВ" or da12 == "Н" or da12 == "М" or da12 == "Б" or da12 == "О" or da12 == "Д" or da12 == "В" or da12 == "У":
                                t.d12 = da12
                                t.save()
                        t.save()
                    if da13:
                        try:
                            da13 = int(da13)
                            if da13 > 0 and da13 <= 23:
                                t.d13 = da13
                        except:
                            da13 = da13.upper()
                            if da13 == "ВВ" or da13 == "Н" or da13 == "М" or da13 == "Б" or da13 == "О" or da13 == "Д" or da13 == "В" or da13 == "У":
                                t.d13 = da13
                                t.save()
                        t.save()
                    if da14:
                        try:
                            da14 = int(da14)
                            if da14 > 0 and da14 <= 23:
                                t.d14 = da14
                        except:
                            da14 = da14.upper()
                            if da14 == "ВВ" or da14 == "Н" or da14 == "М" or da14 == "Б" or da14 == "О" or da14 == "Д" or da14 == "В" or da14 == "У":
                                t.d14 = da14
                                t.save()
                        t.save()
                    if da15:
                        try:
                            da15 = int(da15)
                            if da15 > 0 and da15 <= 23:
                                t.d15 = da15
                        except:
                            da15 = da15.upper()
                            if da15 == "ВВ" or da15 == "Н" or da15 == "М" or da15 == "Б" or da15 == "О" or da15 == "Д" or da15 == "В" or da15 == "У":
                                t.d15 = da15
                                t.save()
                        t.save()
                    if da16:
                        try:
                            da16 = int(da16)
                            if da16 > 0 and da16 <= 23:
                                t.d16 = da16
                        except:
                            da16 = da16.upper()
                            if da16 == "ВВ" or da16 == "Н" or da16 == "М" or da16 == "Б" or da16 == "О" or da16 == "Д" or da16 == "В" or da16 == "У":
                                t.d16 = da16
                                t.save()
                        t.save()
                    if da17:
                        try:
                            da17 = int(da17)
                            if da17 > 0 and da17 <= 23:
                                t.d17 = da17
                        except:
                            da17 = da17.upper()
                            if da17 == "ВВ" or da17 == "Н" or da17 == "М" or da17 == "Б" or da17 == "О" or da17 == "Д" or da17 == "В" or da17 == "У":
                                t.d17 = da17
                                t.save()
                        t.save()
                    if da18:
                        try:
                            da18 = int(da18)
                            if da18 > 0 and da18 <= 23:
                                t.d18 = da18
                        except:
                            da18 = da18.upper()
                            if da18 == "ВВ" or da18 == "Н" or da18 == "М" or da18 == "Б" or da18 == "О" or da18 == "Д" or da18 == "В" or da18 == "У":
                                t.d18 = da18
                                t.save()
                        t.save()
                    if da19:
                        try:
                            da19 = int(da19)
                            if da19 > 0 and da19 <= 23:
                                t.d19 = da19
                        except:
                            da19 = da19.upper()
                            if da19 == "ВВ" or da19 == "Н" or da19 == "М" or da19 == "Б" or da19 == "О" or da19 == "Д" or da19 == "В" or da19 == "У":
                                t.d19 = da19
                                t.save()
                        t.save()
                    if da20:
                        try:
                            da20 = int(da20)
                            if da20 > 0 and da20 <= 23:
                                t.d20 = da20
                        except:
                            da20 = da20.upper()
                            if da20 == "ВВ" or da20 == "Н" or da20 == "М" or da20 == "Б" or da20 == "О" or da20 == "Д" or da20 == "В" or da20 == "У":
                                t.d20 = da20
                                t.save()
                        t.save()
                    if da21:
                        try:
                            da21 = int(da21)
                            if da21 > 0 and da21 <= 23:
                                t.d21 = da21
                        except:
                            da21 = da21.upper()
                            if da21 == "ВВ" or da21 == "Н" or da21 == "М" or da21 == "Б" or da21 == "О" or da21 == "Д" or da21 == "В" or da21 == "У":
                                t.d21 = da21
                                t.save()
                        t.save()
                    if da22:
                        try:
                            da22 = int(da22)
                            if da22 > 0 and da22 <= 23:
                                t.d22 = da22
                        except:
                            da22 = da22.upper()
                            if da22 == "ВВ" or da22 == "Н" or da22 == "М" or da22 == "Б" or da22 == "О" or da22 == "Д" or da22 == "В" or da22 == "У":
                                t.d22 = da22
                                t.save()
                        t.save()
                    if da23:
                        try:
                            da23 = int(da23)
                            if da23 > 0 and da23 <= 23:
                                t.d23 = da23
                        except:
                            da23 = da23.upper()
                            if da23 == "ВВ" or da23 == "Н" or da23 == "М" or da23 == "Б" or da23 == "О" or da23 == "Д" or da23 == "В" or da23 == "У":
                                t.d23 = da23
                                t.save()
                        t.save()
                    if da24:
                        try:
                            da24 = int(da24)
                            if da24 > 0 and da24 <= 23:
                                t.d24 = da24
                        except:
                            da24 = da24.upper()
                            if da24 == "ВВ" or da24 == "Н" or da24 == "М" or da24 == "Б" or da24 == "О" or da24 == "Д" or da24 == "В" or da24 == "У":
                                t.d24 = da24
                                t.save()
                        t.save()
                    if da25:
                        try:
                            da25 = int(da25)
                            if da25 > 0 and da25 <= 23:
                                t.d25 = da25
                        except:
                            da25 = da25.upper()
                            if da25 == "ВВ" or da25 == "Н" or da25 == "М" or da25 == "Б" or da25 == "О" or da25 == "Д" or da25 == "В" or da25 == "У":
                                t.d25 = da25
                                t.save()
                        t.save()
                    if da26:
                        try:
                            da26 = int(da26)
                            if da26 > 0 and da26 <= 23:
                                t.d26 = da26
                        except:
                            da26 = da26.upper()
                            if da26 == "ВВ" or da26 == "Н" or da26 == "М" or da26 == "Б" or da26 == "О" or da26 == "Д" or da26 == "В" or da26 == "У":
                                t.d26 = da26
                                t.save()
                        t.save()
                    if da27:
                        try:
                            da27 = int(da27)
                            if da27 > 0 and da27 <= 23:
                                t.d27 = da27
                        except:
                            da27 = da27.upper()
                            if da27 == "ВВ" or da27 == "Н" or da27 == "М" or da27 == "Б" or da27 == "О" or da27 == "Д" or da27 == "В" or da27 == "У":
                                t.d27 = da27
                                t.save()
                        t.save()
                    if da28:
                        try:
                            da28 = int(da28)
                            if da28 > 0 and da28 <= 23:
                                t.d28 = da28
                        except:
                            da28 = da28.upper()
                            if da28 == "ВВ" or da28 == "Н" or da28 == "М" or da28 == "Б" or da28 == "О" or da28 == "Д" or da28 == "В" or da28 == "У":
                                t.d28 = da28
                                t.save()
                        t.save()
                    if da29:
                        try:
                            da29 = int(da29)
                            if da29 > 0 and da29 <= 23:
                                t.d29 = da29
                        except:
                            da29 = da29.upper()
                            if da29 == "ВВ" or da29 == "Н" or da29 == "М" or da29 == "Б" or da29 == "О" or da29 == "Д" or da29 == "В" or da29 == "У":
                                t.d29 = da29
                                t.save()
                        t.save()
                    if da30:
                        try:
                            da30 = int(da30)
                            if da30 > 0 and da30 <= 23:
                                t.d30 = da30
                        except:
                            da30 = da30.upper()
                            if da30 == "ВВ" or da30 == "Н" or da30 == "М" or da30 == "Б" or da30 == "О" or da30 == "Д" or da30 == "В" or da30 == "У":
                                t.d30 = da30
                                t.save()
                        t.save()
                    if da31:
                        try:
                            da31 = int(da31)
                            if da31 > 0 and da31 <= 23:
                                t.d31 = da31
                        except:
                            da31 = da31.upper()
                            if da31 == "ВВ" or da31 == "Н" or da31 == "М" or da31 == "Б" or da31 == "О" or da31 == "Д" or da31 == "В" or da31 == "У":
                                t.d31 = da31
                                t.save()
                        t.save()
                if t:
                    if t.d1:
                        try:
                            d = int(t.d1)
                            if t.d1 > 10:
                                t.d1c = "#00b300"
                        except:
                            if t.d1 == "ВВ": t.d1c = "#f2e93f"
                            if t.d1 == "Н": t.d1c = "#808080"
                            if t.d1 == "М": t.d1c = "#ffa812"
                            if t.d1 == "Б" or t.d1 == "О" or t.d1 == "Д" or t.d1 == "У": t.d1c = "#ffff00"
                            if t.d1 == "В": t.d1c = "#fa5555"
                        t.save()
                    if t.d2:
                        try:
                            d = int(t.d2)
                            if t.d2 > 10:
                                t.d2c = "#00b300"
                        except:
                            if t.d2 == "ВВ": t.d2c = "#f2e93f"
                            if t.d2 == "Н": t.d2c = "#808080"
                            if t.d2 == "М": t.d2c = "#ffa812"
                            if t.d2 == "Б" or t.d2 == "О" or t.d2 == "Д" or t.d2 == "У": t.d2c = "#ffff00"
                            if t.d2 == "В": t.d2c = "#fa5555"
                        t.save()
                    if t.d3:
                        try:
                            d = int(t.d3)
                            if t.d3 > 10:
                                t.d3c = "#00b300"
                        except:
                            if t.d3 == "ВВ": t.d3c = "#f2e93f"
                            if t.d3 == "Н": t.d3c = "#808080"
                            if t.d3 == "М": t.d3c = "#ffa812"
                            if t.d3 == "Б" or t.d3 == "О" or t.d3 == "Д" or t.d3 == "У": t.d3c = "#ffff00"
                            if t.d3 == "В": t.d3c = "#fa5555"
                        t.save()
                    if t.d4:
                        try:
                            d = int(t.d4)
                            if t.d4 > 10:
                                t.d4c = "#00b300"
                        except:
                            if t.d4 == "ВВ": t.d4c = "#f2e93f"
                            if t.d4 == "Н": t.d4c = "#808080"
                            if t.d4 == "М": t.d4c = "#ffa812"
                            if t.d4 == "Б" or t.d4 == "О" or t.d4 == "Д" or t.d4 == "У": t.d4c = "#ffff00"
                            if t.d4 == "В": t.d4c = "#fa5555"
                        t.save()
                    if t.d5:
                        try:
                            d = int(t.d5)
                            if t.d5 > 10:
                                t.d5c = "#00b300"
                        except:
                            if t.d5 == "ВВ": t.d5c = "#f2e93f"
                            if t.d5 == "Н": t.d5c = "#808080"
                            if t.d5 == "М": t.d5c = "#ffa812"
                            if t.d5 == "Б" or t.d5 == "О" or t.d5 == "Д" or t.d5 == "У": t.d5c = "#ffff00"
                            if t.d5 == "В": t.d5c = "#fa5555"
                        t.save()
                    if t.d6:
                        try:
                            d = int(t.d6)
                            if t.d6 > 10:
                                t.d6c = "#00b300"
                        except:
                            if t.d6 == "ВВ": t.d6c = "#f2e93f"
                            if t.d6 == "Н": t.d6c = "#808080"
                            if t.d6 == "М": t.d6c = "#ffa812"
                            if t.d6 == "Б" or t.d6 == "О" or t.d6 == "Д" or t.d6 == "У": t.d6c = "#ffff00"
                            if t.d6 == "В": t.d6c = "#fa5555"
                        t.save()
                    if t.d7:
                        try:
                            d = int(t.d7)
                            if t.d7 > 10:
                                t.d7c = "#00b300"
                        except:
                            if t.d7 == "ВВ": t.d7c = "#f2e93f"
                            if t.d7 == "Н": t.d7c = "#808080"
                            if t.d7 == "М": t.d7c = "#ffa812"
                            if t.d7 == "Б" or t.d7 == "О" or t.d7 == "Д" or t.d7 == "У": t.d7c = "#ffff00"
                            if t.d7 == "В": t.d7c = "#fa5555"
                        t.save()
                    if t.d8:
                        try:
                            d = int(t.d8)
                            if t.d8 > 10:
                                t.d8c = "#00b300"
                        except:
                            if t.d8 == "ВВ": t.d8c = "#f2e93f"
                            if t.d8 == "Н": t.d8c = "#808080"
                            if t.d8 == "М": t.d8c = "#ffa812"
                            if t.d8 == "Б" or t.d8 == "О" or t.d8 == "Д" or t.d8 == "У": t.d8c = "#ffff00"
                            if t.d8 == "В": t.d8c = "#fa5555"
                        t.save()
                    if t.d9:
                        try:
                            d = int(t.d9)
                            if t.d9 > 10:
                                t.d9c = "#00b300"
                        except:
                            if t.d9 == "ВВ": t.d9c = "#f2e93f"
                            if t.d9 == "Н": t.d9c = "#808080"
                            if t.d9 == "М": t.d9c = "#ffa812"
                            if t.d9 == "Б" or t.d9 == "О" or t.d9 == "Д" or t.d9 == "У": t.d9c = "#ffff00"
                            if t.d9 == "В": t.d9c = "#fa5555"
                        t.save()
                    if t.d10:
                        try:
                            d = int(t.d10)
                            if t.d10 > 10:
                                t.d10c = "#00b300"
                        except:
                            if t.d10 == "ВВ": t.d10c = "#f2e93f"
                            if t.d10 == "Н": t.d10c = "#808080"
                            if t.d10 == "М": t.d10c = "#ffa812"
                            if t.d10 == "Б" or t.d10 == "О" or t.d10 == "Д" or t.d10 == "У": t.d10c = "#ffff00"
                            if t.d10 == "В": t.d10c = "#fa5555"
                        t.save()
                    if t.d11:
                        try:
                            d = int(t.d11)
                            if t.d11 > 10:
                                t.d11c = "#00b300"
                        except:
                            if t.d11 == "ВВ": t.d11c = "#f2e93f"
                            if t.d11 == "Н": t.d11c = "#808080"
                            if t.d11 == "М": t.d11c = "#ffa812"
                            if t.d11 == "Б" or t.d11 == "О" or t.d11 == "Д" or t.d11 == "У": t.d11c = "#ffff00"
                            if t.d11 == "В": t.d11c = "#fa5555"
                        t.save()
                    if t.d12:
                        try:
                            d = int(t.d12)
                            if t.d12 > 10:
                                t.d12c = "#00b300"
                        except:
                            if t.d12 == "ВВ": t.d12c = "#f2e93f"
                            if t.d12 == "Н": t.d12c = "#808080"
                            if t.d12 == "М": t.d12c = "#ffa812"
                            if t.d12 == "Б" or t.d12 == "О" or t.d12 == "Д" or t.d12 == "У": t.d12c = "#ffff00"
                            if t.d12 == "В": t.d12c = "#fa5555"
                        t.save()
                    if t.d13:
                        try:
                            d = int(t.d13)
                            if t.d13 > 10:
                                t.d13c = "#00b300"
                        except:
                            if t.d13 == "ВВ": t.d13c = "#f2e93f"
                            if t.d13 == "Н": t.d13c = "#808080"
                            if t.d13 == "М": t.d13c = "#ffa812"
                            if t.d13 == "Б" or t.d13 == "О" or t.d13 == "Д" or t.d13 == "У": t.d13c = "#ffff00"
                            if t.d13 == "В": t.d13c = "#fa5555"
                        t.save()
                    if t.d14:
                        try:
                            d = int(t.d14)
                            if t.d14 > 10:
                                t.d14c = "#00b300"
                        except:
                            if t.d14 == "ВВ": t.d14c = "#f2e93f"
                            if t.d14 == "Н": t.d14c = "#808080"
                            if t.d14 == "М": t.d14c = "#ffa812"
                            if t.d14 == "Б" or t.d14 == "О" or t.d14 == "Д" or t.d14 == "У": t.d14c = "#ffff00"
                            if t.d14 == "В": t.d14c = "#fa5555"
                        t.save()
                    if t.d15:
                        try:
                            d = int(t.d15)
                            if t.d15 > 10:
                                t.d15c = "#00b300"
                        except:
                            if t.d15 == "ВВ": t.d15c = "#f2e93f"
                            if t.d15 == "Н": t.d15c = "#808080"
                            if t.d15 == "М": t.d15c = "#ffa812"
                            if t.d15 == "Б" or t.d15 == "О" or t.d15 == "Д" or t.d15 == "У": t.d15c = "#ffff00"
                            if t.d15 == "В": t.d15c = "#fa5555"
                        t.save()
                    if t.d16:
                        try:
                            d = int(t.d16)
                            if t.d16 > 10:
                                t.d16c = "#00b300"
                        except:
                            if t.d16 == "ВВ": t.d16c = "#f2e93f"
                            if t.d16 == "Н": t.d16c = "#808080"
                            if t.d16 == "М": t.d16c = "#ffa812"
                            if t.d16 == "Б" or t.d16 == "О" or t.d16 == "Д" or t.d16 == "У": t.d16c = "#ffff00"
                            if t.d16 == "В": t.d16c = "#fa5555"
                        t.save()
                    if t.d17:
                        try:
                            d = int(t.d17)
                            if t.d17 > 10:
                                t.d17c = "#00b300"
                        except:
                            if t.d17 == "ВВ": t.d17c = "#f2e93f"
                            if t.d17 == "Н": t.d17c = "#808080"
                            if t.d17 == "М": t.d17c = "#ffa812"
                            if t.d17 == "Б" or t.d17 == "О" or t.d17 == "Д" or t.d17 == "У": t.d17c = "#ffff00"
                            if t.d17 == "В": t.d17c = "#fa5555"
                        t.save()
                    if t.d18:
                        try:
                            d = int(t.d18)
                            if t.d18 > 10:
                                t.d18c = "#00b300"
                        except:
                            if t.d18 == "ВВ": t.d18c = "#f2e93f"
                            if t.d18 == "Н": t.d18c = "#808080"
                            if t.d18 == "М": t.d18c = "#ffa812"
                            if t.d18 == "Б" or t.d18 == "О" or t.d18 == "Д" or t.d18 == "У": t.d18c = "#ffff00"
                            if t.d18 == "В": t.d18c = "#fa5555"
                        t.save()
                    if t.d19:
                        try:
                            d = int(t.d19)
                            if t.d19 > 10:
                                t.d19c = "#00b300"
                        except:
                            if t.d19 == "ВВ": t.d19c = "#f2e93f"
                            if t.d19 == "Н": t.d19c = "#808080"
                            if t.d19 == "М": t.d19c = "#ffa812"
                            if t.d19 == "Б" or t.d19 == "О" or t.d19 == "Д" or t.d19 == "У": t.d19c = "#ffff00"
                            if t.d19 == "В": t.d19c = "#fa5555"
                        t.save()
                    if t.d20:
                        try:
                            d = int(t.d20)
                            if t.d20 > 10:
                                t.d20c = "#00b300"
                        except:
                            if t.d20 == "ВВ": t.d20c = "#f2e93f"
                            if t.d20 == "Н": t.d20c = "#808080"
                            if t.d20 == "М": t.d20c = "#ffa812"
                            if t.d20 == "Б" or t.d20 == "О" or t.d20 == "Д" or t.d20 == "У": t.d20c = "#ffff00"
                            if t.d20 == "В": t.d20c = "#fa5555"
                        t.save()
                    if t.d21:
                        try:
                            d = int(t.d21)
                            if t.d21 > 10:
                                t.d21c = "#00b300"
                        except:
                            if t.d21 == "ВВ": t.d21c = "#f2e93f"
                            if t.d21 == "Н": t.d21c = "#808080"
                            if t.d21 == "М": t.d21c = "#ffa812"
                            if t.d21 == "Б" or t.d21 == "О" or t.d21 == "Д" or t.d21 == "У": t.d21c = "#ffff00"
                            if t.d21 == "В": t.d21c = "#fa5555"
                        t.save()
                    if t.d22:
                        try:
                            d = int(t.d22)
                            if t.d22 > 10:
                                t.d22c = "#00b300"
                        except:
                            if t.d22 == "ВВ": t.d22c = "#f2e93f"
                            if t.d22 == "Н": t.d22c = "#808080"
                            if t.d22 == "М": t.d22c = "#ffa812"
                            if t.d22 == "Б" or t.d22 == "О" or t.d22 == "Д" or t.d22 == "У": t.d22c = "#ffff00"
                            if t.d22 == "В": t.d22c = "#fa5555"
                        t.save()
                    if t.d23:
                        try:
                            d = int(t.d23)
                            if t.d23 > 10:
                                t.d23c = "#00b300"
                        except:
                            if t.d23 == "ВВ": t.d23c = "#f2e93f"
                            if t.d23 == "Н": t.d23c = "#808080"
                            if t.d23 == "М": t.d23c = "#ffa812"
                            if t.d23 == "Б" or t.d23 == "О" or t.d23 == "Д" or t.d23 == "У": t.d23c = "#ffff00"
                            if t.d23 == "В": t.d23c = "#fa5555"
                        t.save()
                    if t.d24:
                        try:
                            d = int(t.d24)
                            if t.d24 > 10:
                                t.d24c = "#00b300"
                        except:
                            if t.d24 == "ВВ": t.d24c = "#f2e93f"
                            if t.d24 == "Н": t.d24c = "#808080"
                            if t.d24 == "М": t.d24c = "#ffa812"
                            if t.d24 == "Б" or t.d24 == "О" or t.d24 == "Д" or t.d24 == "У": t.d24c = "#ffff00"
                            if t.d24 == "В": t.d24c = "#fa5555"
                        t.save()
                    if t.d25:
                        try:
                            d = int(t.d25)
                            if t.d25 > 10:
                                t.d25c = "#00b300"
                        except:
                            if t.d25 == "ВВ": t.d25c = "#f2e93f"
                            if t.d25 == "Н": t.d25c = "#808080"
                            if t.d25 == "М": t.d25c = "#ffa812"
                            if t.d25 == "Б" or t.d25 == "О" or t.d25 == "Д" or t.d25 == "У": t.d25c = "#ffff00"
                            if t.d25 == "В": t.d25c = "#fa5555"
                        t.save()
                    if t.d26:
                        try:
                            d = int(t.d26)
                            if t.d26 > 10:
                                t.d26c = "#00b300"
                        except:
                            if t.d26 == "ВВ": t.d26c = "#f2e93f"
                            if t.d26 == "Н": t.d26c = "#808080"
                            if t.d26 == "М": t.d26c = "#ffa812"
                            if t.d26 == "Б" or t.d26 == "О" or t.d26 == "Д" or t.d26 == "У": t.d26c = "#ffff00"
                            if t.d26 == "В": t.d26c = "#fa5555"
                        t.save()
                    if t.d27:
                        try:
                            d = int(t.d27)
                            if t.d27 > 10:
                                t.d27c = "#00b300"
                        except:
                            if t.d27 == "ВВ": t.d27c = "#f2e93f"
                            if t.d27 == "Н": t.d27c = "#808080"
                            if t.d27 == "М": t.d27c = "#ffa812"
                            if t.d27 == "Б" or t.d27 == "О" or t.d27 == "Д" or t.d27 == "У": t.d27c = "#ffff00"
                            if t.d27 == "В": t.d27c = "#fa5555"
                        t.save()
                    if t.d28:
                        try:
                            d = int(t.d28)
                            if t.d28 > 10:
                                t.d28c = "#00b300"
                        except:
                            if t.d28 == "ВВ": t.d28c = "#f2e93f"
                            if t.d28 == "Н": t.d28c = "#808080"
                            if t.d28 == "М": t.d28c = "#ffa812"
                            if t.d28 == "Б" or t.d28 == "О" or t.d28 == "Д" or t.d28 == "У": t.d28c = "#ffff00"
                            if t.d28 == "В": t.d28c = "#fa5555"
                        t.save()
                    if t.d29:
                        try:
                            d = int(t.d29)
                            if t.d29 > 10:
                                t.d29c = "#00b300"
                        except:
                            if t.d29 == "ВВ": t.d29c = "#f2e93f"
                            if t.d29 == "Н": t.d29c = "#808080"
                            if t.d29 == "М": t.d29c = "#ffa812"
                            if t.d29 == "Б" or t.d29 == "О" or t.d29 == "Д" or t.d29 == "У": t.d29c = "#ffff00"
                            if t.d29 == "В": t.d29c = "#fa5555"
                        t.save()
                    if t.d30:
                        try:
                            d = int(t.d30)
                            if t.d30 > 10:
                                t.d30c = "#00b300"
                        except:
                            if t.d30 == "ВВ": t.d30c = "#f2e93f"
                            if t.d30 == "Н": t.d30c = "#808080"
                            if t.d30 == "М": t.d30c = "#ffa812"
                            if t.d30 == "Б" or t.d30 == "О" or t.d30 == "Д" or t.d30 == "У": t.d30c = "#ffff00"
                            if t.d30 == "В": t.d30c = "#fa5555"
                        t.save()
                    if t.d31:
                        try:
                            d = int(t.d31)
                            if t.d31 > 10:
                                t.d31c = "#00b300"
                        except:
                            if t.d31 == "ВВ": t.d31c = "#f2e93f"
                            if t.d31 == "Н": t.d31c = "#808080"
                            if t.d31 == "М": t.d31c = "#ffa812"
                            if t.d31 == "Б" or t.d31 == "О" or t.d31 == "Д" or t.d31 == "У": t.d31c = "#ffff00"
                            if t.d31 == "В": t.d31c = "#fa5555"
                        t.save()
                h = 0
                d = 0
                try:
                    try:
                        h1 = int(t.d1)
                        h = h + h1
                        d+=1
                    except:
                        if t.d1 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d+=1
                    try:
                        h2 = int(t.d2)
                        h = h + h2
                        d+=1
                    except:
                        if t.d2 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h3 = int(t.d3)
                        h = h + h3
                        d+=1
                    except:
                        if t.d3 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h4 = int(t.d4)
                        h = h + h4
                        d+=1
                    except:
                        if t.d4 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h5 = int(t.d5)
                        h = h + h5
                        d+=1
                    except:
                        if t.d5 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h6 = int(t.d6)
                        h = h + h6
                        d+=1
                    except:
                        if t.d6 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h7 = int(t.d7)
                        h = h + h7
                        d+=1
                    except:
                        if t.d7 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h8 = int(t.d8)
                        h = h + h8
                        d+=1
                    except:
                        if t.d8 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h9 = int(t.d9)
                        h = h + h9
                        d+=1
                    except:
                        if t.d9 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h10 = int(t.d10)
                        h = h + h10
                        d+=1
                    except:
                        if t.d10 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h11 = int(t.d11)
                        h = h + h11
                        d+=1
                    except:
                        if t.d11 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h12 = int(t.d12)
                        h = h + h12
                        d+=1
                    except:
                        if t.d12 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h13 = int(t.d13)
                        h = h + h13
                        d+=1
                    except:
                        if t.d13 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h14 = int(t.d14)
                        h = h + h14
                        d+=1
                    except:
                        if t.d14 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h15 = int(t.d15)
                        h = h + h15
                        d+=1
                    except:
                        if t.d15 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h16 = int(t.d16)
                        h = h + h16
                        d+=1
                    except:
                        if t.d16 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h17 = int(t.d17)
                        h = h + h17
                        d+=1
                    except:
                        if t.d17 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h18 = int(t.d18)
                        h = h + h18
                        d+=1
                    except:
                        if t.d18 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h19 = int(t.d19)
                        h = h + h19
                        d+=1
                    except:
                        if t.d19 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h20 = int(t.d20)
                        h = h + h20
                        d+=1
                    except:
                        if t.d20 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h21 = int(t.d21)
                        h = h + h21
                        d+=1
                    except:
                        if t.d21 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h22 = int(t.d22)
                        h = h + h22
                        d+=1
                    except:
                        if t.d22 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h23 = int(t.d23)
                        h = h + h23
                        d+=1
                    except:
                        if t.d23 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h24 = int(t.d24)
                        h = h + h24
                        d+=1
                    except:
                        if t.d24 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h25 = int(t.d25)
                        h = h + h25
                        d+=1
                    except:
                        if t.d25 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h26 = int(t.d26)
                        h = h + h26
                        d+=1
                    except:
                        if t.d26 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h27 = int(t.d27)
                        h = h + h27
                        d+=1
                    except:
                        if t.d27 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h28 = int(t.d28)
                        h = h + h28
                        d+=1
                    except:
                        if t.d28 == "ВВ":
                            h = h+int(t.persone.work_hours)
                            d+=1
                    try:
                        h29 = int(t.d29)
                        h = h + h29
                        d += 1
                    except:
                        if t.d29 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h30 = int(t.d30)
                        h = h + h30
                        d += 1
                    except:
                        if t.d30 == "ВВ":
                            h = h + int(t.persone.work_hours)
                            d += 1
                    try:
                        h31 = int(t.d31)
                        h = h + h31
                        d += 1
                    except:
                        if t.d31 == "ВВ":
                            h = h + int(t.persone.work_hours)
                except:
                    pass
                t.hours = int(h)
                t.days = int(d)
                t.save()
                comment = request.POST.get("com"+str(t.id))
                if comment == '':
                    pass
                else:
                    if t.comment == None:
                        t.comment = ''
                        t.save()
                        comment = str(t.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                        t.comment = comment
                        t.save()
                    else:
                        comment = str(t.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                        t.comment = comment
                        t.save()
                try:
                    fine = request.POST.get(str(t.id)+"fine")
                    if fine:
                        t.fine = float(fine)
                    else:
                        t.fine = 0
                except:
                    pass
                try:
                    add = request.POST.get(str(t.id)+"add")
                    if add:
                        t.add = float(add)
                    else:
                        t.add = 0
                except:
                    pass

                f = t.fine
                a = t.add

                t.sum = t.persone.salary/int(per.workdays)*t.days
                #if per.graphic == 1:
                #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                #else:
                #    t.sum = t.persone.salary / (int(per.workdays) * int(t.persone.work_hours)) * t.hours
                #t.salary = int(t.sum) - float(f) + float(a)

                t.save()
            wd = request.POST.get('wd');
            if int(wd) != per.workdays:
                per.workdays = int(wd)
                per.save()
            return redirect('/view_tabel/' + str(id))
        else:
            context = {
                                                'user': user,
                'approvers':approvers,
                'persones': persones,
                'groups': groups,
                'per': per,
                'segment': 'tabel',
                'tabs': tabs,
                'pers': pers,
            }
            return HttpResponse(html_template.render(context, request))


    @login_required(login_url="/login/")
    def group_move_tabel(request, ids, id):
        per_old = period.objects.get(id=ids)
        tabs = tabel_list.objects.all().filter(period=per_old)
        per = period.objects.get(id=id)
        for t in tabs:
            p = persone.objects.get(id=t.persone.id)
            tabel_list(
                period=per,
                persone=p,
                company=p.company,
                snils=p.snils,
                full_name=p.full_name,
                group=p.group,
                position=p.position,
                salary=p.salary,
                output=p.output,
            ).save()
        return redirect('/view_tabel/' + str(id))

        
    @login_required(login_url="/login/")
    def approve_table(request, day, ids):
        per = period.objects.get(id=ids)
        dayapprove(
            day=day,
            per=per,
            approver=request.user,
            time=datetime.datetime.now()
        ).save()
        return redirect('/view_tabel/'+str(ids))

    @login_required(login_url="/login/")
    def approve_tableD(request, day, ids):
        per = driver.objects.get(id=ids)
        dayapproved(
            day=day,
            per=per,
            approver=request.user,
            time=datetime.datetime.now()
        ).save()
        return redirect('/view_driver/' + str(ids))

    @login_required(login_url="/login/")
    def del_tabel(request, id):
        per = period.objects.get(id=id)
        tabs = tabel_list.objects.all().filter(period=per)
        das = dayapprove.objects.all().filter(per=per)
        for d in das:
            d.delete()
        for t in tabs:
            t.delete()
        per.delete()
        return redirect('tabel')

    @login_required(login_url="/login/")
    def del_driver(request, id):
        per = driver.objects.get(id=id)
        tabs = driver_list.objects.all().filter(period=per)
        das = dayapproved.objects.all().filter(per=per)
        for d in das:
            d.delete()
        for t in tabs:
            t.delete()
        per.delete()
        return redirect('driver')


    @login_required(login_url="/login/")
    @cherrypy.expose
    def hexp_tabel(request,name,id):
        per = period.objects.get(id=id)
        tabs = tabel_list.objects.all().filter(period=per)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('templ.xlsx')
        sheet : worksheet = book.worksheets[0]
        m = int(per.date.month)-1
        month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря']
        my = str(month[m])+' '+ str(per.date.strftime("%Y"))
        sheet["E3"].value = 'Табель от '+my
        sheet["K4"].value = str(per.obj)
        i = 8
        for t in tabs:
            sheet.insert_rows(i)
            sheet["A"+str(i)].value = ""+str(t.id)
            sheet["B" + str(i)].value = "" + str(t.group)
            sheet["C" + str(i)].value = ""+str(t.persone.snils)
            if t.persone.company == 1:
                sheet["D" + str(i)].value = "ООО ПМК"
            else:
                sheet["D" + str(i)].value = "Не оформлен"
            sheet["E" + str(i)].value = ""+str(t.persone.full_name)
            sheet["F" + str(i)].value = ""+str(t.persone.position)
            if t.persone.grade:
                sheet["G" + str(i)].value = "" + str(t.persone.grade) + " разряд"
            else:
                sheet["G" + str(i)].value = ""+str(int(float(t.persone.salary)))
            if t.d1:sheet["H" + str(i)].value = ""+str(t.d1)
            if t.d2:sheet["I" + str(i)].value = ""+str(t.d2)
            if t.d3:sheet["J" + str(i)].value = ""+str(t.d3)
            if t.d4:sheet["K" + str(i)].value = ""+str(t.d4)
            if t.d5:sheet["L" + str(i)].value = ""+str(t.d5)
            if t.d6:sheet["M" + str(i)].value = ""+str(t.d6)
            if t.d7:sheet["N" + str(i)].value = ""+str(t.d7)
            if t.d8:sheet["O" + str(i)].value = ""+str(t.d8)
            if t.d9:sheet["P" + str(i)].value = ""+str(t.d9)
            if t.d10:sheet["Q" + str(i)].value = ""+str(t.d10)
            if t.d11:sheet["R" + str(i)].value = ""+str(t.d11)
            if t.d12:sheet["S" + str(i)].value = ""+str(t.d12)
            if t.d13:sheet["T" + str(i)].value = ""+str(t.d13)
            if t.d14:sheet["U" + str(i)].value = ""+str(t.d14)
            if t.d15:sheet["V" + str(i)].value = ""+str(t.d15)
            if t.d16:sheet["W" + str(i)].value = ""+str(t.d16)
            if t.d17:sheet["X" + str(i)].value = ""+str(t.d17)
            if t.d18:sheet["Y" + str(i)].value = ""+str(t.d18)
            if t.d19:sheet["Z" + str(i)].value = ""+str(t.d19)
            if t.d20:sheet["AA" + str(i)].value = ""+str(t.d20)
            if t.d21:sheet["AB" + str(i)].value = ""+str(t.d21)
            if t.d22:sheet["AC" + str(i)].value = ""+str(t.d22)
            if t.d23:sheet["AD" + str(i)].value = ""+str(t.d23)
            if t.d24:sheet["AE" + str(i)].value = ""+str(t.d24)
            if t.d25:sheet["AF" + str(i)].value = ""+str(t.d25)
            if t.d26:sheet["AG" + str(i)].value = ""+str(t.d26)
            if t.d27:sheet["AH" + str(i)].value = ""+str(t.d27)
            if t.d28:sheet["AI" + str(i)].value = ""+str(t.d28)
            if t.d29:sheet["AJ" + str(i)].value = ""+str(t.d29)
            if t.d30:sheet["AK" + str(i)].value = ""+str(t.d30)
            if t.d31:sheet["AL" + str(i)].value = ""+str(t.d31)
            sheet["AM" + str(i)].value = "" + str(t.days)
            if request.user.has_perm('reference.view_temp_file'):
                if t.sum: sheet["AN" + str(i)].value = "" + str(int(float(t.sum)))
                if t.fine:sheet["AO" + str(i)].value = "" + str(int(float(t.fine)))
                if t.add:sheet["AP" + str(i)].value = "" + str(int(float(t.add)))
                if t.salary:sheet["AQ" + str(i)].value = "" + str(int(float(t.salary)))
            else:
                sheet["AN5"].value = None
                sheet["AN6"].value = None
                sheet["AN7"].value = None
                sheet["AO5"].value = None
                sheet["AO6"].value = None
                sheet["AO7"].value = None
                sheet["AP5"].value = None
                sheet["AP6"].value = None
                sheet["AP7"].value = None
                sheet["AQ5"].value = None
                sheet["AQ6"].value = None
                sheet["AQ7"].value = None

                sheet["AN5"].fill = PatternFill(fill_type=None)
                sheet["AN6"].fill = PatternFill(fill_type=None)
                sheet["AN7"].fill = PatternFill(fill_type=None)
                sheet["AO5"].fill = PatternFill(fill_type=None)
                sheet["AO6"].fill = PatternFill(fill_type=None)
                sheet["AO7"].fill = PatternFill(fill_type=None)
                sheet["AP5"].fill = PatternFill(fill_type=None)
                sheet["AP6"].fill = PatternFill(fill_type=None)
                sheet["AP7"].fill = PatternFill(fill_type=None)
                sheet["AQ5"].fill = PatternFill(fill_type=None)
                sheet["AQ6"].fill = PatternFill(fill_type=None)
                sheet["AQ7"].fill = PatternFill(fill_type=None)

                sheet["AN5"].border = None
                sheet["AN6"].border = None
                sheet["AN7"].border = None
                sheet["AO5"].border = None
                sheet["AO6"].border = None
                sheet["AO7"].border = None
                sheet["AP5"].border = None
                sheet["AP6"].border = None
                sheet["AP7"].border = None
                sheet["AQ5"].border = None
                sheet["AQ6"].border = None
                sheet["AQ7"].border = None
            i+=1
        book.title='Табель от '+my
        book.save(response)
        return response
        return redirect('home')

    @login_required(login_url="/login/")
    @cherrypy.expose
    def hexp_driver(request,name,id):
        per = driver.objects.get(id=id)
        tabs = driver_list.objects.all().filter(period=per)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('templd.xlsx')
        sheet : worksheet = book.worksheets[0]
        m = int(per.date.month)-1
        month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря']
        my = str(month[m])+' '+ str(per.date.strftime("%Y"))
        sheet["D3"].value = 'Табель учета использования рабочего времени механизаторов от'+my
        sheet["Q4"].value = str(per.obj)
        i = 8
        for t in tabs:
            sheet.insert_rows(i)
            sheet["A" + str(i)].value = "" + str(t.persone.group)
            sheet["B" + str(i)].value = ""+str(t.persone.snils)
            if t.persone.company == 1:
                sheet["C" + str(i)].value = "ООО ПМК"
            else:
                sheet["C" + str(i)].value = "Не оформлен"
            sheet["D" + str(i)].value = ""+str(t.persone.full_name)
            sheet["E" + str(i)].value = ""+str(t.auto.name)
            sheet["F" + str(i)].value = ""+str(t.auto.number)
            sheet["G" + str(i)].value = ""+str(int(int(t.persone.salary)/300))
            if t.d1:sheet["H" + str(i)].value = ""+str(t.d1)
            if t.d2:sheet["I" + str(i)].value = ""+str(t.d2)
            if t.d3:sheet["J" + str(i)].value = ""+str(t.d3)
            if t.d4:sheet["K" + str(i)].value = ""+str(t.d4)
            if t.d5:sheet["L" + str(i)].value = ""+str(t.d5)
            if t.d6:sheet["M" + str(i)].value = ""+str(t.d6)
            if t.d7:sheet["N" + str(i)].value = ""+str(t.d7)
            if t.d8:sheet["O" + str(i)].value = ""+str(t.d8)
            if t.d9:sheet["P" + str(i)].value = ""+str(t.d9)
            if t.d10:sheet["Q" + str(i)].value = ""+str(t.d10)
            if t.d11:sheet["R" + str(i)].value = ""+str(t.d11)
            if t.d12:sheet["S" + str(i)].value = ""+str(t.d12)
            if t.d13:sheet["T" + str(i)].value = ""+str(t.d13)
            if t.d14:sheet["U" + str(i)].value = ""+str(t.d14)
            if t.d15:sheet["V" + str(i)].value = ""+str(t.d15)
            if t.d16:sheet["W" + str(i)].value = ""+str(t.d16)
            if t.d17:sheet["X" + str(i)].value = ""+str(t.d17)
            if t.d18:sheet["Y" + str(i)].value = ""+str(t.d18)
            if t.d19:sheet["Z" + str(i)].value = ""+str(t.d19)
            if t.d20:sheet["AA" + str(i)].value = ""+str(t.d20)
            if t.d21:sheet["AB" + str(i)].value = ""+str(t.d21)
            if t.d22:sheet["AC" + str(i)].value = ""+str(t.d22)
            if t.d23:sheet["AD" + str(i)].value = ""+str(t.d23)
            if t.d24:sheet["AE" + str(i)].value = ""+str(t.d24)
            if t.d25:sheet["AF" + str(i)].value = ""+str(t.d25)
            if t.d26:sheet["AG" + str(i)].value = ""+str(t.d26)
            if t.d27:sheet["AH" + str(i)].value = ""+str(t.d27)
            if t.d28:sheet["AI" + str(i)].value = ""+str(t.d28)
            if t.d29:sheet["AJ" + str(i)].value = ""+str(t.d29)
            if t.d30:sheet["AK" + str(i)].value = ""+str(t.d30)
            if t.d31:sheet["AL" + str(i)].value = ""+str(t.d31)
            sheet["AM" + str(i)].value = "" + str(t.days)
            sheet["AN" + str(i)].value = "" + str(t.hours)
            if request.user.has_perm('reference.view_temp_file'):
                if t.sum: sheet["AQ" + str(i)].value = "" + str(int(float(t.sum)))
                if t.salary:sheet["AR" + str(i)].value = "" + str(int(float(t.salary)))
            else:
                sheet["AP5"].value = None
                sheet["AP6"].value = None
                sheet["AP7"].value = None
                sheet["AQ5"].value = None
                sheet["AQ6"].value = None
                sheet["AQ7"].value = None
                sheet["AR5"].value = None
                sheet["AR6"].value = None
                sheet["AR7"].value = None


                sheet["AP5"].fill = PatternFill(fill_type=None)
                sheet["AP6"].fill = PatternFill(fill_type=None)
                sheet["AP7"].fill = PatternFill(fill_type=None)
                sheet["AQ5"].fill = PatternFill(fill_type=None)
                sheet["AQ6"].fill = PatternFill(fill_type=None)
                sheet["AQ7"].fill = PatternFill(fill_type=None)
                sheet["AR5"].fill = PatternFill(fill_type=None)
                sheet["AR6"].fill = PatternFill(fill_type=None)
                sheet["AR7"].fill = PatternFill(fill_type=None)


                sheet["AQ5"].border = None
                sheet["AQ6"].border = None
                sheet["AQ7"].border = None
                sheet["AR5"].border = None
                sheet["AR6"].border = None
                sheet["AR7"].border = None
                sheet["AP5"].border = None
                sheet["AP6"].border = None
                sheet["AP7"].border = None

            i+=1
        book.title='Табель водителей от '+my
        book.save(response)
        return response
        return redirect('view_driver/'+str(id))

    # views.py
    from django.http import JsonResponse
    from django.views.decorators.csrf import csrf_exempt

    @csrf_exempt
    def save_input_value(request):
        if request.method == 'POST' and request.is_ajax():
            input_id = request.POST.get('inputId')
            input_value = request.POST.get('value')
            # Здесь вы можете сохранить значение в базе данных или файле
            # Например:
            # YourModel.objects.create(input_id=input_id, value=input_value)
            return JsonResponse({'message': 'Значение успешно сохранено'})
        else:
            return JsonResponse({'message': 'Недопустимый запрос'}, status=400)


class reference:
    @login_required(login_url="/login/")
    def ref(request):
        context = {'segment': 'ref'}
        html_template = loader.get_template('reference/ref.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def objs(request):
        obj = objs.objects.all()
        context = {'obj': obj, 'segment': 'task'}
        html_template = loader.get_template('reference/object/objects.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_object(request):
        obj = objs.objects.all()
        forms = ObjsForm
        if request.method == 'POST':
            forms = ObjsForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.save()
                return redirect('objs')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                           'obj': obj,
                           'forms': forms,
                           'erorr': erorr
                           }
                html_template = loader.get_template('reference/object/new_object.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                       'obj': obj,
                       'forms': forms
                       }
            html_template = loader.get_template('reference/object/new_object.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_object(request, id):
        r = objs.objects.get(id=id)
        context = {'r': r, 'segment': 'ref'}
        html_template = loader.get_template('reference/object/view_object.html')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
            name = request.POST.get("name")
            if name == str(r.name):
                pass
            else:
                r.name = name
                r.save()
            status = request.POST.get('status')
            if status == str(r.status):
                pass
            else:
                r.status = status
                r.save()
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def auto(request):
        auto = automobile.objects.all()
        context = {'auto': auto, 'segment': 'ref'}
        html_template = loader.get_template('reference/auto/auto.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_auto(request):
        auto = automobile.objects.all()
        forms = AutoForm
        if request.method == 'POST':
            forms = AutoForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.save()
                return redirect('auto')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                            'auto': auto,
                            'forms': forms,
                            'erorr': erorr
                            }
                html_template = loader.get_template('reference/auto/new_auto.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                        'auto': auto,
                        'forms': forms
                        }
            html_template = loader.get_template('reference/auto/new_auto.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_auto(request, id):
        r = automobile.objects.get(id=id)
        context = {'r': r, 'segment': 'ref'}
        html_template = loader.get_template('reference/auto/view_auto.html')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
            name = request.POST.get("name")
            owner = request.POST.get("owner")
            if name != str(r.name):
                r.name = name
            if owner != str(r.owner):
                r.owner = owner
            r.save()
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def group(request):
        groups = group.objects.all()
        context = {'group': groups, 'segment': 'ref'}
        html_template = loader.get_template('reference/group/group.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_group(request):
        groups = group.objects.all()
        forms = GroupForm
        if request.method == 'POST':
            forms = GroupForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.save()
                return redirect('group')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                           'group': groups,
                           'forms': forms,
                           'erorr': erorr
                           }
                html_template = loader.get_template('reference/group/new_group.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                       'group': groups,
                       'forms': forms
                       }
            html_template = loader.get_template('reference/group/new_group.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_group(request, id):
        r = group.objects.get(id=id)
        context = {'r': r, 'segment': 'ref'}
        html_template = loader.get_template('reference/group/view_group.html')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
            name = request.POST.get("name")
            if name == str(r.name):
                pass
            else:
                r.name = name
                r.save()
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def persones(request):
        pers = persone.objects.all().order_by('full_name')
        groups = group.objects.all()
        context = {'pers': pers,'groups': groups, 'segment': 'ref'}
        html_template = loader.get_template('reference/persone/persones.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_persone(request):
        pers = persone.objects.all()
        groups = group.objects.all().order_by('name')
        forms = PersoneForm
        if request.method == 'POST':
            forms = PersoneForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                p = persone.objects.all().last()
                post.id = int(p.id)+1
                post.save()
                return redirect('persones')
            else:
                erorr = 'Некорректные данные'
                context = {'segment': 'ref',
                           'pers': pers,
                           'forms': forms,
                           'erorr': erorr,
                           'groups':groups
                           }
                html_template = loader.get_template('reference/persone/new_persone.html')
                return HttpResponse(html_template.render(context, request))
        else:
            context = {'segment': 'ref',
                       'pers': pers,
                       'forms': forms,
                       'groups': groups
                       }
            html_template = loader.get_template('reference/persone/new_persone.html')
            return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_persone(request, id):
        r = persone.objects.get(id=id)
        groups = group.objects.all().order_by('name')
        persones = persone.objects.all()
        context = {'r': r,'groups': groups, 'segment': 'ref','persones':persones}
        html_template = loader.get_template('reference/persone/view_persone.html')
        if request.method == 'POST':
            comment = request.POST.get("comm")
            if comment == '':
                pass
            else:
                if r.comment == None:
                    r.comment = ''
                    r.save()
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
                else:
                    comment = str(r.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment) + '\r\n'
                    r.comment = comment
                    r.save()
            try:
                cat = request.POST.get("category")
                if cat:
                    cat = category.objects.get(id=int(cat))
                    r.category == cat
                    r.save()
            except:
                pass
            try:
                date_a = request.POST.get("date_a")
                if date_a:
                    if date_a == r.date_accept:
                        pass
                    else:
                        r.date_leave = date_a
                        r.save()
                else:
                    if date_a == r.date_accept:
                        pass
                    else:
                        r.date_accept = None
                        r.save()
            except:
                pass
            try:
                date_l = request.POST.get("date_l")
                if date_l:
                    if date_l == r.date_leave:
                        pass
                    else:
                        r.date_leave = date_l
                        r.save()
                else:
                    if date_l == r.date_leave:
                        pass
                    else:
                        r.date_leave = None
                        r.save()
            except:
                pass
            try:
                company = request.POST.get("company")
                if company == str(r.company):
                    pass
                else:
                    r.company = company
                    r.save()
            except:
                pass
            try:
                snils = request.POST.get('snils')
                if snils == str(r.snils):
                    pass
                else:
                    r.snils = snils
                    r.save()
            except:
                pass
            try:
                inn = request.POST.get('inn')
                if snils == str(r.inn):
                    pass
                else:
                    r.inn = inn
                    r.save()
            except:
                pass
            try:
                full_name = request.POST.get('full_name')
                if full_name == str(r.full_name):
                    pass
                else:
                    r.full_name = full_name
                    r.save()
            except:
                pass
            try:
                phone_number = request.POST.get('phone_number')
                r.phone_number = phone_number
                r.save()
            except:
                pass
            try:
                position = request.POST.get('position')
                if position == str(r.position):
                    pass
                else:
                    r.position = position
                    r.save()
            except:
                pass
            try:
                workgroup = request.POST.get('workgroup')
                if workgroup == str(r.workgroup):
                    pass
                else:
                    r.workgroup = workgroup
                    r.save()
            except:
                pass
            try:
                g = request.POST.get('group')
                if r.group:
                    if int(g) == r.group.id:
                        pass
                    else:
                        g = int(g)
                        r.group = group.objects.get(id=g)
                        r.save()
                else:
                    g = int(g)
                    r.group = group.objects.get(id=g)
                    r.save()
            except:
                pass
            try:
                grade = request.POST.get('grade')
                if grade == str(r.grade):
                    pass
                else:
                    r.grade = grade
                    r.save()
            except:
                pass
            try:
                salary = request.POST.get('salary')
                if salary == str(r.salary):
                    pass
                else:
                    try:
                        salary = int(salary)
                    except:
                        salary = salary.rpartition(',')[0]
                        salary = int(salary)
                    r.salary = float(salary)
                    r.save()
            except:
                pass
            try:
                graphic = request.POST.get('graphic')
                if graphic == str(r.graphic):
                    pass
                else:
                    r.graphic = int(graphic)
                    r.save()
            except:
                pass
            try:
                output = request.POST.get('output')
                if output == str(r.output):
                    pass
                else:
                    r.output = output
                    r.save()
            except:
                pass
            try:
                driver = request.POST.get('driver')
                if driver == str(r.driver):
                    pass
                else:
                    r.driver = int(driver)
                    r.save()
            except:
                pass
            try:
                leaved = request.POST.get('leaved')
                if leaved == str(r.leaved):
                    pass
                else:
                    r.leaved = int(leaved)
                    r.save()
                work_hours = request.POST.get('work_hours')
                if work_hours == str(r.work_hours):
                    pass
                else:
                    r.work_hours = work_hours
                    r.save()
            except:
                pass
            redirect('view_persone/'+str(id))
        return HttpResponse(html_template.render(context, request))

class task:
    @login_required(login_url="/login/")
    def task_main(request):
        tasks = task_head.objects.all()
        context = {'segment': 'task', 'tasks': tasks}
        html_template = loader.get_template('task/task_main.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_task(request):
        tb = task_body.objects.all()
        obj = objs.objects.all()
        forms = TaskForm
        if request.method == 'POST':
            forms = TaskForm(request.POST)
            if forms.is_valid():
                post = forms.save(commit=False)
                post.date_create = datetime.datetime.now()
                post.creator = request.user
                n = int(request.POST.get('tbn'))
                i = 0
                post.obj = objs.objects.get(id=int(request.POST.get('obj')))
                post.save()
                if n == 0:
                    task_body(
                        head=post,
                        name=request.POST.get('tn' + str(i)),
                        number=request.POST.get('tc' + str(i)),
                        ci=request.POST.get('te' + str(i)),
                        need_date=request.POST.get('td' + str(i))
                    ).save()
                    post.body = 1
                else:
                    n = n-1
                    while i <= n:
                        task_body(
                            head=post,
                            name=request.POST.get('tn'+str(i)),
                            number=request.POST.get('tc'+str(i)),
                            ci=request.POST.get('te'+str(i)),
                            need_date=request.POST.get('td' + str(i))
                        ).save()
                        i=i+1
                        post.body = i
                post.save()
                return redirect('task_main')
            else:
                error = 'Некорректные данные'
                context = {'segment': 'task', 'forms': forms,'error': error}
                html_template = loader.get_template('task/new_task.html')
                return HttpResponse(html_template.render(context, request))
        context = {'segment': 'task', 'forms':forms, 'obj':obj}
        html_template = loader.get_template('task/new_task.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def view_task(request, id):
        th = task_head.objects.get(id=id)
        tb = task_body.objects.all().filter(head=th)
        if request.method == 'POST':

            try:
                stat = request.POST.get('status')
                if stat:
                    if th.status != stat:
                        th.status = stat
                        th.status_changer = request.user
                    th.save()
            except:
                pass
            s = th.status
            comment = request.POST.get('comm')
            if comment != '':
                if th.comment == None:
                    th.comment = ' '
                    th.save()
                comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M"))+' '+ str(request.user.last_name)+' '+ str(request.user.first_name) + ': ' + str(comment)+'\r\n'
                th.comment = comment
                th.status = s
                th.save()
            try:
                ph = 0
                i = 0
                for t in tb:
                    tch = request.POST.get('tch' + str(t.id))
                    tac = request.POST.get('tac' + str(t.id))
                    tag = request.POST.get('tagreem'+str(t.id))
                    if tag:
                        tag = int(tag)
                        if tag == 1:
                            t.agreem = 2
                            t.agreemer = None
                            comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' изменил статус одобрения запроса "' + str(t.name) + '" - "На рассмотрение"' + '\r\n'
                            th.comment = comment
                            th.save()
                        if tag == 2:
                            t.agreem = 1
                            t.agreemer = request.user
                            t.save()
                            comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' изменил статус одобрения запроса "' + str(t.name) + '" - "Одобрено"' + '\r\n'
                            th.comment = comment
                            th.save()
                    else:
                        if t.agreem == 1:
                            t.agreem = 2
                            t.agreemer = None
                            comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name)+' '+ str(request.user.first_name) + ' изменил статус одобрения запроса "' + str(t.name) + '" - "На рассмотрение"' + '\r\n'
                            th.comment = comment
                            th.save()
                    if tch:
                        t.change = int(tch)
                    else:
                        t.change = None
                    try:
                        if tac:
                            tac = int(tac)
                            if tac == 1:
                                t.accept = 2
                                comment = str(th.comment) + str(
                                    datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' изменил статус приемки "' + str(
                                    t.name) + '" - "В ожидании"' + '\r\n'
                                th.comment = comment
                                th.save()
                            if tac == 2:
                                t.accept = 1
                                t.save()
                                comment = str(th.comment) + str(
                                    datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' изменил статус приемки "' + str(
                                    t.name) + '" - "Получено"' + '\r\n'
                                th.comment = comment
                                th.save()
                        else:
                            if t.accept == 1:
                                t.accept = 2
                                comment = str(th.comment) + str(datetime.datetime.now().strftime("%d.%m.%y %H:%M")) + ' Cистема: ' + str(request.user.last_name) + ' ' + str(request.user.first_name) + ' изменил статус приемки "' + str(
                                    t.name) + '" - "В ожидании"' + '\r\n'
                                th.comment = comment
                                th.save()
                    except:
                        pass
                    try:
                        if t.procent <= 99:
                            TestDay = datetime.date.today()
                            if TestDay >= t.need_date:
                                t.expired = 1
                                th.expired = 1
                                t.save()
                    except:
                        pass
                    t.save()
                    ph = ph + t.procent
                    i +=1
                endph = 0
                for t in tb:
                    if t.procent <= 100:
                        endph=endph + t.procent
                    else:
                        endph=endph + 100
                th.procent = endph/int(th.body)
                th.save()
            except:
                pass
            return redirect('/view_task/'+str(id))
        else:
            context = {'segment': 'task', 'tb':tb, 'th': th}
            html_template = loader.get_template('task/view_task.html')
            return HttpResponse(html_template.render(context, request))



    @login_required(login_url="/login/")
    def del_task(request, id):
        th = task_head.objects.get(id=id)
        tb = task_body.objects.all().filter(head=th)
        for t in tb:
            t.delete()
        th.delete()
        response = redirect('task_main')
        return response

    @login_required(login_url="/login/")
    def close_task(request, id):
        th = task_head.objects.get(id=id)
        th.closed = 1
        th.save()
        response = redirect('task_main')
        return response

    @login_required(login_url="/login/")
    def open_task(request, id):
        th = task_head.objects.get(id=id)
        th.closed = None
        th.save()
        response = redirect('task_main')
        return response

class Agreement:
    @login_required(login_url="/login/")
    def agreement_main(request):
        agreements = agreement.objects.all()
        context = {'segment': 'agreement', 'agreements': agreements}
        html_template = loader.get_template('agreement/agreement_main.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def new_agreement(request):
        if request.method == 'POST':
            form = AgreementForm(request.POST, request.FILES)
            if form.is_valid():
                post = form.save(commit=False)
                post.date_create = datetime.datetime.now()
                post.name = request.POST.get('name')
                post.creator = request.user
                post.status = 'На проверке'
                post.save()
                agreem = agreement.objects.get(id=post.id)
                agreement_step(
                    step='Проверка',
                    agreem=agreem,
                    accepted=0
                ).save()
                accepters = accepter_list.objects.all()
                for accepter in accepters:
                    accept_user = User.objects.get(id=accepter.user.id)
                    step = agreement_step.objects.all().last()
                    agreement_accepter(accepting=step,user=accept_user).save()
                return HttpResponseRedirect(reverse('agreements'))
        else:
            forms = AgreementForm
            agreem = agreement.objects.all()
            context = {'segment': 'agreement', 'agreem': agreem,'forms': forms}
            html_template = loader.get_template('agreement/new_agreement.html')
            return HttpResponse(html_template.render(context, request))


    @login_required(login_url="/login/")
    def view_agreement(request, id):
        agreem = agreement.objects.get(id=id)
        if request.method == 'POST':
            agreem.comment = str(agreem.comment)+'\r\n'+ str(datetime.datetime.now())+' '+ str(request.user.last_name)+' '+ str(request.user.first_name) +':'+str(request.POST.get('com'))
            agreem.save()
        forms = TempFileForm
        steps = agreement_step.objects.all().filter(agreem=agreem)
        accepters = agreement_accepter.objects.all()
        context = {'segment': 'agreement', 'agreem': agreem,'steps':steps,'accepters':accepters,'forms':forms}
        html_template = loader.get_template('agreement/view_agreement.html')
        return HttpResponse(html_template.render(context, request))

    @login_required(login_url="/login/")
    def correct_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        creator = User.objects.get(id=agreem.creator.id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted += 1
        close.save()
        agreement_step(
            step = 'Корректировка',
            agreem = agreem,
            accepted = 0,
            need = 1
        ).save()
        step = agreement_step.objects.all().last()
        agreement_accepter(
            accepting=step,
            user=creator,
        ).save()
        agreem.status = 'На корректировке'
        agreem.save()
        return redirect('/view_agreement/'+str(id))

    @login_required(login_url="/login/")
    def accept_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        creator = User.objects.get(id=agreem.creator.id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted+=1
        close.save()
        if close.accepted == close.need:
            n = 0
            list = agreementer_list.objects.all()
            for a in list:
                n+=1
            agreement_step(
                step='Согласование',
                agreem=agreem,
                accepted=0,
                need = n
            ).save()
            step = agreement_step.objects.all().last()
            for a in list:
                agreement_accepter(
                    accepting=step,
                    user=creator,
                ).save()
            agreem.status = 'На согласование'
            agreem.save()
        return redirect('/view_agreement/' + str(id))

    @login_required(login_url="/login/")
    def agreemented(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        creator = User.objects.get(id=agreem.creator.id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted += 1

        close.save()
        if close.accepted == close.need:
            n = 0
            list = ender_list.objects.all()
            for a in list:
                n += 1
            agreement_step(
                step='Завершение',
                agreem=agreem,
                accepted=0,
                need=n
            ).save()
            step = agreement_step.objects.all().last()
            for a in list:
                agreement_accepter(
                    accepting=step,
                    user=creator,
                ).save()
            agreem.status = 'Завершен'
            agreem.save()
        return redirect('/view_agreement/' + str(id))

    @login_required(login_url="/login/")
    def close_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        accept = agreement_accepter.objects.get(id=ida)
        accept.accept = 1
        accept.save()
        close = agreement_step.objects.get(id=ids)
        close.accepted += 1
        close.save()
        agreem.closed = 1
        agreem.closer = request.user
        agreem.save()
        agreem.status = 'Завершен'
        agreem.save()
        return redirect('/view_agreement/' + str(id))



    @login_required(login_url="/login/")
    def new_file_agreement(request, id, ids, ida):
        agreem = agreement.objects.get(id=id)
        if request.method == 'POST':
            form = TempFileForm(request.POST, request.FILES)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                agreem.file = post.temp
                agreem.status = 'Проверка'
                agreem.save()
                accept = agreement_accepter.objects.get(id=ida)
                accept.accept = 1
                accept.save()
                close = agreement_step.objects.get(id=ids)
                close.accepted += 1
                close.save()
                n=0
                accepters = accepter_list.objects.all()
                for accepter in accepters:
                    n+=1
                agreement_step(
                    step='Проверка',
                    agreem=agreem,
                    accepted=0,
                    need=n
                ).save()
                for accepter in accepters:
                    accept_user = User.objects.get(id=accepter.user.id)
                    step = agreement_step.objects.all().last()
                    agreement_accepter(accepting=step, user=accept_user).save()
                return redirect('/view_agreement/' + str(id))
        form = TempFileForm
        context = {'segment': 'agreement', 'form': form}
        html_template = loader.get_template('agreement/corrector_agreement.html')
        return HttpResponse(html_template.render(context, request))





 #   close_agreement
    @login_required(login_url="/login/")
    def agreement_file(request, id):
        document = get_object_or_404(agreement, id=id)
        response = HttpResponse(document.file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{document.file.name}"'
        return response

class misc:
    @login_required(login_url="/login/")
    def month_info(request, years, month, name):
        date = datetime.date(years,month,1)
        pers = period.objects.all().filter(date=date)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('mix.xlsx')
        t_list: worksheet = book.worksheets[0]
        i=8
        for p in pers:
            tls = tabel_list.objects.all().filter(period=p)
            m = int(p.date.month)-1
            month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября',
                     'Ноября', 'Декабря']
            my = str(month[m]) + ' ' + str(p.date.strftime("%Y"))
            for t in tls:
                t_list.insert_rows(i)
                t_list["A" + str(i)].value = "" + str(t.company)
                t_list["B" + str(i)].value = "" + str(t.full_name)
                t_list["C" + str(i)].value = "" + str(t.position)
                t_list["D" + str(i)].value = "" + str(t.period.obj)
                if t.d1:t_list["E" + str(i)].value = "" + str(t.d1)
                if t.d2:t_list["F" + str(i)].value = "" + str(t.d2)
                if t.d3:t_list["G" + str(i)].value = "" + str(t.d3)
                if t.d4:t_list["H" + str(i)].value = "" + str(t.d4)
                if t.d5:t_list["I" + str(i)].value = "" + str(t.d5)
                if t.d6:t_list["J" + str(i)].value = "" + str(t.d6)
                if t.d7:t_list["K" + str(i)].value = "" + str(t.d7)
                if t.d8:t_list["L" + str(i)].value = "" + str(t.d8)
                if t.d9:t_list["M" + str(i)].value = "" + str(t.d9)
                if t.d10:t_list["N" + str(i)].value = "" + str(t.d10)
                if t.d11:t_list["O" + str(i)].value = "" + str(t.d11)
                if t.d12:t_list["P" + str(i)].value = "" + str(t.d12)
                if t.d13:t_list["Q" + str(i)].value = "" + str(t.d13)
                if t.d14:t_list["R" + str(i)].value = "" + str(t.d14)
                if t.d15:t_list["S" + str(i)].value = "" + str(t.d15)
                if t.d16:t_list["T" + str(i)].value = "" + str(t.d16)
                if t.d17:t_list["U" + str(i)].value = "" + str(t.d17)
                if t.d18:t_list["V" + str(i)].value = "" + str(t.d18)
                if t.d19:t_list["W" + str(i)].value = "" + str(t.d19)
                if t.d20:t_list["X" + str(i)].value = "" + str(t.d20)
                if t.d21:t_list["Y" + str(i)].value = "" + str(t.d21)
                if t.d22:t_list["Z" + str(i)].value = "" + str(t.d22)
                if t.d23:t_list["AA" + str(i)].value = "" + str(t.d23)
                if t.d24:t_list["AB" + str(i)].value = "" + str(t.d24)
                if t.d25:t_list["AC" + str(i)].value = "" + str(t.d25)
                if t.d26:t_list["AD" + str(i)].value = "" + str(t.d26)
                if t.d27:t_list["AE" + str(i)].value = "" + str(t.d27)
                if t.d28:t_list["AF" + str(i)].value = "" + str(t.d28)
                if t.d29:t_list["AG" + str(i)].value = "" + str(t.d29)
                if t.d30:t_list["AH" + str(i)].value = "" + str(t.d30)
                if t.d31:t_list["AI" + str(i)].value = "" + str(t.d31)
                t_list["AJ" + str(i)].value = "" + str(t.days)
                i += 1
        book.title = 'Отчет общий'
        book.save(response)
        return response
        return redirect('/tabels/')

    @login_required(login_url="/login/")
    def month_info_driver(request, years, month, name):
        date = datetime.date(years, month, 1)
        pers = driver.objects.all().filter(date=date)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;  filename=%s' % str(name)
        book = openpyxl.load_workbook('mix.xlsx')
        t_list: worksheet = book.worksheets[0]
        i = 8
        for p in pers:
            tls = tabel_list.objects.all().filter(period=p)
            m = int(p.date.month) - 1
            month = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября',
                     'Ноября', 'Декабря']
            my = str(month[m]) + ' ' + str(p.date.strftime("%Y"))
            for t in tls:
                t_list.insert_rows(i)
                t_list["A" + str(i)].value = "" + str(t.company)
                t_list["B" + str(i)].value = "" + str(t.full_name)
                t_list["C" + str(i)].value = "" + str(t.position)
                t_list["D" + str(i)].value = "" + str(t.period.obj)
                if t.d1: t_list["E" + str(i)].value = "" + str(t.d1)
                if t.d2: t_list["F" + str(i)].value = "" + str(t.d2)
                if t.d3: t_list["G" + str(i)].value = "" + str(t.d3)
                if t.d4: t_list["H" + str(i)].value = "" + str(t.d4)
                if t.d5: t_list["I" + str(i)].value = "" + str(t.d5)
                if t.d6: t_list["J" + str(i)].value = "" + str(t.d6)
                if t.d7: t_list["K" + str(i)].value = "" + str(t.d7)
                if t.d8: t_list["L" + str(i)].value = "" + str(t.d8)
                if t.d9: t_list["M" + str(i)].value = "" + str(t.d9)
                if t.d10: t_list["N" + str(i)].value = "" + str(t.d10)
                if t.d11: t_list["O" + str(i)].value = "" + str(t.d11)
                if t.d12: t_list["P" + str(i)].value = "" + str(t.d12)
                if t.d13: t_list["Q" + str(i)].value = "" + str(t.d13)
                if t.d14: t_list["R" + str(i)].value = "" + str(t.d14)
                if t.d15: t_list["S" + str(i)].value = "" + str(t.d15)
                if t.d16: t_list["T" + str(i)].value = "" + str(t.d16)
                if t.d17: t_list["U" + str(i)].value = "" + str(t.d17)
                if t.d18: t_list["V" + str(i)].value = "" + str(t.d18)
                if t.d19: t_list["W" + str(i)].value = "" + str(t.d19)
                if t.d20: t_list["X" + str(i)].value = "" + str(t.d20)
                if t.d21: t_list["Y" + str(i)].value = "" + str(t.d21)
                if t.d22: t_list["Z" + str(i)].value = "" + str(t.d22)
                if t.d23: t_list["AA" + str(i)].value = "" + str(t.d23)
                if t.d24: t_list["AB" + str(i)].value = "" + str(t.d24)
                if t.d25: t_list["AC" + str(i)].value = "" + str(t.d25)
                if t.d26: t_list["AD" + str(i)].value = "" + str(t.d26)
                if t.d27: t_list["AE" + str(i)].value = "" + str(t.d27)
                if t.d28: t_list["AF" + str(i)].value = "" + str(t.d28)
                if t.d29: t_list["AG" + str(i)].value = "" + str(t.d29)
                if t.d30: t_list["AH" + str(i)].value = "" + str(t.d30)
                if t.d31: t_list["AI" + str(i)].value = "" + str(t.d31)
                t_list["AJ" + str(i)].value = "" + str(t.days)
                i += 1
        book.title = 'Отчет общий'
        book.save(response)
        return response
        return redirect('/tabels/')


    @login_required(login_url="/login/")
    def convert_tabel(request):
        b = 0
        while b<13:
            if b==0:
                book = openpyxl.load_workbook('tabel (1).xlsx')
            if b==1:
                book = openpyxl.load_workbook('tabel (2).xlsx')
            if b==2:
                book = openpyxl.load_workbook('tabel (3).xlsx')
            if b==3:
                book = openpyxl.load_workbook('tabel (4).xlsx')
            if b==4:
                book = openpyxl.load_workbook('tabel (5).xlsx')
            if b==5:
                book = openpyxl.load_workbook('tabel (6).xlsx')
            if b==6:
                book = openpyxl.load_workbook('tabel (7).xlsx')
            if b==7:
                book = openpyxl.load_workbook('tabel (8).xlsx')
            if b==8:
                book = openpyxl.load_workbook('tabel (9).xlsx')
            if b==9:
                book = openpyxl.load_workbook('tabel (10).xlsx')
            if b==10:
                book = openpyxl.load_workbook('tabel (11).xlsx')
            if b==11:
                book = openpyxl.load_workbook('tabel.xlsx')
            t_list: worksheet = book.worksheets[0]
            i = 8
            while i<200:
                Check = str(t_list["A"+str(i)].value)
                if Check:
                    i+=1
            i+=1
            x = 8
            years = "2024"
            month = "2"
            d = datetime.date(int(years),int(month),1)
            try:
                o = objs.objects.get(name=""+str(t_list["T4"].value))
            except:
                objs(name=str(t_list["T4"].value),status="Открыт").save()
                o = objs.objects.get(name="" + str(t_list["T4"].value))
            post = period(type=3,date=d,graphic=5,obj=o,workdays=21)
            da = d
            if post:
                try:
                    da = da.replace(day=1)
                    post.dw1 = da
                except:
                    pass
                try:
                    da = da.replace(day=2)
                    post.dw2 = da
                except:
                    pass
                try:
                    da = da.replace(day=3)
                    post.dw3 = da
                except:
                    pass
                try:
                    da = da.replace(day=4)
                    post.dw4 = da
                except:
                    pass
                try:
                    da = da.replace(day=5)
                    post.dw5 = da
                except:
                    pass
                try:
                    da = da.replace(day=6)
                    post.dw6 = da
                except:
                    pass
                try:
                    da = da.replace(day=7)
                    post.dw7 = da
                except:
                    pass
                try:
                    da = da.replace(day=8)
                    post.dw8 = da
                except:
                    pass
                try:
                    da = da.replace(day=9)
                    post.dw9 = da
                except:
                    pass
                try:
                    da = da.replace(day=10)
                    post.dw10 = da
                except:
                    pass
                try:
                    da = da.replace(day=11)
                    post.dw11 = da
                except:
                    pass
                try:
                    da = da.replace(day=12)
                    post.dw12 = da
                except:
                    pass
                try:
                    da = da.replace(day=13)
                    post.dw13 = da
                except:
                    pass
                try:
                    da = da.replace(day=14)
                    post.dw14 = da
                except:
                    pass
                try:
                    da = da.replace(day=15)
                    post.dw15 = da
                except:
                    pass
                try:
                    da = da.replace(day=16)
                    post.dw16 = da
                except:
                    pass
                try:
                    da = da.replace(day=17)
                    post.dw17 = da
                except:
                    pass
                try:
                    da = da.replace(day=18)
                    post.dw18 = da
                except:
                    pass
                try:
                    da = da.replace(day=19)
                    post.dw19 = da
                except:
                    pass
                try:
                    da = da.replace(day=20)
                    post.dw20 = da
                except:
                    pass
                try:
                    da = da.replace(day=21)
                    post.dw21 = da
                except:
                    pass
                try:
                    da = da.replace(day=22)
                    post.dw22 = da
                except:
                    pass
                try:
                    da = da.replace(day=23)
                    post.dw23 = da
                except:
                    pass
                try:
                    da = da.replace(day=24)
                    post.dw24 = da
                except:
                    pass
                try:
                    da = da.replace(day=25)
                    post.dw25 = da
                except:
                    pass
                try:
                    da = da.replace(day=26)
                    post.dw26 = da
                except:
                    pass
                try:
                    da = da.replace(day=27)
                    post.dw27 = da
                except:
                    pass
                try:
                    da = da.replace(day=28)
                    post.dw28 = da
                except:
                    pass
                try:
                    da = da.replace(day=29)
                    post.dw29 = da
                except:
                    pass
                try:
                    da = da.replace(day=30)
                    post.dw30 = da
                except:
                    pass
                try:
                    da = da.replace(day=31)
                    post.dw31 = da
                except:
                    pass
                post.save()
            while x<i:
                try:
                    int(t_list["A"+str(x)].value)
                    org = str(t_list["A"+str(x)].value)
                    if org == "ПМК":
                        org = "ООО 'ПМК'"
                    else:
                        org = "Неоформленный"
                    name = str(t_list["G"+str(x)].value)
                    p = persone.objects.get(full_name=str(name),date_leave=None)
                    tabel_list(
                        period=post,
                        persone=p,
                        company=str(org),
                        snils=str(t_list["D"+str(x)].value),
                        full_name=str(t_list["G" + str(x)].value),
                        group=g,
                        position=str(t_list["H" + str(x)].value),
                        d1=t_list["K" + str(x)].value,
                        d2=t_list["L" + str(x)].value,
                        d3=t_list["M" + str(x)].value,
                        d4=t_list["N" + str(x)].value,
                        d5=t_list["O" + str(x)].value,
                        d6=t_list["P" + str(x)].value,
                        d7=t_list["Q" + str(x)].value,
                        d8=t_list["R" + str(x)].value,
                        d9=t_list["S" + str(x)].value,
                        d10=t_list["T" + str(x)].value,
                        d11=t_list["U" + str(x)].value,
                        d12=t_list["V" + str(x)].value,
                        d13=t_list["W" + str(x)].value,
                        d14=t_list["X" + str(x)].value,
                        d15=t_list["Y" + str(x)].value,
                        d16=t_list["Z" + str(x)].value,
                        d17=t_list["AA" + str(x)].value,
                        d18=t_list["AB" + str(x)].value,
                        d19=t_list["AC" + str(x)].value,
                        d20=t_list["AD" + str(x)].value,
                        d21=t_list["AE" + str(x)].value,
                        d22=t_list["AF" + str(x)].value,
                        d23=t_list["AG" + str(x)].value,
                        d24=t_list["AH" + str(x)].value,
                        d25=t_list["AI" + str(x)].value,
                        d26=t_list["AJ" + str(x)].value,
                        d27=t_list["AK" + str(x)].value,
                        d28=t_list["AL" + str(x)].value,
                        d29=t_list["AM" + str(x)].value,
                        #d30=t_list["AN" + str(x)].value,
                    ).save()
                    x += 1
                except:
                    try:
                        gn = str(t_list["A"+str(x)].value)
                        g = group.objects.get(name=str(gn))
                        x += 1
                    except:
                        n = len(str(t_list["A"+str(x)].value))
                        if n>2:
                            g = group.objects.all().last()
                            id = int(g.id)+1
                            group(id=id,name=str(t_list["A"+str(x)].value)).save()
                            g = group.objects.get(name=str(t_list["A" + str(x)].value))
                        x += 1
            b+=1
        return redirect('/profile.html')

    @login_required(login_url="/login/")
    def convert_driver(request):
        b = 0
        while b < 3:
            if b == 0:
                book = openpyxl.load_workbook('Driver.xlsx')
            if b == 1:
                book = openpyxl.load_workbook('Driver (1).xlsx')
            if b == 2:
                book = openpyxl.load_workbook('Driver (2).xlsx')
            t_list: worksheet = book.worksheets[0]
            i = 18
            x = 9
            years = "2024"
            month = "1"
            d = datetime.date(int(years), int(month), 1)
            n = str(t_list["T4"].value)
            n = n[:-4]
            try:
                o = objs.objects.get(name="" + n)
            except:
                objs(name=str(n), status="Открыт").save()
                o = objs.objects.get(name="" + str(n))
            post = driver(date=d, obj=o)
            da = d
            if post:
                try:
                    da = da.replace(day=1)
                    post.dw1 = da
                except:
                    pass
                try:
                    da = da.replace(day=2)
                    post.dw2 = da
                except:
                    pass
                try:
                    da = da.replace(day=3)
                    post.dw3 = da
                except:
                    pass
                try:
                    da = da.replace(day=4)
                    post.dw4 = da
                except:
                    pass
                try:
                    da = da.replace(day=5)
                    post.dw5 = da
                except:
                    pass
                try:
                    da = da.replace(day=6)
                    post.dw6 = da
                except:
                    pass
                try:
                    da = da.replace(day=7)
                    post.dw7 = da
                except:
                    pass
                try:
                    da = da.replace(day=8)
                    post.dw8 = da
                except:
                    pass
                try:
                    da = da.replace(day=9)
                    post.dw9 = da
                except:
                    pass
                try:
                    da = da.replace(day=10)
                    post.dw10 = da
                except:
                    pass
                try:
                    da = da.replace(day=11)
                    post.dw11 = da
                except:
                    pass
                try:
                    da = da.replace(day=12)
                    post.dw12 = da
                except:
                    pass
                try:
                    da = da.replace(day=13)
                    post.dw13 = da
                except:
                    pass
                try:
                    da = da.replace(day=14)
                    post.dw14 = da
                except:
                    pass
                try:
                    da = da.replace(day=15)
                    post.dw15 = da
                except:
                    pass
                try:
                    da = da.replace(day=16)
                    post.dw16 = da
                except:
                    pass
                try:
                    da = da.replace(day=17)
                    post.dw17 = da
                except:
                    pass
                try:
                    da = da.replace(day=18)
                    post.dw18 = da
                except:
                    pass
                try:
                    da = da.replace(day=19)
                    post.dw19 = da
                except:
                    pass
                try:
                    da = da.replace(day=20)
                    post.dw20 = da
                except:
                    pass
                try:
                    da = da.replace(day=21)
                    post.dw21 = da
                except:
                    pass
                try:
                    da = da.replace(day=22)
                    post.dw22 = da
                except:
                    pass
                try:
                    da = da.replace(day=23)
                    post.dw23 = da
                except:
                    pass
                try:
                    da = da.replace(day=24)
                    post.dw24 = da
                except:
                    pass
                try:
                    da = da.replace(day=25)
                    post.dw25 = da
                except:
                    pass
                try:
                    da = da.replace(day=26)
                    post.dw26 = da
                except:
                    pass
                try:
                    da = da.replace(day=27)
                    post.dw27 = da
                except:
                    pass
                try:
                    da = da.replace(day=28)
                    post.dw28 = da
                except:
                    pass
                try:
                    da = da.replace(day=29)
                    post.dw29 = da
                except:
                    pass
                try:
                    da = da.replace(day=30)
                    post.dw30 = da
                except:
                    pass
                try:
                    da = da.replace(day=31)
                    post.dw31 = da
                except:
                    pass
                post.save()
            while x < i:
                    org = str(t_list["E" + str(x)].value)
                    if org == "ПМК":
                        org = 1
                    else:
                        org = 2
                    name = str(t_list["G" + str(x)].value)
                    try:
                        p = persone.objects.get(full_name=str(name), date_leave=None)
                        p.driver = 1
                        p.save()
                    except:
                        try:
                            p = persone.objects.get(full_name=str(name)+"(*)", date_leave=None)
                            p.driver = 1
                            p.save()
                        except:
                            if str(t_list["C" + str(i)].value):
                                try:
                                    cat = category.objects.get(name=str(t_list["C" + str(i)].value))
                                except:
                                    category(name=str(t_list["C" + str(i)].value)).save()
                                    cat = category.objects.get(name=str(t_list["C" + str(i)].value))
                            gr=group.objects.get(name="ОГМ")
                            persone(
                                company=org,
                                snils=str(t_list["D" + str(x)].value),
                                full_name=str(t_list["G" + str(x)].value)+"(*)",
                                category=cat,
                                group=gr,
                                date_accept=datetime.date.today(),
                                driver=1,
                            ).save()
                            p = persone.objects.get(full_name=str(name)+"(*)")
                    try:
                        try:
                          auto = automobile.objects.get(name=str(t_list["H" + str(x)].value))
                        except:
                          auto = automobile.objects.get(number=str(t_list["I" + str(x)].value))
                    except:
                        automobile(name=str(t_list["H" + str(x)].value), owner='ООО "ПМК"', number=str(t_list["I" + str(x)].value)).save()
                        auto = automobile.objects.get(name=str(t_list["H" + str(x)].value),)
                    driver_list(
                        period=post,
                        persone=p,
                        company=str(org),
                        full_name=str(t_list["G" + str(x)].value),
                        position=p.position,
                        auto=auto,
                        d1=t_list["K" + str(x)].value,
                        d2=t_list["L" + str(x)].value,
                        d3=t_list["M" + str(x)].value,
                        d4=t_list["N" + str(x)].value,
                        d5=t_list["O" + str(x)].value,
                        d6=t_list["P" + str(x)].value,
                        d7=t_list["Q" + str(x)].value,
                        d8=t_list["R" + str(x)].value,
                        d9=t_list["S" + str(x)].value,
                        d10=t_list["T" + str(x)].value,
                        d11=t_list["U" + str(x)].value,
                        d12=t_list["V" + str(x)].value,
                        d13=t_list["W" + str(x)].value,
                        d14=t_list["X" + str(x)].value,
                        d15=t_list["Y" + str(x)].value,
                        d16=t_list["Z" + str(x)].value,
                        d17=t_list["AA" + str(x)].value,
                        d18=t_list["AB" + str(x)].value,
                        d19=t_list["AC" + str(x)].value,
                        d20=t_list["AD" + str(x)].value,
                        d21=t_list["AE" + str(x)].value,
                        d22=t_list["AF" + str(x)].value,
                        d23=t_list["AG" + str(x)].value,
                        d24=t_list["AH" + str(x)].value,
                        d25=t_list["AI" + str(x)].value,
                        d26=t_list["AJ" + str(x)].value,
                        d27=t_list["AK" + str(x)].value,
                        d28=t_list["AL" + str(x)].value,
                        d29=t_list["AM" + str(x)].value,
                        #d30=t_list["AN" + str(x)].value,
                    ).save()
                    x += 1
            b += 1
        return redirect('/profile.html')
    def convert3(request):
        book = openpyxl.load_workbook('workers.xlsx')
        workers = book.worksheets[0]

        # Получаем связанные объекты за пределами цикла
        possible_groups = group.objects.all()  # Предполагая, что вам нужны все группы
        for row in workers.iter_rows(min_row=2, max_row=699):
            if row[0].value:  # Проверяем пустую ячейку в столбце A
                persones = persone.objects.all()
                for p in persones:
                    if p.full_name == str(row[3].value):
                        if p.salary == 0:
                            grades = str(row[11].value),
                            if int(grades[0]) == 2:
                                sal = 60900
                                p.grade = 2
                            elif int(grades[0]) == 3:
                                sal = 67800
                                p.grade = 2
                            elif int(grades[0]) == 4:
                                sal = 72700
                                p.grade = 2
                            elif int(grades[0]) == 5:
                                sal = 80000
                                p.grade = 2
                            elif int(grades[0]) == 6:
                                sal = 86250
                                p.grade = 2
                            p.salary = sal
                            p.save()
        return redirect('/profile.html')


    @login_required(login_url="/login/")
    def convert2(request):
        book = openpyxl.load_workbook('workers.xlsx')
        workers = book.worksheets[0]

        # Получаем связанные объекты за пределами цикла
        possible_groups = group.objects.all()  # Предполагая, что вам нужны все группы

        for row in workers.iter_rows(min_row=2, max_row=699):
            if row[0].value:  # Проверяем пустую ячейку в столбце A
                org = 1 if row[1].value == 'ПМК' else 2
                cat, created = category.objects.get_or_create(
                    name__iexact=row[4].value)  # Получаем или создаем категорию
                gr = possible_groups.filter(name=row[5].value).first()  # Получаем группу, если она существует
                sal = int(str(row[10].value)) or int(str(row[9].value)) or int(str(row[8].value))  # Получаем зарплату
                grades = str(row[11].value),
                if sal:
                    pass
                else:
                    if int(grades[0]) == 2:
                        sal = 60900
                    elif int(grades[0]) == 3:
                        sal = 67800
                    elif int(grades[0]) == 4:
                        sal = 72700
                    elif int(grades[0]) == 5:
                        sal = 80000
                    elif int(grades[0]) == 6:
                        sal = 86250
                # Проверяем существование объекта перед созданием
                if not persone.objects.filter(full_name=str(row[3].value),date_accept=dateutil.parser.parse(str(row[13].value)).strftime("%Y-%m-%d") if row[13].value else None,).exists():
                    persone.objects.create(
                        company=org,
                        snils=str(row[2].value),
                        full_name=str(row[3].value),
                        category=cat,
                        group=gr,
                        position=str(row[7].value),
                        salary=float(sal),
                        graphic=str(row[12].value),
                        grade=grades[0],
                        date_accept=dateutil.parser.parse(str(row[13].value)).strftime("%Y-%m-%d") if row[13].value else None,
                        date_leave=dateutil.parser.parse(str(row[14].value)).strftime("%Y-%m-%d") if row[14].value else None,
                        leaved=1 if row[14].value else 0
                    )



        return redirect('/profile.html')

    @login_required(login_url="/login/")
    def convert(request):
        book = openpyxl.load_workbook('workers.xlsx')
        workers: worksheet = book.worksheets[0]
        i = 2
        while i<699:
            if workers["A" + str(i)].value:
                if str(workers["F"+str(i)].value):
                    gr = group.objects.get(name=str(workers["F"+str(i)].value))
                else:
                    gr = None
                if str(workers["B"+str(i)].value) == 'ПМК':
                    org = 1
                else:
                    org = 2
                if str(workers["E"+str(i)].value):
                    try:
                        cat = category.objects.get(name=str(workers["E"+str(i)].value))
                    except:
                        category(name=str(workers["E"+str(i)].value)).save()
                        cat = category.objects.get(name=str(workers["E"+str(i)].value))
                da = str(workers["N" + str(i)].value)
                if da:
                    try:
                        da = datetime.datetime.strptime(da, "%d.%m.%Y").strftime("%Y-%m-%d")
                    except:
                        da = None
                else:
                    da = None
                dl = str(workers["O" + str(i)].value)
                if dl:
                    try:
                        dl = datetime.datetime.strptime(dl, "%d.%m.%Y").strftime("%Y-%m-%d")
                    except:
                        dl = None
                else:
                    dl = None
                sal = int(str(workers["K" + str(i)].value))
                if sal == 0:
                    sal = int(str(workers["J" + str(i)].value))
                    if sal == 0:
                        sal = int(str(workers["I" + str(i)].value))
                persone(
                    company=org,
                    snils=str(workers["C" + str(i)].value),
                    full_name=str(workers["D" + str(i)].value),
                    category=cat,
                    group=gr,
                    position=str(workers["H" + str(i)].value),
                    salary=float(int(sal)),
                    graphic=str(workers["M" + str(i)].value),
                    date_accept=da,
                    date_leave=dl
                ).save()
            i+=1
        return redirect('/profile.html')

    @login_required(login_url="/login/")
    def auto_number(request):
        autos = automobile.objects.all()
        for auto in autos:
            try:
                auto.number = str(auto.comment)
                auto.save()
            except:
                pass
        return redirect('/profile.html')



@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))
